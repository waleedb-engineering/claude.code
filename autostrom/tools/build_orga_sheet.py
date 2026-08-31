# -*- coding: utf-8 -*-
"""Legt das Organigramm auf einem eigenen Reiter "Organigramm" an.

Der Reiter enthaelt drei Dinge:
  * die Zeichenflaeche, auf der inject_shapes.py die Boxen absetzt
  * die Textquellen (Spalte Z), mit denen die Boxen verknuepft sind
  * darunter dieselbe Struktur als Zellen, die sich automatisch einfaerbt

Alle Werte stammen aus der Komponententabelle; hier wird nichts gepflegt.

Aufruf:  python3 build_orga_sheet.py <mappe.xlsx>
"""
import json, sys
import openpyxl
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle

K_START, K_ENDE = 6, 29
BOX_B, ABSTAND = 200, 16
RASTER = BOX_B + ABSTAND
EBENEN = dict(wurzel=(0, 52), ve=(86, 48), partner=(166, 48), art=(246, 32))
Y_KOMP, H_KOMP, RASTER_KOMP = 302, 70, 80

C_TITEL, C_VE, C_PARTNER, C_ART = 'FF1F3864', 'FF1F4E78', 'FF2E75B6', 'FFBDD7EE'
STATUSFARBEN = [('erfasst', 'FFE7E6E6', 'FF404040', None),
                ('bewertet', 'FFD9EAF7', 'FF1F3864', None),
                ('erfüllt', 'FFE2F0D9', 'FF006100', True),
                ('trifft nicht zu', 'FFD9D9D9', 'FF666666', None),
                ('außerhalb Umfang', 'FFEDEDED', 'FF9C6500', None),
                ('keine Anforderungen', 'FFF2F2F2', 'FF777777', None)]
RAHMEN = Border(*[Side(style='thin', color='FFBFBFBF')] * 4)


def regel(formel, fuellung, farbe=None, fett=None):
    stil = DifferentialStyle(font=Font(color=farbe, b=fett) if (farbe or fett) else None,
                             fill=PatternFill(bgColor=fuellung))
    return Rule(type='expression', formula=[formel], dxf=stil, stopIfTrue=False)


def hierarchie(ws):
    """Gruppiert die Komponententabelle logisch nach Vergabeeinheit, Partner und Art."""
    gruppen, reihenfolge = {}, []
    for r in range(K_START, K_ENDE + 1):
        if not ws.cell(r, 4).value:
            continue
        schluessel = (ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value)
        if schluessel not in gruppen:
            gruppen[schluessel], _ = [], reihenfolge.append(schluessel)
        gruppen[schluessel].append(r)
    ve_rang, partner_rang = {}, {}
    for ve, partner, _ in reihenfolge:
        ve_rang.setdefault(ve, len(ve_rang))
        partner_rang.setdefault((ve, partner), len(partner_rang))
    reihenfolge.sort(key=lambda s: (ve_rang[s[0]], partner_rang[s[:2]]))
    return gruppen, reihenfolge


def main():
    pfad = sys.argv[1]
    wb = openpyxl.load_workbook(pfad)
    komp = wb['Komponenten']
    gruppen, reihenfolge = hierarchie(komp)
    spalte_von = {s: i for i, s in enumerate(reihenfolge)}
    max_komp = max(len(z) for z in gruppen.values())

    # --- Komponentenreiter aufräumen: alles unterhalb der Tabelle entfernt ---------
    for rng in [str(x) for x in komp.merged_cells.ranges]:
        if int(rng.split(':')[0][1:] or 0) >= 31 or ':' in rng and int(
                ''.join(c for c in rng.split(':')[0] if c.isdigit()) or 0) >= 31:
            komp.unmerge_cells(rng)
    for r in range(31, komp.max_row + 2):
        for c in list(range(1, 14)):          # Spalte N (Partnerliste) bleibt unberührt
            komp.cell(r, c).value = None
    komp.cell(31, 1).value = ('Das Organigramm liegt im eigenen Reiter „Organigramm“. '
                              'Es rechnet vollständig aus dieser Tabelle.')
    komp.cell(31, 1).font = Font(name='Arial', size=10, italic=True, color='FF808080')
    komp.sheet_view.topLeftCell = 'A1'

    # --- Reiter anlegen -----------------------------------------------------------
    if 'Organigramm' in wb.sheetnames:
        del wb['Organigramm']
    ws = wb.create_sheet('Organigramm', wb.sheetnames.index('Komponenten') + 1)
    ws.sheet_view.showGridLines = False
    for spalte, breite in zip('ABCDE', (22, 20, 14, 30, 13)):
        ws.column_dimensions[spalte].width = breite
    ws.column_dimensions['Z'].width = 46

    ws['A1'] = 'Organigramm – Vergabe- und Komponentenstruktur'
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color=C_TITEL)
    ws['A2'] = ('SmartInfra → Vergabeeinheit → Partner → Art → Komponente. Jede Box ist mit einer '
                'Formelzelle in Spalte Z verknüpft und aktualisiert sich automatisch aus dem '
                'Reiter „Komponenten“. Gepflegt wird ausschließlich dort.')
    ws['A2'].font = Font(name='Arial', size=9, color='FF808080')
    ws.merge_cells('A2:K2')
    ws['A2'].alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 8

    dia_hoehe = Y_KOMP + max_komp * RASTER_KOMP + 12
    start = 4
    zeilen = -(-dia_hoehe // 24) + 1
    for r in range(start, start + zeilen):
        ws.row_dimensions[r].height = 18

    # --- Textquellen (Spalte Z) ---------------------------------------------------
    knoten, z_zeile = [], start
    ws.cell(start - 1, 26).value = 'Textquellen der Boxen (Formeln – bitte nicht löschen)'
    ws.cell(start - 1, 26).font = Font(name='Arial', size=9, bold=True, color='FF808080')

    def quelle(formel, ebene, **rest):
        nonlocal z_zeile
        zelle = f'Z{z_zeile}'
        ws[zelle] = formel
        ws[zelle].font = Font(name='Arial', size=8, color='FF808080')
        knoten.append(dict(zelle=zelle, ebene=ebene, **rest))
        z_zeile += 1

    K = f"Komponenten!$A${K_START}:$A${K_ENDE}"
    P = f"Komponenten!$B${K_START}:$B${K_ENDE}"
    A = f"Komponenten!$C${K_START}:$C${K_ENDE}"
    quelle(f'="SMARTINFRA"&CHAR(10)&COUNTA(Komponenten!$D${K_START}:$D${K_ENDE})&'
           f'" Komponenten / Schnittstellen"', 'wurzel',
           spalte_von=0, spalte_bis=len(reihenfolge) - 1)
    stufen = [('ve', 1, 'A', lambda r: f'COUNTIF({K},Komponenten!$A{r})',
               ' Komponente', ' Komponenten'),
              ('partner', 2, 'B',
               lambda r: f'COUNTIFS({K},Komponenten!$A{r},{P},Komponenten!$B{r})',
               ' Komponente', ' Komponenten'),
              ('art', 3, 'C',
               lambda r: (f'COUNTIFS({K},Komponenten!$A{r},{P},Komponenten!$B{r},'
                          f'{A},Komponenten!$C{r})'), ' Eintrag', ' Einträge')]
    for ebene, tiefe, feld, zaehler, einzahl, mehrzahl in stufen:
        gesehen = []
        for s in reihenfolge:
            if s[:tiefe] in gesehen:
                continue
            gesehen.append(s[:tiefe])
            r = gruppen[s][0]
            idx = [spalte_von[x] for x in reihenfolge if x[:tiefe] == s[:tiefe]]
            z = zaehler(r)
            quelle(f'=Komponenten!${feld}{r}&CHAR(10)&{z}&IF({z}=1,"{einzahl}","{mehrzahl}")',
                   ebene, spalte_von=min(idx), spalte_bis=max(idx))
    for s in reihenfolge:
        for platz, r in enumerate(gruppen[s]):
            quelle(f'=IF(Komponenten!$D{r}="","",Komponenten!$D{r}&CHAR(10)&Komponenten!$E{r}&'
                   f'IF(Komponenten!$E{r}=1," Anforderung"," Anforderungen")&" · "&'
                   f'Komponenten!$H{r}&"/"&Komponenten!$E{r}&" erfüllt"&CHAR(10)&"Status: "&'
                   f'Komponenten!$K{r})',
                   'komponente', spalte_von=spalte_von[s], spalte_bis=spalte_von[s],
                   platz=platz, status_zelle=f'Komponenten!$K${r}')

    # --- Statusansicht in Zellen --------------------------------------------------
    s0 = start + zeilen + 1
    ws.cell(s0, 1).value = 'Statusansicht – dieselbe Struktur, Farbe folgt dem Komponentenstatus'
    ws.cell(s0, 1).font = Font(name='Arial', size=12, bold=True, color=C_TITEL)
    ws.merge_cells(start_row=s0, start_column=1, end_row=s0, end_column=5)
    for i in range(K_ENDE - K_START + 1):
        o, t, v = s0 + 1 + i, K_START + i, K_START + i - 1
        ws.cell(o, 1).value = (f'=IF(Komponenten!$D{t}="","",IF(Komponenten!$A{t}='
                               f'Komponenten!$A{v},"",Komponenten!$A{t}))')
        ws.cell(o, 2).value = (f'=IF(Komponenten!$D{t}="","",IF(AND(Komponenten!$A{t}='
                               f'Komponenten!$A{v},Komponenten!$B{t}=Komponenten!$B{v}),"",'
                               f'"└ "&Komponenten!$B{t}))')
        ws.cell(o, 3).value = (f'=IF(Komponenten!$D{t}="","",IF(AND(Komponenten!$A{t}='
                               f'Komponenten!$A{v},Komponenten!$B{t}=Komponenten!$B{v},'
                               f'Komponenten!$C{t}=Komponenten!$C{v}),"","└ "&Komponenten!$C{t}))')
        ws.merge_cells(start_row=o, start_column=4, end_row=o, end_column=5)
        ws.cell(o, 4).value = (f'=IF(Komponenten!$D{t}="","","└ "&Komponenten!$D{t}&CHAR(10)&'
                               f'Komponenten!$E{t}&" Anforderungen · "&Komponenten!$H{t}&"/"&'
                               f'Komponenten!$E{t}&" erfüllt"&CHAR(10)&"Status: "&'
                               f'Komponenten!$K{t})')
        for c in range(1, 6):
            zelle = ws.cell(o, c)
            zelle.border = RAHMEN
            zelle.alignment = Alignment(wrap_text=True, vertical='center',
                                        horizontal='center' if c <= 3 else 'left')
            zelle.font = Font(name='Arial', size=9, bold=(c == 1),
                              color='FFFFFFFF' if c == 1 else 'FF000000')
        ws.row_dimensions[o].height = 46
    s_ende = s0 + (K_ENDE - K_START + 1)
    ws.conditional_formatting = ConditionalFormattingList()
    for spalte, farbe, schrift in (('A', C_VE, 'FFFFFFFF'), ('B', C_PARTNER, 'FFFFFFFF'),
                                   ('C', C_ART, C_TITEL)):
        ws.conditional_formatting.add(f'{spalte}{s0 + 1}:{spalte}{s_ende}',
                                      regel(f'${spalte}{s0 + 1}<>""', farbe, schrift, True))
    for wert, farbe, schrift, fett in STATUSFARBEN:
        ws.conditional_formatting.add(
            f'D{s0 + 1}:E{s_ende}',
            regel(f'Komponenten!$K${K_START}="{wert}"', farbe, schrift, fett))

    l0 = s_ende + 2
    ws.cell(l0, 1).value = 'Legende Statusfarben'
    ws.cell(l0, 1).font = Font(name='Arial', size=10, bold=True, color=C_TITEL)
    for i, (wert, farbe, _s, _f) in enumerate(STATUSFARBEN):
        zelle = ws.cell(l0 + 1, i + 1)
        zelle.value = wert
        zelle.fill = PatternFill('solid', fgColor=farbe)
        zelle.border = RAHMEN
        zelle.font = Font(name='Arial', size=9)
        zelle.alignment = Alignment(horizontal='center', vertical='center')

    bauplan = dict(blatt='Organigramm', box_breite=BOX_B, raster=RASTER,
                   dia_breite=len(reihenfolge) * RASTER, dia_hoehe=dia_hoehe,
                   ebenen={**{k: list(v) for k, v in EBENEN.items()},
                           'komponente': [Y_KOMP, H_KOMP, RASTER_KOMP]},
                   y_versatz_pt=sum(ws.row_dimensions[r].height for r in range(1, start)),
                   knoten=knoten)
    json.dump(bauplan, open('organigramm.json', 'w'), ensure_ascii=False, indent=1)
    wb.save(pfad)
    print(f'Reiter „Organigramm“ angelegt: {len(reihenfolge)} Gruppen, {len(knoten)} Boxen, '
          f'Zeichenfläche Zeile {start}-{start + zeilen - 1}, Statusansicht ab {s0}')


if __name__ == '__main__':
    main()
