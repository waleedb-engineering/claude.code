# -*- coding: utf-8 -*-
"""Schreibt das Organigramm als echte Excel-Zeichenobjekte in die Mappe.

Jede Box ist ein Zeichenobjekt (abgerundetes Rechteck), dessen Beschriftung
ueber das Attribut textlink mit einer Formelzelle verbunden ist - genau die
Verknuepfung, die Excel anlegt, wenn man bei einer markierten Form "=Zelle" in
die Bearbeitungsleiste schreibt. Damit aktualisiert sich die Beschriftung
automatisch, sobald sich Name, Anzahl oder Status in der Tabelle aendern.

Aufruf:  python3 inject_shapes.py <mappe.xlsx>
Voraussetzung: organigramm.json aus build.py und eine bereits neu berechnete
Mappe (die Formelwerte dienen als Anzeigetext, bis Excel neu rechnet).
"""
import json, re, shutil, sys, zipfile
from xml.sax.saxutils import escape

import openpyxl

PX = 9525                                     # 1 Pixel in EMU
PT = 12700                                    # 1 Punkt in EMU

FARBE = {
    'wurzel':     ('1F3864', 'FFFFFF', 900, True),
    've':         ('1F4E78', 'FFFFFF', 900, True),
    'partner':    ('2E75B6', 'FFFFFF', 900, True),
    'art':        ('BDD7EE', '1F3864', 800, True),
}
STATUSFARBE = {                               # Fuellung der Komponentenboxen
    'erfasst': ('E7E6E6', '404040'),
    'bewertet': ('D9EAF7', '1F3864'),
    'erfüllt': ('E2F0D9', '006100'),
    'trifft nicht zu': ('D9D9D9', '666666'),
    'keine Anforderungen': ('F2F2F2', '777777'),
}
LINIE = '8EA9DB'
RAND = '1F4E78'


class Zeichnung:
    """Sammelt Zeichenobjekte und gibt das drawing-XML aus."""

    def __init__(self):
        self.teile = []
        self.id = 1

    def _naechste_id(self):
        self.id += 1
        return self.id

    def box(self, x, y, b, h, text, textlink, fuellung, schriftfarbe, groesse, fett):
        i = self._naechste_id()
        guid = '{%08X-1B1B-4E4E-9A9A-%012X}' % (i * 7919, i * 104729)
        anzeige = escape(text)
        self.teile.append(f'''<xdr:absoluteAnchor>
<xdr:pos x="{int(x * PX)}" y="{int(y * PX)}"/><xdr:ext cx="{int(b * PX)}" cy="{int(h * PX)}"/>
<xdr:sp macro="" textlink="{textlink}">
<xdr:nvSpPr><xdr:cNvPr id="{i}" name="Organigramm {i}"/><xdr:cNvSpPr/></xdr:nvSpPr>
<xdr:spPr><a:xfrm><a:off x="{int(x * PX)}" y="{int(y * PX)}"/>\
<a:ext cx="{int(b * PX)}" cy="{int(h * PX)}"/></a:xfrm>
<a:prstGeom prst="roundRect"><a:avLst><a:gd name="adj" fmla="val 8000"/></a:avLst></a:prstGeom>
<a:solidFill><a:srgbClr val="{fuellung}"/></a:solidFill>
<a:ln w="9525"><a:solidFill><a:srgbClr val="{RAND}"/></a:solidFill></a:ln></xdr:spPr>
<xdr:txBody><a:bodyPr vertOverflow="overflow" horzOverflow="overflow" wrap="square" \
lIns="27432" tIns="9144" rIns="27432" bIns="9144" anchor="ctr"/><a:lstStyle/>
<a:p><a:pPr algn="ctr"/><a:fld id="{guid}" type="TxLink">
<a:rPr lang="de-DE" sz="{groesse}" b="{1 if fett else 0}">\
<a:solidFill><a:srgbClr val="{schriftfarbe}"/></a:solidFill>\
<a:latin typeface="Arial"/></a:rPr>
<a:pPr algn="ctr"/><a:t>{anzeige}</a:t></a:fld></a:p></xdr:txBody></xdr:sp>
<xdr:clientData/></xdr:absoluteAnchor>''')

    def strich(self, x, y, b, h):
        i = self._naechste_id()
        b, h = max(b, 1), max(h, 1)
        self.teile.append(f'''<xdr:absoluteAnchor>
<xdr:pos x="{int(x * PX)}" y="{int(y * PX)}"/><xdr:ext cx="{int(b * PX)}" cy="{int(h * PX)}"/>
<xdr:sp macro="">
<xdr:nvSpPr><xdr:cNvPr id="{i}" name="Verbinder {i}"/><xdr:cNvSpPr/></xdr:nvSpPr>
<xdr:spPr><a:xfrm><a:off x="{int(x * PX)}" y="{int(y * PX)}"/>\
<a:ext cx="{int(b * PX)}" cy="{int(h * PX)}"/></a:xfrm>
<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>
<a:solidFill><a:srgbClr val="{LINIE}"/></a:solidFill><a:ln><a:noFill/></a:ln></xdr:spPr>
<xdr:txBody><a:bodyPr/><a:lstStyle/><a:p><a:endParaRPr lang="de-DE"/></a:p></xdr:txBody>\
</xdr:sp><xdr:clientData/></xdr:absoluteAnchor>''')

    def xml(self):
        return ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/'
                'spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/'
                '2006/main">' + ''.join(self.teile) + '</xdr:wsDr>')


def baue_zeichnung(plan, texte, statuswerte):
    z = Zeichnung()
    raster, box_b = plan['raster'], plan['box_breite']
    y0 = plan['y_versatz_pt'] * PT / PX + 6            # Zeichenflaeche beginnt hier
    x0 = 6
    eb = plan['ebenen']

    def mitte(k):
        links = k['spalte_von'] * raster
        rechts = k['spalte_bis'] * raster + box_b
        return x0 + (links + rechts) / 2

    nach_ebene = {}
    for k in plan['knoten']:
        nach_ebene.setdefault(k['ebene'], []).append(k)

    breite = {'wurzel': 300, 've': 260, 'partner': box_b, 'art': box_b - 40}
    for ebene in ('wurzel', 've', 'partner', 'art'):
        y, h = eb[ebene][0], eb[ebene][1]
        f, s, g, fett = FARBE[ebene]
        for k in nach_ebene.get(ebene, []):
            b = breite[ebene]
            z.box(mitte(k) - b / 2, y0 + y, b, h, texte[k['zelle']],
                  f"Komponenten!${k['zelle'][0]}${k['zelle'][1:]}", f, s, g, fett)

    # Komponentenboxen haengend unter der jeweiligen Art
    y_k, h_k, raster_k = eb['komponente']
    komp_b = box_b - 24
    for k in nach_ebene.get('komponente', []):
        x = x0 + k['spalte_von'] * raster + 24
        y = y0 + y_k + k['platz'] * raster_k
        status = statuswerte.get(k['status_zelle']) or 'erfasst'
        f, s = STATUSFARBE.get(status, STATUSFARBE['erfasst'])
        name = texte[k['zelle']].split('\n')[0]
        groesse = 700 if len(name) > 42 else 800
        z.box(x, y, komp_b, h_k, texte[k['zelle']],
              f"Komponenten!${k['zelle'][0]}${k['zelle'][1:]}", f, s, groesse, False)

    # ---- Verbinder -------------------------------------------------------
    def bus(oben, unten_liste, y_oben_unterkante, y_unten_oberkante):
        """Senkrecht runter, waagerecht verteilen, senkrecht zu jedem Kind."""
        y_bus = (y_oben_unterkante + y_unten_oberkante) / 2
        z.strich(mitte(oben) - 1, y_oben_unterkante, 2, y_bus - y_oben_unterkante)
        mitten = [mitte(u) for u in unten_liste]
        if len(mitten) > 1:
            z.strich(min(mitten), y_bus - 1, max(mitten) - min(mitten), 2)
        for m in mitten:
            z.strich(m - 1, y_bus, 2, y_unten_oberkante - y_bus)

    wurzel = nach_ebene['wurzel'][0]
    bus(wurzel, nach_ebene['ve'], y0 + eb['wurzel'][0] + eb['wurzel'][1], y0 + eb['ve'][0])
    for ve in nach_ebene['ve']:
        kinder = [p for p in nach_ebene['partner']
                  if p['spalte_von'] >= ve['spalte_von'] and p['spalte_bis'] <= ve['spalte_bis']]
        bus(ve, kinder, y0 + eb['ve'][0] + eb['ve'][1], y0 + eb['partner'][0])
    for p in nach_ebene['partner']:
        kinder = [a for a in nach_ebene['art']
                  if a['spalte_von'] >= p['spalte_von'] and a['spalte_bis'] <= p['spalte_bis']]
        bus(p, kinder, y0 + eb['partner'][0] + eb['partner'][1], y0 + eb['art'][0])
    for a in nach_ebene['art']:
        kinder = [k for k in nach_ebene['komponente'] if k['spalte_von'] == a['spalte_von']]
        if not kinder:
            continue
        spine_x = x0 + a['spalte_von'] * raster + 12
        y_von = y0 + eb['art'][0] + eb['art'][1]
        y_bis = y0 + y_k + (len(kinder) - 1) * raster_k + h_k / 2
        z.strich(spine_x - 1, y_von, 2, y_bis - y_von)
        for k in kinder:
            y_mitte = y0 + y_k + k['platz'] * raster_k + h_k / 2
            z.strich(spine_x, y_mitte - 1, 12, 2)
    return z


def einfuegen(pfad, drawing_xml, blattdatei):
    """Legt das drawing-Teil an und verknuepft es mit dem Arbeitsblatt."""
    tmp = pfad + '.tmp'
    with zipfile.ZipFile(pfad) as quelle:
        namen = quelle.namelist()
        nummer = 1
        while f'xl/drawings/drawing{nummer}.xml' in namen:
            nummer += 1
        teil = f'xl/drawings/drawing{nummer}.xml'
        rels_name = f'xl/worksheets/_rels/{blattdatei}.rels'
        blatt = quelle.read(f'xl/worksheets/{blattdatei}').decode('utf-8')
        rels = (quelle.read(rels_name).decode('utf-8') if rels_name in namen else
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n<Relationships '
                'xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '</Relationships>')
        ct = quelle.read('[Content_Types].xml').decode('utf-8')

        rid = 'rIdOrga'
        rels = rels.replace('</Relationships>', (
            f'<Relationship Id="{rid}" Type="http://schemas.openxmlformats.org/officeDocument/'
            f'2006/relationships/drawing" Target="../drawings/drawing{nummer}.xml"/>'
            '</Relationships>'))
        if '<drawing ' not in blatt:
            # Reihenfolge laut Schema: drawing steht vor tableParts und extLst
            marke = blatt.rfind('<extLst>')
            if marke == -1:
                marke = blatt.rfind('</worksheet>')
            blatt = blatt[:marke] + f'<drawing r:id="{rid}"/>' + blatt[marke:]
        if f'/xl/drawings/drawing{nummer}.xml' not in ct:
            ct = ct.replace('</Types>', (
                f'<Override PartName="/{teil}" ContentType="application/vnd.openxmlformats-'
                f'officedocument.drawing+xml"/></Types>'))

        ersetzt = {f'xl/worksheets/{blattdatei}': blatt, rels_name: rels,
                   '[Content_Types].xml': ct}
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as ziel:
            for eintrag in quelle.infolist():
                daten = ersetzt.get(eintrag.filename)
                ziel.writestr(eintrag, daten.encode('utf-8') if daten is not None
                              else quelle.read(eintrag.filename))
            if rels_name not in namen:
                ziel.writestr(rels_name, rels.encode('utf-8'))
            ziel.writestr(teil, drawing_xml.encode('utf-8'))
    shutil.move(tmp, pfad)
    return teil


def blattdatei_finden(pfad, blattname):
    wb = openpyxl.load_workbook(pfad)
    index = wb.sheetnames.index(blattname)
    with zipfile.ZipFile(pfad) as z:
        wbxml = z.read('xl/workbook.xml').decode('utf-8')
        relxml = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')
    rid = re.findall(r'<sheet[^>]*r:id="([^"]+)"', wbxml)[index]
    ziel = re.search(rf'Id="{rid}"[^>]*Target="([^"]+)"', relxml).group(1)
    return ziel.split('/')[-1]


def main():
    pfad = sys.argv[1]
    plan = json.load(open('organigramm.json', encoding='utf-8'))
    wb = openpyxl.load_workbook(pfad, data_only=True)
    ws = wb[plan['blatt']]
    texte = {k['zelle']: str(ws[k['zelle']].value or '') for k in plan['knoten']}
    fehlend = [z for z, t in texte.items() if not t]
    if fehlend:
        raise SystemExit(f'Textquellen ohne berechneten Wert: {fehlend} – '
                         'bitte die Mappe zuerst neu berechnen lassen.')
    statuswerte = {k['status_zelle']: ws[k['status_zelle']].value
                   for k in plan['knoten'] if 'status_zelle' in k}
    zeichnung = baue_zeichnung(plan, texte, statuswerte)
    blattdatei = blattdatei_finden(pfad, plan['blatt'])
    teil = einfuegen(pfad, zeichnung.xml(), blattdatei)
    print(f'{zeichnung.id - 1} Zeichenobjekte in {teil} ({blattdatei}) geschrieben')


if __name__ == '__main__':
    main()
