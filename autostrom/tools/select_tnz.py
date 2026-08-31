# -*- coding: utf-8 -*-
"""Setzt den Status "trifft nicht zu" fuer eindeutig nicht einschlaegige Zeilen.

Nur drei Gruppen, die sich objektiv am Text nachweisen lassen:
  A  Antwort des Auftraggebers ist ein reiner Verweis auf eine andere Bieterfrage
  B  Begriffsbestimmungen aus § 4 des Betreibervertrags
  C  Regelungen, die ausschliesslich den Auftraggeber betreffen

Bestehende Eintraege werden nie ueberschrieben: geaendert wird nur, was auf
"erfasst" steht und noch keinen Nachweis traegt.

Aufruf:  python3 select_tnz.py <mappe.xlsx>
"""
import json, re, sys
import openpyxl

SP = dict(id=2, dokument=3, fundstelle=4, themenfeld=5, kernaussage=6, text=7,
          komponente=10, status=11, datum=12, nachweis=13)
VERMERK = 'Automatisch ausgewertet: '


KATALOG = {x['nr']: x for x in json.load(open('bieterfragen.json', encoding='utf-8'))} \
    if __import__('os').path.exists('bieterfragen.json') else {}
ZUSATZ = re.compile(r'(^|\s)(Ja,|Nein,|Verständnis|vgl\.|Zu i|Zu 1|Zu 2|Danach)')


def reine_ankuendigung(fundstelle):
    """Wahr, wenn die Antwort nur eine Unterlagenanpassung ankuendigt.

    Enthaelt sie darueber hinaus eine eigene Aussage, bleibt die Zeile offen.
    """
    treffer = re.match(r'Nr\. (\d+) /', fundstelle) or re.match(r'§ (\d+)$', fundstelle)
    antwort = KATALOG.get(treffer.group(1), {}).get('antwort', '') if treffer else ''
    return bool(antwort) and len(antwort) <= 200 and not ZUSATZ.search(antwort)


def gruppe(dokument, fundstelle, themenfeld, kernaussage, text, komponente):
    """Liefert (Kurzname, Begruendung) oder None."""
    treffer = re.match(r'Verweis des Auftraggebers auf die Antwort zu Bieterfrage ([\d,\s.und-]+)',
                       kernaussage)
    if treffer:
        nummer = treffer.group(1).strip().rstrip('.')
        return ('Verweis', f'{VERMERK}Die Antwort des Auftraggebers verweist ausschließlich auf '
                           f'Bieterfrage {nummer}. Keine eigenständige Anforderung – der Inhalt '
                           f'ist unter der dort erfassten Zeile nachzuhalten.')
    if dokument == 'Betreibervertrag' and fundstelle.startswith('§ 4 ('):
        return ('Begriffsbestimmung',
                f'{VERMERK}Begriffsbestimmung aus § 4 des Betreibervertrags. Definition ohne '
                f'eigene Leistungspflicht des Auftragnehmers.')
    if kernaussage.startswith('Anpassung der Unterlagen:') and reine_ankuendigung(fundstelle):
        return ('Unterlagenanpassung',
                f'{VERMERK}Die Antwort kündigt lediglich eine Anpassung der Vergabeunterlagen an. '
                f'Die Anforderung selbst steht im geänderten Dokument und ist dort nachzuhalten.')
    if themenfeld == 'Vergabeverfahren & Angebot' and komponente == '–':
        return ('Vergabeverfahren',
                f'{VERMERK}Betrifft das Vergabeverfahren (Angebot, Wertung, Unterlagen), nicht die '
                f'zu erbringende Leistung. Keiner Komponente zugeordnet.')
    return None


def main():
    pfad = sys.argv[1]
    wb = openpyxl.load_workbook(pfad)
    ws = wb['Anforderungen']
    gezaehlt, uebergangen = {}, 0
    for r in range(7, 2100):
        if not ws.cell(r, SP['id']).value:
            continue
        if ws.cell(r, SP['status']).value != 'erfasst':
            continue
        if ws.cell(r, SP['nachweis']).value:              # eigene Notiz nicht anfassen
            uebergangen += 1
            continue
        werte = [str(ws.cell(r, SP[k]).value or '')
                 for k in ('dokument', 'fundstelle', 'themenfeld', 'kernaussage', 'text',
                           'komponente')]
        ergebnis = gruppe(*werte)
        if not ergebnis:
            continue
        name, begruendung = ergebnis
        ws.cell(r, SP['status']).value = 'trifft nicht zu'
        ws.cell(r, SP['nachweis']).value = begruendung
        gezaehlt[name] = gezaehlt.get(name, 0) + 1
    print('auf „trifft nicht zu" gesetzt:', sum(gezaehlt.values()), gezaehlt)
    print('wegen vorhandener Notiz übergangen:', uebergangen)
    wb.save(pfad)
    print('gespeichert:', pfad)


if __name__ == '__main__':
    main()
