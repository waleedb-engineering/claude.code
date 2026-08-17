# -*- coding: utf-8 -*-
"""Prueft die fertige Mappe nach dem Neuberechnen."""
import json, sys
import openpyxl

datei = sys.argv[1] if len(sys.argv) > 1 else 'check.xlsx'
wb = openpyxl.load_workbook(datei, data_only=True)
wf = openpyxl.load_workbook(datei)

print('=== COCKPIT ===')
cp = wb['Cockpit']
for r in range(6, 20):
    if cp.cell(r, 1).value:
        print(f'  {cp.cell(r,1).value:42s} {cp.cell(r,2).value}')

print('\n=== KOMPONENTEN (Tabelle) ===')
k = wb['Komponenten']
print('  {:16s} {:22s} {:13s} {:44s} {:>5s} {:>4s} {:>4s} {:>4s} {:>4s}  {}'.format(
    'Vergabeeinheit', 'Partner', 'Art', 'Komponente', 'ges', 'erf', 'bew', 'erf', 'tnz', 'Status'))
for r in range(6, 30):
    if k.cell(r, 4).value:
        print('  {:16s} {:22s} {:13s} {:44s} {:>5} {:>4} {:>4} {:>4} {:>4}  {}'.format(
            str(k.cell(r, 1).value)[:16], str(k.cell(r, 2).value)[:22],
            str(k.cell(r, 3).value)[:13], str(k.cell(r, 4).value)[:44],
            k.cell(r, 5).value, k.cell(r, 6).value, k.cell(r, 7).value,
            k.cell(r, 8).value, k.cell(r, 9).value, k.cell(r, 10).value))

print('\n=== ORGANIGRAMM ===')
print(' Wurzel:', repr(k.cell(34, 1).value))
for r in range(36, 60):
    zeilen = [str(k.cell(r, c).value or '').replace('\n', ' | ') for c in range(1, 5)]
    if any(zeilen):
        print(f'  {r:3d}  {zeilen[0]:34s} {zeilen[1]:30s} {zeilen[2]:24s} {zeilen[3]}')

print('\n=== AUSWERTUNG ===')
a = wb['Auswertung']
for r in range(1, a.max_row + 1):
    v = [a.cell(r, c).value for c in range(1, 8)]
    if v[0] is not None:
        print('  ' + ' | '.join('' if x is None else (f'{x:.0%}' if isinstance(x, float) and c == 6
                                                      else str(x))
                                for c, x in enumerate(v)))

print('\n=== FEHLERPRUEFUNG ===')
fehler = []
for sh in wb.worksheets:
    for row in sh.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith('#') and c.value.endswith(('!', '?')):
                fehler.append((sh.title, c.coordinate, c.value))
print('  Formelfehler:', fehler if fehler else 'keine')

print('\n=== STICHPROBE gegen PDF ===')
daten = {d['nr']: d for d in json.load(open('bieterfragen.json', encoding='utf-8'))}
ws = wf['Anforderungen']
index = {}
for r in range(7, 713):
    nr = ws.cell(r, 13).value
    if nr is not None:
        index[str(nr)] = r
for nr in ['1', '33', '174', '351', '441', '506', '587', '659']:
    r = index.get(nr)
    if not r:
        print(f'  Nr {nr}: NICHT gefunden')
        continue
    d = daten[nr]
    ok_frage = (ws.cell(r, 6).value or '')[:60] == d['frage'][:60]
    ok_antwort = (ws.cell(r, 14).value or '')[:60] == d['antwort'][:60]
    print(f"  Nr {nr:>3s} Zeile {r:>3d} | Frage-Text {'OK' if ok_frage else 'ABWEICHUNG (Bestand)'}"
          f" | Antwort {'OK' if ok_antwort else 'ABWEICHUNG'}")
    print(f'      Themenfeld: {ws.cell(r,4).value}')
    print(f'      Kernaussage: {str(ws.cell(r,5).value)[:150]}')
    print(f'      Komponente: {ws.cell(r,9).value}   Partner: {ws.cell(r,8).value}'
          f'   Status: {ws.cell(r,10).value}')

print('\n=== ZAEHLUNGEN ===')
ids = [ws.cell(r, 1).value for r in range(7, 800) if ws.cell(r, 1).value]
print('  Anforderungen gesamt:', len(ids), '| erste:', ids[0], '| letzte:', ids[-1])
print('  eindeutige IDs:', len(set(ids)) == len(ids))
bf = [ws.cell(r, 13).value for r in range(7, 800) if ws.cell(r, 13).value is not None]
print('  Datensätze mit Bieterfrage-Nr.:', len(bf), '| eindeutig:', len(set(bf)) == len(bf))
nummern = sorted(int(x) for x in bf if isinstance(x, int))
print('  Fragennummern:', nummern[0], '-', nummern[-1], '| Anzahl:', len(nummern),
      '| lückenlos:', nummern == list(range(1, 660)))
print('  Sondereinträge:', [x for x in bf if not isinstance(x, int)])
