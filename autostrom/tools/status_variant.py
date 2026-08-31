# -*- coding: utf-8 -*-
"""Loest die Vermischung ueber einen zusaetzlichen Statuswert statt ueber eine Spalte.

Der Statusprozess bekommt einen fuenften Wert "außerhalb Umfang". Er wirkt wie
"trifft nicht zu" auf den Erfuellungsgrad, bleibt davon aber unterscheidbar:
"trifft nicht zu" heisst "keine Anforderung", "außerhalb Umfang" heisst
"nicht unser Paket".

Angepasst werden Register, Cockpit, Auswertung und Komponententabelle.

Aufruf:  python3 status_variant.py <mappe.xlsx>
"""
import copy, sys
import openpyxl
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

NEU = 'außerhalb Umfang'
TNZ = 'trifft nicht zu'
STATUS = ['erfasst', 'bewertet', 'erfüllt', TNZ, NEU]
FARBEN = {'erfasst': ('FFE7E6E6', None, None), 'bewertet': ('FFD9EAF7', None, None),
          'erfüllt': ('FFE2F0D9', 'FF006100', True), TNZ: ('FFD9D9D9', 'FF666666', None),
          NEU: ('FFEDEDED', 'FF9C6500', None)}
GRENZE = 2000
NOTIZ = ('Automatisch ausgewertet: Außerhalb des betrachteten Leistungsumfangs '
         '(Schranken, Kennzeichenerfassung, dynamische Anzeige).')


def dxf(fuellung, farbe=None, fett=None):
    return DifferentialStyle(font=Font(color=farbe, b=fett) if (farbe or fett) else None,
                             fill=PatternFill(bgColor=fuellung))


def main():
    pfad = sys.argv[1]
    wb = openpyxl.load_workbook(pfad)

    # ---------------------------------------------------------------- Register
    ws = wb['Anforderungen']
    gesetzt, letzte = 0, 6
    for r in range(7, GRENZE + 1):
        if not ws.cell(r, 2).value:
            continue
        letzte = r
        if ws.cell(r, 14).value == 'außerhalb' and ws.cell(r, 11).value == 'erfasst':
            ws.cell(r, 11).value = NEU
            ws.cell(r, 13).value = NOTIZ
            gesetzt += 1
        ws.cell(r, 14).value = None                  # Hilfsspalte wird nicht mehr gebraucht
    ws.cell(6, 14).value = None
    print(f'auf „{NEU}" gesetzt: {gesetzt}')

    alt = ws.tables['tblAnforderungen']
    kopf = [c.name for c in alt.tableColumns if c.name != 'Umfang']
    stil_tab = alt.tableStyleInfo
    del ws.tables['tblAnforderungen']
    tab = Table(displayName='tblAnforderungen', ref=f'B6:M{letzte}',
                tableColumns=[TableColumn(id=i + 1, name=n) for i, n in enumerate(kopf)])
    if stil_tab is not None:
        tab.tableStyleInfo = TableStyleInfo(
            name=stil_tab.name, showFirstColumn=stil_tab.showFirstColumn,
            showLastColumn=stil_tab.showLastColumn, showRowStripes=stil_tab.showRowStripes,
            showColumnStripes=stil_tab.showColumnStripes)
    ws.add_table(tab)

    ws.data_validations.dataValidation = [
        d for d in ws.data_validations.dataValidation if 'N7' not in str(d.sqref)]
    for d in ws.data_validations.dataValidation:
        if d.type == 'list':
            d.formula1 = '"' + ','.join(STATUS) + '"'
    regeln = [(str(b.sqref), list(v)) for b, v in ws.conditional_formatting._cf_rules.items()]
    ws.conditional_formatting = ConditionalFormattingList()
    for bereich, liste in regeln:
        if bereich.startswith('N'):
            continue
        for regel in liste:
            ws.conditional_formatting.add(bereich, regel)
    fuellung, farbe, fett = FARBEN[NEU]
    ws.conditional_formatting.add(f'K7:K{GRENZE}', Rule(
        type='expression', formula=[f'K7="{NEU}"'], dxf=dxf(fuellung, farbe, fett)))

    # ---------------------------------------------------------------- Cockpit
    cp = wb['Cockpit']
    ohne = (f'COUNTIF(Anforderungen!$K$7:$K${GRENZE},"{TNZ}")'
            f'+COUNTIF(Anforderungen!$K$7:$K${GRENZE},"{NEU}")')
    for r in range(1, cp.max_row + 1):
        wert = cp.cell(r, 2).value
        if isinstance(wert, str) and 'Erfüllungsgrad' in str(cp.cell(r, 1).value or ''):
            cp.cell(r, 2).value = (
                f'=IFERROR(COUNTIF(Anforderungen!$K$7:$K${GRENZE},"erfüllt")/'
                f'(COUNTA(Anforderungen!$B$7:$B${GRENZE})-({ohne})),0)')
        if str(cp.cell(r, 1).value or '') == 'Anforderungen im betrachteten Umfang':
            cp.cell(r, 1).value = 'Außerhalb Umfang'
            cp.cell(r, 2).value = f'=COUNTIF(Anforderungen!$K$7:$K${GRENZE},"{NEU}")'
    zeile = max(r for r in range(1, cp.max_row + 2) if cp.cell(r, 1).value) + 1
    cp.cell(zeile, 1).value = 'Anwendbar (ohne „trifft nicht zu“ und „außerhalb Umfang“)'
    cp.cell(zeile, 1)._style = copy.copy(cp.cell(zeile - 1, 1)._style)
    cp.cell(zeile, 2).value = f'=COUNTA(Anforderungen!$B$7:$B${GRENZE})-({ohne})'
    cp.cell(zeile, 2)._style = copy.copy(cp.cell(zeile - 1, 2)._style)
    cp.cell(zeile, 2).number_format = '0'

    # ---------------------------------------------------------------- Auswertung
    aus = wb['Auswertung']
    for kopfzeile in [r for r in range(1, aus.max_row + 1) if aus.cell(r, 2).value == 'Gesamt']:
        aus.cell(kopfzeile, 7).value = 'Außerhalb Umfang'
        aus.cell(kopfzeile, 8).value = 'Erfüllungsgrad'
        aus.cell(kopfzeile, 8)._style = copy.copy(aus.cell(kopfzeile, 7)._style)
        aus.column_dimensions['H'].width = 15
        r = kopfzeile + 1
        while aus.cell(r, 2).value:
            aus.cell(r, 7).value = str(aus.cell(r, 6).value).replace(TNZ, NEU)
            aus.cell(r, 7)._style = copy.copy(aus.cell(r, 6)._style)
            aus.cell(r, 8).value = f'=IFERROR(E{r}/(B{r}-F{r}-G{r}),0)'
            aus.cell(r, 8)._style = copy.copy(aus.cell(r, 6)._style)
            aus.cell(r, 8).number_format = '0%'
            r += 1
    for r in range(1, aus.max_row + 1):                 # Gesamtstatus um den fuenften Wert
        if aus.cell(r, 1).value == TNZ and str(aus.cell(r, 2).value or '').startswith('=COUNTIF'):
            aus.cell(r + 1, 1).value = NEU
            aus.cell(r + 1, 1)._style = copy.copy(aus.cell(r, 1)._style)
            aus.cell(r + 1, 2).value = f'=COUNTIF(Anforderungen!$K$7:$K${GRENZE},A{r + 1})'
            aus.cell(r + 1, 2)._style = copy.copy(aus.cell(r, 2)._style)
            aus.cell(r + 1, 3).value = (f'=IFERROR(B{r + 1}/COUNTA(Anforderungen!'
                                        f'$B$7:$B${GRENZE}),0)')
            aus.cell(r + 1, 3)._style = copy.copy(aus.cell(r, 3)._style)
            aus.cell(r + 1, 3).number_format = '0%'
            break
    for r in range(1, aus.max_row + 1):
        if str(aus.cell(r, 1).value or '').startswith('Anwendbare Anforderungen'):
            aus.cell(r, 1).value = 'Anwendbare Anforderungen'
            aus.cell(r, 2).value = f'=COUNTA(Anforderungen!$B$7:$B${GRENZE})-({ohne})'
            aus.cell(r, 3).value = 'Gesamt abzüglich „trifft nicht zu“ und „außerhalb Umfang“'

    # ---------------------------------------------------------------- Komponenten
    komp = wb['Komponenten']
    komp.cell(5, 12).value = komp.cell(5, 11).value                 # Bemerkung nach L
    komp.cell(5, 12)._style = copy.copy(komp.cell(5, 11)._style)
    komp.cell(5, 11).value = 'Status'
    komp.cell(5, 10).value = NEU
    komp.cell(5, 10)._style = copy.copy(komp.cell(5, 9)._style)
    for breite, spalte in ((13, 'J'), (18, 'K'), (40, 'L')):
        komp.column_dimensions[spalte].width = breite
    for r in range(6, 30):
        komp.cell(r, 12).value = komp.cell(r, 11).value             # alte Bemerkung
        komp.cell(r, 12)._style = copy.copy(komp.cell(r, 11)._style)
        komp.cell(r, 10).value = str(komp.cell(r, 9).value).replace(TNZ, NEU)
        komp.cell(r, 10)._style = copy.copy(komp.cell(r, 9)._style)
        komp.cell(r, 11).value = (
            f'=IF($D{r}="","",IF($E{r}=0,"keine Anforderungen",IF($F{r}>0,"erfasst",'
            f'IF($G{r}>0,"bewertet",IF($H{r}>0,"erfüllt",IF($I{r}>0,"{TNZ}","{NEU}"))))))')
        komp.cell(r, 11)._style = copy.copy(komp.cell(r, 10)._style)
    regeln = [(str(b.sqref), list(v)) for b, v in komp.conditional_formatting._cf_rules.items()]
    komp.conditional_formatting = ConditionalFormattingList()
    for bereich, liste in regeln:
        for regel in liste:
            komp.conditional_formatting.add(bereich.replace('J6:J29', 'K6:K29'), regel)
    for wert in (NEU, 'keine Anforderungen'):
        fuellung, farbe, fett = FARBEN.get(wert, ('FFF2F2F2', 'FF777777', None))
        komp.conditional_formatting.add('K6:K29', Rule(
            type='expression', formula=[f'K6="{wert}"'], dxf=dxf(fuellung, farbe, fett)))

    # ---------------------------------------------------------------- Lesehilfe
    lese = wb['Lesehilfe']
    for r in range(1, lese.max_row + 1):
        if lese.cell(r, 1).value == 'Umfang':
            lese.cell(r, 1).value = f'Status „{NEU}“'
            lese.cell(r, 2).value = (
                'Fünfter Statuswert. Er bedeutet: die Anforderung gilt für das Projekt, gehört '
                'aber nicht zum betrachteten Paket (Schranken, Kennzeichenerfassung, dynamische '
                'Anzeige). Für den Erfüllungsgrad zählt sie wie „trifft nicht zu“ nicht mit, '
                'bleibt davon aber unterscheidbar: „trifft nicht zu“ heißt „keine Anforderung“.')
    wb.save(pfad)
    print('gespeichert:', pfad)


if __name__ == '__main__':
    main()
