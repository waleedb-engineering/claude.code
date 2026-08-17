# -*- coding: utf-8 -*-
"""Erweitert die bestehende Anforderungs-Mappe um den Bieterfragenkatalog,
die neue Komponentenstruktur (Vergabeeinheit | Partner | Art | ...) und ein
formelbasiertes Organigramm.

Die bestehende Datei wird geladen und gezielt ergaenzt - es wird keine neue
Mappe aufgebaut. Reiter, Design, Statuslogik und Formelsprache bleiben erhalten.
"""
import json, re, copy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import Rule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

QUELLE = 'original.xlsx'
ZIEL = '20260817_Autostrom_LKWLaden_Anforderungen_mit_Bieterfragenkatalog.xlsx'
MAXZ = 800           # Auswertungsgrenze im Register "Anforderungen"
ARIAL = 'Arial'

daten = json.load(open('bieterfragen.json', encoding='utf-8'))
wb = openpyxl.load_workbook(QUELLE)

# --------------------------------------------------------------------------
# Farben (uebernommen aus der bestehenden Datei)
# --------------------------------------------------------------------------
C_DUNKEL = 'FF1F4E78'
C_TITEL = 'FF1F3864'
C_KOPF = 'FF274B7A'
C_HELL = 'FFF2F2F2'
C_ERFASST = 'FFE7E6E6'
C_BEWERTET = 'FFD9EAF7'
C_ERFUELLT = 'FFE2F0D9'
C_TNZ = 'FFD9D9D9'
C_KEINE = 'FFF2F2F2'
C_PRUEFEN = 'FFFFF2CC'
C_VE = 'FF1F4E78'
C_PARTNER = 'FF2E75B6'
C_ART = 'FFBDD7EE'

duenn = Side(style='thin', color='FFBFBFBF')
RAHMEN = Border(left=duenn, right=duenn, top=duenn, bottom=duenn)


from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


def sauber(v):
    """Entfernt Steuerzeichen, die Excel in Zellen nicht zulaesst."""
    if isinstance(v, str):
        return ILLEGAL_CHARACTERS_RE.sub(' ', v)
    return v


def fill(hexcode):
    return PatternFill('solid', fgColor=hexcode)


def dxf_regel(formel, fuellung=None, schriftfarbe=None, fett=None):
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.styles import Font as F, PatternFill as P
    ds = DifferentialStyle(
        font=F(color=schriftfarbe, b=fett) if (schriftfarbe or fett) else None,
        fill=P(bgColor=fuellung) if fuellung else None)
    return Rule(type='expression', formula=[formel], dxf=ds, stopIfTrue=False)


# ==========================================================================
# 1  Register "Anforderungen"
# ==========================================================================
ws = wb['Anforderungen']

# 1a  Verwaiste Restzeilen ohne ID entfernen (verfaelschten die Statuszaehlung)
entfernt = []
for r in range(7, ws.max_row + 1):
    if ws.cell(r, 1).value in (None, '') and any(
            ws.cell(r, c).value not in (None, '') for c in range(2, 15)):
        entfernt.append((r, [ws.cell(r, c).value for c in range(1, 15)]))
        for c in range(1, 15):
            ws.cell(r, c).value = None
print('verwaiste Zeilen bereinigt:', entfernt)

# 1b  bereits vorhandene Bieterfragen ermitteln (Dokument = "Bieterfrage")
vorhandene_bf = {}
for r in range(7, 59):
    if str(ws.cell(r, 2).value or '').strip().lower().startswith('bieterfrage'):
        m = re.search(r'(\d+)', str(ws.cell(r, 3).value or ''))
        if m:
            vorhandene_bf[m.group(1)] = r
print('bereits vorhandene Bieterfragen:', sorted(vorhandene_bf, key=int))

nach_nr = {d['nr']: d for d in daten}

# 1c  Spaltenueberschriften: Bestand behutsam schaerfen + zwei neue Spalten
ws['F6'] = 'Anforderung / Frage (Originaltext)'
ws['M6'] = 'Bieterfrage-Nr.'
ws['N6'] = 'Antwort Auftraggeber (Original)'
for spalte in ('M', 'N'):
    ws[f'{spalte}6']._style = copy.copy(ws['L6']._style)
ws.merge_cells('M5:N5')
ws['M5'] = 'Bieterfragenkatalog (Original)'
ws['M5']._style = copy.copy(ws['J5']._style)
ws.column_dimensions['M'].width = 14
ws.column_dimensions['N'].width = 58

# 1d  Stilvorlagen aus einer bestehenden Datenzeile
stil_text = copy.copy(ws['A7']._style)
stil_status = copy.copy(ws['J7']._style)
stil_datum = copy.copy(ws['K7']._style)

# vorhandene Bieterfragen: Quelle vereinheitlichen + Original-Antwort ergaenzen
for nr, r in vorhandene_bf.items():
    ws.cell(r, 2).value = 'Bieterfragenkatalog'
    ws.cell(r, 13).value = int(nr)
    ws.cell(r, 13)._style = copy.copy(stil_text)
    d = nach_nr.get(nr)
    if d:
        ws.cell(r, 14).value = sauber(d['antwort'])
        ws.cell(r, 14)._style = copy.copy(stil_text)

# 1e  Restlichen Katalog unten anhaengen
start = 59
lfd = 52
zeile = start
angehaengt = 0
for d in daten:
    if d['nr'] in vorhandene_bf:
        continue
    lfd += 1
    if d['art'] == 'Frage':
        fundstelle = f"Nr. {d['nr']} / S. {d['seite']}"
        bfnr = int(d['nr'])
    else:
        fundstelle = f"{d['nr']} / S. {d['seite']}"
        bfnr = d['nr']
    bemerkung = f"Bezug lt. Katalog: {d['bezug'] or '–'}"
    if d['datum']:
        bemerkung += f"\nKatalogdatum: {d['datum']}"
    if 'Zuordnung prüfen' in (d['komponenten'], d['partner']):
        bemerkung += "\nZuordnung prüfen (Komponente/Partner nicht eindeutig)"
    werte = [f'ANF-{lfd:03d}', 'Bieterfragenkatalog', fundstelle, d['themenfeld'],
             d['kernaussage'], d['frage'], bemerkung, d['partner'], d['komponenten'],
             'erfasst', None, None, bfnr, d['antwort']]
    for c, v in enumerate(werte, start=1):
        zelle = ws.cell(zeile, c)
        zelle.value = sauber(v)
        if c == 10:
            zelle._style = copy.copy(stil_status)
        elif c == 11:
            zelle._style = copy.copy(stil_datum)
        else:
            zelle._style = copy.copy(stil_text)
    ws.row_dimensions[zeile].height = 48
    zeile += 1
    angehaengt += 1
letzte_zeile = zeile - 1
print(f'angehaengte Bieterfragen: {angehaengt}  (Zeilen {start}-{letzte_zeile})')

# 1f  Datumsformat im Erfuellungsblock vereinheitlichen
for r in range(7, MAXZ + 1):
    ws.cell(r, 11).number_format = 'DD.MM.YYYY'

# 1g  Tabelle neu aufspannen (14 Spalten)
kopf = [ws.cell(6, c).value for c in range(1, 15)]
alt = ws.tables['tblAnforderungen']
stil = alt.tableStyleInfo
del ws.tables['tblAnforderungen']
tab = Table(displayName='tblAnforderungen', ref=f'A6:N{letzte_zeile}',
            tableColumns=[TableColumn(id=i + 1, name=n) for i, n in enumerate(kopf)])
tab.tableStyleInfo = TableStyleInfo(name=stil.name, showFirstColumn=stil.showFirstColumn,
                                    showLastColumn=stil.showLastColumn,
                                    showRowStripes=stil.showRowStripes,
                                    showColumnStripes=stil.showColumnStripes)
ws.add_table(tab)

# 1h  Dropdown + Datumspruefung auf den gesamten Arbeitsbereich ziehen
ws.data_validations.dataValidation = []
dv_status = DataValidation(type='list', formula1='"erfasst,bewertet,erfüllt,trifft nicht zu"',
                           allow_blank=False, showDropDown=False,
                           showErrorMessage=True, showInputMessage=True)
dv_status.error = 'Bitte einen der vier Status auswählen.'
dv_status.errorTitle = 'Ungültiger Status'
dv_status.promptTitle = 'Status'
dv_status.prompt = 'erfasst → bewertet → erfüllt oder trifft nicht zu'
ws.add_data_validation(dv_status)
dv_status.add(f'J7:J{MAXZ}')
dv_datum = DataValidation(type='date', operator='greaterThanOrEqual',
                          formula1='DATE(2020,1,1)', allow_blank=True,
                          showErrorMessage=True, showInputMessage=True)
dv_datum.error = 'Bitte ein gültiges Datum für „Erfüllt am“ auswählen oder eingeben.'
dv_datum.errorTitle = 'Ungültiges Datum'
dv_datum.promptTitle = 'Erfüllt am'
dv_datum.prompt = 'Datum auswählen bzw. im Format TT.MM.JJJJ eingeben.'
ws.add_data_validation(dv_datum)
dv_datum.add(f'K7:K{MAXZ}')

# 1i  Bedingte Formatierung des Erfuellungsblocks neu spannen (Logik unveraendert)
ws.conditional_formatting = ConditionalFormattingList()
for wert, farbe, schrift, fett in [('erfasst', C_ERFASST, None, None),
                                   ('bewertet', C_BEWERTET, None, None),
                                   ('erfüllt', C_ERFUELLT, 'FF006100', True),
                                   ('trifft nicht zu', C_TNZ, 'FF666666', False)]:
    ws.conditional_formatting.add(f'J7:J{MAXZ}',
                                  dxf_regel(f'J7="{wert}"', farbe, schrift, fett))
for formel, farbe, schrift, fett in [
        ('AND($J7="erfüllt",$K7="")', 'FFF4CCCC', 'FF9C0006', True),
        ('AND($J7="erfüllt",$K7<>"")', C_ERFUELLT, None, None),
        ('AND($J7<>"erfüllt",$K7<>"")', 'FFFCE4D6', 'FF9C5700', False)]:
    ws.conditional_formatting.add(f'K7:K{MAXZ}', dxf_regel(formel, farbe, schrift, fett))
for formel, farbe, schrift, fett in [
        ('AND($J7="erfüllt",$L7="")', 'FFF4CCCC', 'FF9C0006', True),
        ('AND($J7="erfüllt",$L7<>"")', C_ERFUELLT, None, None),
        ('AND($J7="trifft nicht zu",$L7="")', 'FFF4CCCC', 'FF9C0006', True),
        ('AND($J7="trifft nicht zu",$L7<>"")', C_ERFASST, None, None)]:
    ws.conditional_formatting.add(f'L7:L{MAXZ}', dxf_regel(formel, farbe, schrift, fett))
# offene Zuordnungen sichtbar machen
for spalte in ('H', 'I'):
    ws.conditional_formatting.add(
        f'{spalte}7:{spalte}{MAXZ}',
        dxf_regel(f'{spalte}7="Zuordnung prüfen"', C_PRUEFEN, 'FF7F6000', False))

ws['A3'] = ('Logik: erfasst → bewertet → erfüllt. Bei „erfüllt“ sind Datum und Nachweis/Notiz '
            'erforderlich, bei „trifft nicht zu“ eine Begründung. Gelb markierte Zuordnungen '
            '(Partner/Komponente) sind fachlich zu prüfen.')
ws.freeze_panes = 'A7'
ws.auto_filter.ref = None

# ==========================================================================
# 2  Register "Komponenten": neue Struktur + Organigramm
# ==========================================================================
KOMP = wb['Komponenten']
for zeile_ in list(KOMP.merged_cells.ranges):
    KOMP.unmerge_cells(str(zeile_))
KOMP.conditional_formatting = ConditionalFormattingList()
for r in range(1, KOMP.max_row + 5):
    for c in range(1, 12):
        z = KOMP.cell(r, c)
        z.value = None
        z._style = copy.copy(wb['Lesehilfe']['B21']._style)
    if r in KOMP.row_dimensions:
        KOMP.row_dimensions[r].height = None

# Komponentenstamm: Werte stammen 1:1 aus der bisherigen Komponententabelle.
# Vergabeeinheit ist in keiner Quelle belegt -> "Zuordnung prüfen".
KOMPONENTEN = [
    # (Partner, Art, Komponente, Bemerkung)
    ('Aplitronic', 'Komponente', 'Ladeeinrichtung', ''),
    ('eRound', 'Komponente', 'Abrechnungskomponente', ''),
    ('eRound', 'Komponente', 'Authentifizierungskomponente', ''),
    ('eRound', 'Komponente', 'Monitoring', ''),
    ('eRound', 'Komponente', 'OCPI Buchungskalender', ''),
    ('Reservierungsplattform', 'Komponente', 'Anbindung', ''),
    ('Reservierungsplattform', 'Komponente', 'Reservierungsplattform', ''),
    ('SCS', 'Komponente', 'Belegungsdetektion', ''),
    ('SCS', 'Komponente', 'Dynamische Anzeige', ''),
    ('SCS', 'Komponente', 'Kennzeichenerfassung', ''),
    ('SCS', 'Komponente', 'Pönalisierungssystem', ''),
    ('SCS', 'Komponente', 'Schrankensteuerung', ''),
    ('tbd', 'Komponente', 'Beschilderung', ''),
    ('tbd', 'Komponente', 'Schrankenanlage', ''),
    ('eRound / SCS', 'Schnittstelle',
     'OCPI Buchungskalender / Abrechnungskomponente / Belegungsdetektion / Pönalisierungssystem',
     'Schnittstelle: zählt Anforderungen, die alle genannten Komponenten betreffen.'),
    ('eRound / SCS', 'Schnittstelle', 'OCPI Buchungskalender / Dynamische Anzeige',
     'Schnittstelle: zählt Anforderungen, die alle genannten Komponenten betreffen.'),
    ('eRound / SCS', 'Schnittstelle',
     'OCPI Buchungskalender / Kennzeichenerfassung / Pönalisierungssystem',
     'Schnittstelle: zählt Anforderungen, die alle genannten Komponenten betreffen.'),
    ('eRound / tbd', 'Schnittstelle', 'Authentifizierungskomponente / Schrankensteuerung',
     'Schnittstelle: zählt Anforderungen, die alle genannten Komponenten betreffen.'),
]
RESERVE = 6                                   # freie Zeilen fuer neue Komponenten
K_START = 6
K_ENDE = K_START + len(KOMPONENTEN) + RESERVE - 1

KOMP['A1'] = 'Komponenten & Schnittstellen'
KOMP['A1'].font = Font(name=ARIAL, size=16, bold=True, color=C_TITEL)
KOMP['A2'] = ('Struktur Vergabeeinheit → Partner → Art → Komponente. Anzahl und Status der '
              'Anforderungen werden automatisch aus dem Register „Anforderungen“ abgeleitet.')
KOMP['A2'].font = Font(name=ARIAL, size=10, color='FF595959')
KOMP['A3'] = ('Statuslogik: erfasst hat Vorrang vor bewertet; erst wenn keine offenen Vorstufen '
              'mehr vorhanden sind, wird erfüllt bzw. trifft nicht zu angezeigt. '
              'Nur die Felder Vergabeeinheit, Partner, Art, Komponente und Bemerkung werden '
              'gepflegt – alles Übrige rechnet sich selbst.')
KOMP['A3'].font = Font(name=ARIAL, size=9, color='FF808080')
KOMP.merge_cells('A2:K2')
KOMP.merge_cells('A3:K3')
KOMP['A2'].alignment = Alignment(wrap_text=True, vertical='center')
KOMP['A3'].alignment = Alignment(wrap_text=True, vertical='center')
KOMP.row_dimensions[1].height = 22
KOMP.row_dimensions[2].height = 28
KOMP.row_dimensions[3].height = 28

KOPF = ['Vergabeeinheit', 'Partner', 'Art', 'Komponente', 'Anforderungen', 'erfasst',
        'bewertet', 'erfüllt', 'trifft nicht zu', 'Status', 'Bemerkung / Schnittstelle']
BREITEN = [20, 20, 14, 34, 13, 10, 10, 10, 13, 18, 40]
for i, (t, b) in enumerate(zip(KOPF, BREITEN), start=1):
    z = KOMP.cell(5, i)
    z.value = t
    z.font = Font(name=ARIAL, size=10, bold=True, color='FFFFFFFF')
    z.fill = fill(C_DUNKEL)
    z.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    z.border = RAHMEN
    KOMP.column_dimensions[get_column_letter(i)].width = b
KOMP.row_dimensions[5].height = 32

ANF_I = f'Anforderungen!$I$7:$I${MAXZ}'
ANF_J = f'Anforderungen!$J$7:$J${MAXZ}'
ANF_A = f'Anforderungen!$A$7:$A${MAXZ}'


def zaehlformel(r, teile, status=None):
    """Zaehlt Anforderungen, deren Komponentenfeld alle Teilbegriffe enthaelt."""
    if teile is None:                       # generisch ueber den Zellwert
        bed = f'ISNUMBER(SEARCH($D{r},{ANF_I}))'
    else:
        bed = '*'.join(f'ISNUMBER(SEARCH("{t}",{ANF_I}))' for t in teile)
    zusatz = f'*({ANF_J}="{status}")' if status else f'*({ANF_A}<>"")'
    return f'=IF($D{r}="","",SUMPRODUCT({bed}{zusatz}))'


for i, (partner, art, komponente, bemerkung) in enumerate(KOMPONENTEN):
    r = K_START + i
    teile = [t.strip() for t in komponente.split(' / ')] if art == 'Schnittstelle' else None
    KOMP.cell(r, 1).value = 'Zuordnung prüfen'
    KOMP.cell(r, 2).value = partner
    KOMP.cell(r, 3).value = art
    KOMP.cell(r, 4).value = komponente
    KOMP.cell(r, 5).value = zaehlformel(r, teile)
    KOMP.cell(r, 6).value = zaehlformel(r, teile, 'erfasst')
    KOMP.cell(r, 7).value = zaehlformel(r, teile, 'bewertet')
    KOMP.cell(r, 8).value = zaehlformel(r, teile, 'erfüllt')
    KOMP.cell(r, 9).value = zaehlformel(r, teile, 'trifft nicht zu')
    KOMP.cell(r, 10).value = (
        f'=IF($D{r}="","",IF($E{r}=0,"keine Anforderungen",'
        f'IF($F{r}>0,"erfasst",IF($G{r}>0,"bewertet",'
        f'IF($H{r}>0,"erfüllt","trifft nicht zu")))))')
    KOMP.cell(r, 11).value = bemerkung or None

for r in range(K_START, K_ENDE + 1):
    if r >= K_START + len(KOMPONENTEN):      # Reservezeilen mit Formeln vorbereiten
        KOMP.cell(r, 5).value = zaehlformel(r, None)
        KOMP.cell(r, 6).value = zaehlformel(r, None, 'erfasst')
        KOMP.cell(r, 7).value = zaehlformel(r, None, 'bewertet')
        KOMP.cell(r, 8).value = zaehlformel(r, None, 'erfüllt')
        KOMP.cell(r, 9).value = zaehlformel(r, None, 'trifft nicht zu')
        KOMP.cell(r, 10).value = (
            f'=IF($D{r}="","",IF($E{r}=0,"keine Anforderungen",'
            f'IF($F{r}>0,"erfasst",IF($G{r}>0,"bewertet",'
            f'IF($H{r}>0,"erfüllt","trifft nicht zu")))))')
    KOMP.row_dimensions[r].height = 30
    for c in range(1, 12):
        z = KOMP.cell(r, c)
        z.font = Font(name=ARIAL, size=10, bold=(c == 10))
        z.fill = fill(C_HELL) if c <= 4 else fill('FFFFFFFF')
        z.border = RAHMEN
        z.alignment = Alignment(wrap_text=True, vertical='center',
                                horizontal='center' if 5 <= c <= 10 else 'left')

KOMP.conditional_formatting.add(
    f'E{K_START}:E{K_ENDE}',
    Rule(type='dataBar', dataBar=openpyxl.formatting.rule.DataBar(
        cfvo=[openpyxl.formatting.rule.FormatObject(type='min'),
              openpyxl.formatting.rule.FormatObject(type='max')],
        color='FF4472C4', showValue=True, minLength=None, maxLength=None)))
for wert, farbe, schrift, fett in [('erfasst', C_ERFASST, None, None),
                                   ('bewertet', C_BEWERTET, None, None),
                                   ('erfüllt', C_ERFUELLT, 'FF006100', True),
                                   ('trifft nicht zu', C_TNZ, 'FF666666', None),
                                   ('keine Anforderungen', C_KEINE, 'FF777777', None)]:
    KOMP.conditional_formatting.add(f'J{K_START}:J{K_ENDE}',
                                    dxf_regel(f'J{K_START}="{wert}"', farbe, schrift, fett))
KOMP.conditional_formatting.add(f'A{K_START}:A{K_ENDE}',
                                dxf_regel(f'A{K_START}="Zuordnung prüfen"', C_PRUEFEN,
                                          'FF7F6000', None))

# --------------------------------------------------------------------------
# 2b  Organigramm (formelbasiert, direkt aus der Tabelle oberhalb)
# --------------------------------------------------------------------------
O_TITEL = K_ENDE + 2
O_WURZEL = O_TITEL + 3
O_START = O_WURZEL + 2                       # erste Komponentenzeile
VERSATZ = O_START - K_START                  # Organigrammzeile -> Tabellenzeile

KOMP.cell(O_TITEL, 1).value = 'Organigramm – automatische Struktur- und Statusansicht'
KOMP.cell(O_TITEL, 1).font = Font(name=ARIAL, size=14, bold=True, color=C_TITEL)
KOMP.cell(O_TITEL + 1, 1).value = (
    'SmartInfra → Vergabeeinheit → Partner → Art → Komponente. Jede Box rechnet direkt aus der '
    'Tabelle oberhalb; die Farbe folgt dem Komponentenstatus. Hinweis: Die Tabelle nach '
    'Vergabeeinheit, Partner und Art sortieren, damit die Gruppierung sauber zusammenfällt.')
KOMP.cell(O_TITEL + 1, 1).font = Font(name=ARIAL, size=9, color='FF808080')
KOMP.merge_cells(start_row=O_TITEL + 1, start_column=1, end_row=O_TITEL + 1, end_column=11)
KOMP.cell(O_TITEL + 1, 1).alignment = Alignment(wrap_text=True, vertical='center')
KOMP.row_dimensions[O_TITEL].height = 24
KOMP.row_dimensions[O_TITEL + 1].height = 28

# Wurzelknoten
KOMP.merge_cells(start_row=O_WURZEL, start_column=1, end_row=O_WURZEL, end_column=5)
wurzel = KOMP.cell(O_WURZEL, 1)
wurzel.value = (f'="SMARTINFRA"&CHAR(10)&COUNTA($D${K_START}:$D${K_ENDE})&" Komponenten / '
                f'Schnittstellen  ·  "&SUM($E${K_START}:$E${K_ENDE})&" Anforderungszuordnungen"')
wurzel.font = Font(name=ARIAL, size=12, bold=True, color='FFFFFFFF')
wurzel.fill = fill(C_TITEL)
wurzel.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
wurzel.border = RAHMEN
KOMP.row_dimensions[O_WURZEL].height = 40
KOMP.cell(O_WURZEL + 1, 1).value = '│'
KOMP.cell(O_WURZEL + 1, 1).font = Font(name=ARIAL, size=10, color='FF808080')
KOMP.row_dimensions[O_WURZEL + 1].height = 12

for i in range(len(KOMPONENTEN) + RESERVE):
    o = O_START + i
    t = K_START + i                          # zugehoerige Tabellenzeile
    v = t - 1                                # Vorgaengerzeile in der Tabelle
    KOMP.cell(o, 1).value = (
        f'=IF($D{t}="","",IF($A{t}=$A{v},"",$A{t}&CHAR(10)&'
        f'COUNTIF($A${K_START}:$A${K_ENDE},$A{t})&" Komponenten"))')
    KOMP.cell(o, 2).value = (
        f'=IF($D{t}="","",IF(AND($A{t}=$A{v},$B{t}=$B{v}),"","└ "&$B{t}&"  ("&'
        f'COUNTIFS($A${K_START}:$A${K_ENDE},$A{t},$B${K_START}:$B${K_ENDE},$B{t})&")"))')
    KOMP.cell(o, 3).value = (
        f'=IF($D{t}="","",IF(AND($A{t}=$A{v},$B{t}=$B{v},$C{t}=$C{v}),"","└ "&$C{t}&"  ("&'
        f'COUNTIFS($A${K_START}:$A${K_ENDE},$A{t},$B${K_START}:$B${K_ENDE},$B{t},'
        f'$C${K_START}:$C${K_ENDE},$C{t})&")"))')
    KOMP.merge_cells(start_row=o, start_column=4, end_row=o, end_column=5)
    KOMP.cell(o, 4).value = (
        f'=IF($D{t}="","","└ "&$D{t}&CHAR(10)&$E{t}&" Anforderungen  ·  "&$H{t}&"/"&$E{t}&'
        f'" erfüllt"&CHAR(10)&"Status: "&$J{t})')
    for c in range(1, 6):
        z = KOMP.cell(o, c)
        z.border = RAHMEN
        z.alignment = Alignment(wrap_text=True, vertical='center',
                                horizontal='center' if c <= 3 else 'left')
        z.font = Font(name=ARIAL, size=9,
                      bold=(c == 1), color='FFFFFFFF' if c == 1 else 'FF000000')
    KOMP.row_dimensions[o].height = 46
O_ENDE = O_START + len(KOMPONENTEN) + RESERVE - 1

KOMP.conditional_formatting.add(f'A{O_START}:A{O_ENDE}',
                                dxf_regel(f'$A{O_START}<>""', C_VE, 'FFFFFFFF', True))
KOMP.conditional_formatting.add(f'B{O_START}:B{O_ENDE}',
                                dxf_regel(f'$B{O_START}<>""', C_PARTNER, 'FFFFFFFF', True))
KOMP.conditional_formatting.add(f'C{O_START}:C{O_ENDE}',
                                dxf_regel(f'$C{O_START}<>""', C_ART, 'FF1F3864', True))
for wert, farbe, schrift, fett in [('erfasst', C_ERFASST, 'FF404040', None),
                                   ('bewertet', C_BEWERTET, 'FF1F3864', None),
                                   ('erfüllt', C_ERFUELLT, 'FF006100', True),
                                   ('trifft nicht zu', C_TNZ, 'FF666666', None),
                                   ('keine Anforderungen', C_KEINE, 'FF777777', None)]:
    KOMP.conditional_formatting.add(
        f'D{O_START}:E{O_ENDE}', dxf_regel(f'$J{K_START}="{wert}"', farbe, schrift, fett))

# Legende
L0 = O_ENDE + 2
KOMP.cell(L0, 1).value = 'Legende Statusfarben'
KOMP.cell(L0, 1).font = Font(name=ARIAL, size=10, bold=True, color=C_TITEL)
for i, (text, farbe) in enumerate([('erfasst', C_ERFASST), ('bewertet', C_BEWERTET),
                                   ('erfüllt', C_ERFUELLT), ('trifft nicht zu', C_TNZ),
                                   ('keine Anforderungen', C_KEINE)]):
    z = KOMP.cell(L0 + 1, i + 1)
    z.value = text
    z.fill = fill(farbe)
    z.border = RAHMEN
    z.font = Font(name=ARIAL, size=9)
    z.alignment = Alignment(horizontal='center', vertical='center')
KOMP.row_dimensions[L0 + 1].height = 20
KOMP.cell(L0 + 3, 1).value = (
    'Pflegehinweis: Es werden ausschließlich Vergabeeinheit, Partner, Art, Komponente und '
    'Bemerkung eingetragen. Anzahl, Statusverteilung, Gesamtstatus und das gesamte Organigramm '
    'aktualisieren sich daraus automatisch. „Zuordnung prüfen“ bedeutet: aus den vorliegenden '
    'Unterlagen nicht eindeutig ableitbar.')
KOMP.cell(L0 + 3, 1).font = Font(name=ARIAL, size=9, color='FF808080')
KOMP.merge_cells(start_row=L0 + 3, start_column=1, end_row=L0 + 3, end_column=11)
KOMP.cell(L0 + 3, 1).alignment = Alignment(wrap_text=True, vertical='top')
KOMP.row_dimensions[L0 + 3].height = 30
KOMP.sheet_view.showGridLines = False
KOMP.freeze_panes = 'A6'

print(f'Komponenten: Tabelle {K_START}-{K_ENDE}, Organigramm {O_START}-{O_ENDE}')

# ==========================================================================
# 3  Register "Auswertung" neu aufbauen (gleiche Optik, korrigierte Formeln)
# ==========================================================================
AUS = wb['Auswertung']
kopf_stil = copy.copy(AUS['A1']._style)
unter_stil = copy.copy(AUS['A2']._style)
sektion_stil = copy.copy(AUS['A5']._style)
spalte_stil = copy.copy(AUS['A6']._style)
label_stil = copy.copy(AUS['A7']._style)
zahl_stil = copy.copy(AUS['B7']._style)

for rng in list(AUS.merged_cells.ranges):
    AUS.unmerge_cells(str(rng))
AUS.conditional_formatting = ConditionalFormattingList()
for r in range(1, AUS.max_row + 2):
    for c in range(1, 8):
        AUS.cell(r, c).value = None
    if r in AUS.row_dimensions:
        AUS.row_dimensions[r].height = None

AUS['A1'] = 'Auswertung – Anforderungen'
AUS['A1']._style = kopf_stil
AUS['A2'] = ('Automatische Auswertung des Statusprozesses erfasst → bewertet → erfüllt bzw. '
             'trifft nicht zu – inklusive Bieterfragenkatalog.')
AUS['A2']._style = unter_stil
AUS['A3'] = ('Erfüllungsgrad = erfüllte Anforderungen / anwendbare Anforderungen '
             '(Status „trifft nicht zu“ wird aus dem Nenner ausgeschlossen).')
AUS['A3']._style = unter_stil
for r in (1, 2, 3):
    AUS.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)

ANF_B = f'Anforderungen!$B$7:$B${MAXZ}'
ANF_D = f'Anforderungen!$D$7:$D${MAXZ}'
ANF_H = f'Anforderungen!$H$7:$H${MAXZ}'

THEMEN = sorted({ws.cell(r, 4).value for r in range(7, letzte_zeile + 1)
                 if ws.cell(r, 4).value})
PARTNER = ['Aplitronic', 'eRound', 'SCS', 'Reservierungsplattform', 'tbd', 'offen',
           'Zuordnung prüfen']
KOMPLISTE = [k[2] for k in KOMPONENTEN if k[1] == 'Komponente'] + ['Zuordnung prüfen']
DOKUMENTE = sorted({ws.cell(r, 2).value for r in range(7, letzte_zeile + 1)
                    if ws.cell(r, 2).value})


def sektion(zeile, titel, eintraege, modus):
    """Schreibt einen Auswertungsblock; modus 'exakt' (COUNTIF) oder 'teil' (SUMPRODUCT)."""
    AUS.cell(zeile, 1).value = titel
    for c in range(1, 8):
        AUS.cell(zeile, c)._style = copy.copy(sektion_stil)
    AUS.merge_cells(start_row=zeile, start_column=1, end_row=zeile, end_column=7)
    kopfzeile = zeile + 1
    for c, t in enumerate(['Merkmal', 'Gesamt', 'Erfasst', 'Bewertet', 'Erfüllt',
                           'Trifft nicht zu', 'Erfüllungsgrad'], start=1):
        AUS.cell(kopfzeile, c).value = t
        AUS.cell(kopfzeile, c)._style = copy.copy(spalte_stil)
    for i, name in enumerate(eintraege):
        r = kopfzeile + 1 + i
        AUS.cell(r, 1).value = name
        AUS.cell(r, 1)._style = copy.copy(label_stil)
        if modus == 'exakt':
            bereich = {'Themenfeld': ANF_D, 'Dokument': ANF_B}[titel.split()[-1]]
            AUS.cell(r, 2).value = f'=COUNTIF({bereich},$A{r})'
            for k, status in enumerate(['erfasst', 'bewertet', 'erfüllt', 'trifft nicht zu']):
                AUS.cell(r, 3 + k).value = f'=COUNTIFS({bereich},$A{r},{ANF_J},"{status}")'
        else:
            bereich = ANF_H if 'Partner' in titel else ANF_I
            AUS.cell(r, 2).value = (f'=SUMPRODUCT(ISNUMBER(SEARCH($A{r},{bereich}))'
                                    f'*({ANF_A}<>""))')
            for k, status in enumerate(['erfasst', 'bewertet', 'erfüllt', 'trifft nicht zu']):
                AUS.cell(r, 3 + k).value = (f'=SUMPRODUCT(ISNUMBER(SEARCH($A{r},{bereich}))'
                                            f'*({ANF_J}="{status}"))')
        AUS.cell(r, 7).value = f'=IFERROR(E{r}/(B{r}-F{r}),0)'
        for c in range(2, 8):
            AUS.cell(r, c)._style = copy.copy(zahl_stil)
        AUS.cell(r, 7).number_format = '0%'
    ende = kopfzeile + len(eintraege)
    AUS.conditional_formatting.add(
        f'B{kopfzeile + 1}:B{ende}',
        Rule(type='dataBar', dataBar=openpyxl.formatting.rule.DataBar(
            cfvo=[openpyxl.formatting.rule.FormatObject(type='min'),
                  openpyxl.formatting.rule.FormatObject(type='max')],
            color='FF4472C4', showValue=True, minLength=None, maxLength=None)))
    return ende + 3


z = 5
z = sektion(z, '1  Nach Themenfeld', THEMEN, 'exakt')
z = sektion(z, '2  Nach Partner', PARTNER, 'teil')
z = sektion(z, '3  Nach Komponente', KOMPLISTE, 'teil')
z = sektion(z, '4  Nach Dokument', DOKUMENTE, 'exakt')

# 5 Gesamtstatus
AUS.cell(z, 1).value = '5  Gesamtstatus'
for c in range(1, 8):
    AUS.cell(z, c)._style = copy.copy(sektion_stil)
AUS.merge_cells(start_row=z, start_column=1, end_row=z, end_column=7)
for c, t in enumerate(['Status', 'Anforderungen', 'Anteil'], start=1):
    AUS.cell(z + 1, c).value = t
    AUS.cell(z + 1, c)._style = copy.copy(spalte_stil)
for i, status in enumerate(['erfasst', 'bewertet', 'erfüllt', 'trifft nicht zu']):
    r = z + 2 + i
    AUS.cell(r, 1).value = status
    AUS.cell(r, 1)._style = copy.copy(label_stil)
    AUS.cell(r, 2).value = f'=COUNTIF({ANF_J},A{r})'
    AUS.cell(r, 3).value = f'=IFERROR(B{r}/COUNTA({ANF_A}),0)'
    for c in (2, 3):
        AUS.cell(r, c)._style = copy.copy(zahl_stil)
    AUS.cell(r, 3).number_format = '0%'
z = z + 7

# 6 Vollstaendigkeit
AUS.cell(z, 1).value = '6  Vollständigkeit Erfüllung / Nachweis'
for c in range(1, 8):
    AUS.cell(z, c)._style = copy.copy(sektion_stil)
AUS.merge_cells(start_row=z, start_column=1, end_row=z, end_column=7)
for c, t in enumerate(['Prüfung', 'Anforderungen', 'Hinweis'], start=1):
    AUS.cell(z + 1, c).value = t
    AUS.cell(z + 1, c)._style = copy.copy(spalte_stil)
ANF_K = f'Anforderungen!$K$7:$K${MAXZ}'
ANF_L = f'Anforderungen!$L$7:$L${MAXZ}'
pruefungen = [
    ('Erfüllt + Datum + Nachweis',
     f'=COUNTIFS({ANF_J},"erfüllt",{ANF_K},"<>",{ANF_L},"<>")', 'vollständig dokumentiert'),
    ('Erfüllt ohne Datum', f'=COUNTIFS({ANF_J},"erfüllt",{ANF_K},"")',
     'Datum in „Erfüllt am“ ergänzen'),
    ('Erfüllt ohne Nachweis / Notiz', f'=COUNTIFS({ANF_J},"erfüllt",{ANF_L},"")',
     'Nachweis oder kurze Notiz ergänzen'),
    ('Trifft nicht zu ohne Begründung', f'=COUNTIFS({ANF_J},"trifft nicht zu",{ANF_L},"")',
     'Begründung in „Nachweis / Notiz“ ergänzen'),
    ('Anwendbare Anforderungen',
     f'=COUNTA({ANF_A})-COUNTIF({ANF_J},"trifft nicht zu")', 'Gesamt abzüglich „trifft nicht zu“'),
    ('Offene Zuordnung Komponente',
     f'=COUNTIF({ANF_I},"Zuordnung prüfen")', 'Komponente fachlich zuordnen'),
    ('Offene Zuordnung Partner',
     f'=COUNTIF({ANF_H},"Zuordnung prüfen")', 'Partner fachlich zuordnen'),
]
for i, (name, formel, hinweis) in enumerate(pruefungen):
    r = z + 2 + i
    AUS.cell(r, 1).value = name
    AUS.cell(r, 1)._style = copy.copy(label_stil)
    AUS.cell(r, 2).value = formel
    AUS.cell(r, 2)._style = copy.copy(zahl_stil)
    AUS.cell(r, 3).value = hinweis
    AUS.cell(r, 3)._style = copy.copy(label_stil)
for c, b in zip('ABCDEFG', [34, 14, 12, 12, 12, 15, 15]):
    AUS.column_dimensions[c].width = b
AUS.sheet_view.showGridLines = False

# ==========================================================================
# 4  Cockpit: Bezuege nachziehen
# ==========================================================================
CP = wb['Cockpit']
for r in range(1, CP.max_row + 1):
    for c in range(1, 3):
        v = CP.cell(r, c).value
        if isinstance(v, str) and v.startswith('='):
            CP.cell(r, c).value = v.replace('$280', f'${MAXZ}')
CP['B17'] = f'=COUNTIF(Komponenten!$J${K_START}:$J${K_ENDE},"erfüllt")'
CP['A18'] = 'Bieterfragen im Register'
CP['B18'] = f'=COUNTIF({ANF_B},"Bieterfragenkatalog")'
CP['A19'] = 'Offene Zuordnungen (Komponente/Partner)'
CP['B19'] = (f'=COUNTIF({ANF_I},"Zuordnung prüfen")+COUNTIF({ANF_H},"Zuordnung prüfen")')
for r in (18, 19):
    CP.cell(r, 1)._style = copy.copy(CP['A17']._style)
    CP.cell(r, 2)._style = copy.copy(CP['B17']._style)
    CP.cell(r, 2).number_format = '0'
CP['A3'] = ('Alle Kennzahlen rechnen direkt aus den Registern „Anforderungen“ und '
            '„Komponenten“ – inklusive des vollständigen Bieterfragenkatalogs.')

# ==========================================================================
# 5  Lesehilfe ergaenzen
# ==========================================================================
LH = wb['Lesehilfe']
stil_a = copy.copy(LH['A20']._style)
stil_b = copy.copy(LH['B20']._style)
neu = [
    ('Bieterfragenkatalog', 'Der vollständige Bieterfragenkatalog (Stand 13.05.2026) ist als '
     'Dokument „Bieterfragenkatalog“ direkt im Register „Anforderungen“ enthalten – eine Frage '
     'je Zeile. Die Fundstelle nennt Fragennummer und Seite im PDF.'),
    ('Kernaussage', 'Kurze fachliche Zusammenfassung aus Frage und Antwort des Auftraggebers. '
     'Sie gibt die Antwort nicht verschärft wieder; eine reine Klarstellung bleibt eine '
     'Klarstellung. Die Originalfrage steht in Spalte F, die Originalantwort in Spalte N.'),
    ('Bieterfrage-Nr. / Antwort', 'Zwei neue Spalten am Ende der Tabelle: die Nummer der '
     'Bieterfrage und die vollständige Originalantwort des Auftraggebers.'),
    ('Zuordnung prüfen', 'Kennzeichnet Partner- oder Komponentenzuordnungen, die sich aus den '
     'vorliegenden Unterlagen nicht eindeutig ergeben. Bewusst offen gelassen statt geraten – '
     'diese Felder sind gelb markiert.'),
    ('Vergabeeinheit', 'Die Vergabeeinheit ist in keiner der vorliegenden Quellen benannt und '
     'steht daher überall auf „Zuordnung prüfen“. Sobald sie im Register „Komponenten“ '
     'eingetragen wird, gruppiert sich das Organigramm automatisch neu.'),
    ('Organigramm', 'Im Register „Komponenten“ unterhalb der Tabelle: SmartInfra → '
     'Vergabeeinheit → Partner → Art → Komponente. Vollständig formelbasiert, Farbe je '
     'Komponentenstatus. Tabelle nach Vergabeeinheit, Partner und Art sortieren.'),
]
r = 23
for a, b in neu:
    LH.cell(r, 1).value = a
    LH.cell(r, 1)._style = stil_a
    LH.cell(r, 2).value = b
    LH.cell(r, 2)._style = stil_b
    LH.row_dimensions[r].height = 42
    r += 1

wb.save(ZIEL)
print('gespeichert:', ZIEL)
