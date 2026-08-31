# -*- coding: utf-8 -*-
"""Trennt "gehoert nicht zu meinem Paket" von "ist keine Anforderung".

Bisher stand beides als Status "trifft nicht zu" nebeneinander. Dieses Skript

  * legt die Spalte N "Umfang" an (im Umfang / außerhalb),
  * setzt die allein wegen der Themenauswahl ausgeblendeten Zeilen auf "erfasst"
    zurueck und entfernt deren automatische Notiz,
  * laesst die sachlich geprueften "trifft nicht zu" unberuehrt.

Aufruf:  python3 add_scope_column.py <mappe.xlsx>
"""
import copy, re, sys
import openpyxl
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

SP = dict(id=2, kernaussage=6, text=7, bemerkung=8, komponente=10, status=11, nachweis=13,
          umfang=14)
KOMPONENTEN = ('Schrankenanlage', 'Schrankensteuerung', 'Kennzeichenerfassung',
               'Dynamische Anzeige')
STICHWORTE = re.compile(
    r'(schranke|poller|zufahrtsbeschr|kennzeichen|\bkamera|\banpr\b|nummernschild|'
    r'dynamisch\w* anzeige|\bled\b|anzeigetafel|\bdisplay)', re.I)
IM_UMFANG, AUSSERHALB = 'im Umfang', 'außerhalb'
THEMENNOTIZ = 'Außerhalb des betrachteten Leistungsumfangs'
GRENZE = 2000


def main():
    pfad = sys.argv[1]
    wb = openpyxl.load_workbook(pfad)
    ws = wb['Anforderungen']

    ws.cell(6, SP['umfang']).value = 'Umfang'
    ws.cell(6, SP['umfang'])._style = copy.copy(ws.cell(6, SP['nachweis'])._style)
    ws.column_dimensions['N'].width = 14
    stil = copy.copy(ws.cell(7, SP['id'])._style)

    im, ausser, zurueck, letzte = 0, 0, 0, 6
    for r in range(7, GRENZE + 1):
        if not ws.cell(r, SP['id']).value:
            continue
        letzte = r
        volltext = ' '.join(str(ws.cell(r, SP[k]).value or '')
                            for k in ('kernaussage', 'text', 'bemerkung'))
        drin = (any(k in str(ws.cell(r, SP['komponente']).value or '') for k in KOMPONENTEN)
                or bool(STICHWORTE.search(volltext)))
        ws.cell(r, SP['umfang']).value = IM_UMFANG if drin else AUSSERHALB
        ws.cell(r, SP['umfang'])._style = copy.copy(stil)
        im, ausser = im + drin, ausser + (not drin)
        notiz = str(ws.cell(r, SP['nachweis']).value or '')
        if THEMENNOTIZ in notiz:                  # war nur wegen der Themenauswahl gesetzt
            ws.cell(r, SP['status']).value = 'erfasst'
            ws.cell(r, SP['nachweis']).value = None
            zurueck += 1
    print(f'Spalte „Umfang": {im} im Umfang, {ausser} außerhalb')
    print(f'Status zurückgesetzt auf „erfasst": {zurueck}')

    alt = ws.tables['tblAnforderungen']
    kopf = [c.name for c in alt.tableColumns] + ['Umfang']
    stil_tab = alt.tableStyleInfo
    del ws.tables['tblAnforderungen']
    tab = Table(displayName='tblAnforderungen', ref=f'B6:N{letzte}',
                tableColumns=[TableColumn(id=i + 1, name=n) for i, n in enumerate(kopf)])
    if stil_tab is not None:
        tab.tableStyleInfo = TableStyleInfo(
            name=stil_tab.name, showFirstColumn=stil_tab.showFirstColumn,
            showLastColumn=stil_tab.showLastColumn, showRowStripes=stil_tab.showRowStripes,
            showColumnStripes=stil_tab.showColumnStripes)
    ws.add_table(tab)

    pruefung = DataValidation(type='list', formula1=f'"{IM_UMFANG},{AUSSERHALB}"',
                              allow_blank=False, showErrorMessage=True)
    ws.add_data_validation(pruefung)
    pruefung.add(f'N7:N{GRENZE}')
    ws.conditional_formatting.add(f'N7:N{GRENZE}', Rule(
        type='expression', formula=[f'$N7="{AUSSERHALB}"'],
        dxf=DifferentialStyle(font=Font(color='FF808080'),
                              fill=PatternFill(bgColor='FFF2F2F2'))))
    ws.conditional_formatting.add(f'N7:N{GRENZE}', Rule(
        type='expression', formula=[f'$N7="{IM_UMFANG}"'],
        dxf=DifferentialStyle(font=Font(color='FF1F4E78', b=True),
                              fill=PatternFill(bgColor='FFDDEBF7'))))

    cp = wb['Cockpit']
    zeile = max(r for r in range(1, cp.max_row + 2) if cp.cell(r, 1).value) + 1
    cp.cell(zeile, 1).value = 'Anforderungen im betrachteten Umfang'
    cp.cell(zeile, 1)._style = copy.copy(cp.cell(zeile - 1, 1)._style)
    cp.cell(zeile, 2).value = f'=COUNTIF(Anforderungen!$N$7:$N${GRENZE},"{IM_UMFANG}")'
    cp.cell(zeile, 2)._style = copy.copy(cp.cell(zeile - 1, 2)._style)
    cp.cell(zeile, 2).number_format = '0'

    lese = wb['Lesehilfe']
    zeile = max(r for r in range(1, lese.max_row + 2) if lese.cell(r, 1).value) + 1
    lese.cell(zeile, 1).value = 'Umfang'
    lese.cell(zeile, 1)._style = copy.copy(lese.cell(zeile - 1, 1)._style)
    lese.cell(zeile, 2).value = (
        'Trennt die Frage „gehört das zu unserem Paket?" von der Frage „ist das überhaupt eine '
        'Anforderung?". „im Umfang“ = betrifft Schranken, Kennzeichenerfassung oder dynamische '
        'Anzeige. Der Status bleibt davon unberührt und bildet weiterhin nur den '
        'Erfüllungsprozess ab.')
    lese.cell(zeile, 2)._style = copy.copy(lese.cell(zeile - 1, 2)._style)
    lese.row_dimensions[zeile].height = 42

    wb.save(pfad)
    print('gespeichert:', pfad)


if __name__ == '__main__':
    main()
