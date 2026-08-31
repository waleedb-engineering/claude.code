# -*- coding: utf-8 -*-
"""Nimmt automatisch gesetzte "trifft nicht zu" zurueck, die der Pruefung nicht standhalten.

Zurueckgenommen werden:
  * die Gruppe "nur Auftraggeber" vollstaendig - ein Recht des Auftraggebers begruendet
    regelmaessig eine Duldungs- oder Umsetzungspflicht des Auftragnehmers
    (z. B. § 25 Abs. 3: Anpassung der Mindestvoraussetzungen mit zwei Monaten Vorlauf)
  * die Gruppe "nicht Gegenstand" - die Antwort des Auftraggebers enthaelt entgegen der
    Kernaussage eine Anbindungspflicht
  * Unterlagenanpassungen, deren Antwort ueber die reine Ankuendigung hinausgeht

Aufruf:  python3 revert_tnz.py <mappe.xlsx>
"""
import json, re, sys
import openpyxl

SP = dict(id=2, fundstelle=4, kernaussage=6, status=11, nachweis=13)
INHALT = re.compile(r'(^|\s)(Ja,|Nein,|Verständnis|vgl\.|Zu i|Zu 1|Zu 2|Danach)')


KATALOG = {x['nr']: x for x in json.load(open('bieterfragen.json', encoding='utf-8'))}


def antwort_zu(fundstelle):
    """Volltext der Auftraggeber-Antwort zu einer Bieterfragen-Fundstelle."""
    treffer = re.match(r'Nr\. (\d+) /', fundstelle) or re.match(r'§ (\d+)$', fundstelle)
    return KATALOG.get(treffer.group(1), {}).get('antwort', '') if treffer else ''


def zuruecknehmen(nachweis, kernaussage, fundstelle):
    if 'richtet sich ausschließlich an den Auftraggeber' in nachweis:
        return 'nur Auftraggeber'
    if 'nicht Gegenstand' in nachweis:
        return 'nicht Gegenstand'
    if 'Anpassung der Vergabeunterlagen' in nachweis:
        antwort = antwort_zu(fundstelle)
        # Eine reine Ankuendigung ist kurz und enthaelt keine eigene Aussage
        if len(antwort) > 200 or INHALT.search(antwort):
            return 'Unterlagenanpassung mit Inhalt'
    return None


def main():
    pfad = sys.argv[1]
    wb = openpyxl.load_workbook(pfad)
    ws = wb['Anforderungen']
    gezaehlt = {}
    for r in range(7, 2100):
        if not ws.cell(r, SP['id']).value:
            continue
        nachweis = str(ws.cell(r, SP['nachweis']).value or '')
        if not nachweis.startswith('Automatisch ausgewertet:'):
            continue
        grund = zuruecknehmen(nachweis, str(ws.cell(r, SP['kernaussage']).value or ''),
                              str(ws.cell(r, SP['fundstelle']).value or ''))
        if not grund:
            continue
        ws.cell(r, SP['status']).value = 'erfasst'
        ws.cell(r, SP['nachweis']).value = None
        gezaehlt[grund] = gezaehlt.get(grund, 0) + 1
    print('zurückgenommen:', sum(gezaehlt.values()), gezaehlt)
    wb.save(pfad)
    print('gespeichert:', pfad)


if __name__ == '__main__':
    main()
