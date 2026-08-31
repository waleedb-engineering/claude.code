# -*- coding: utf-8 -*-
"""Legt den Reiter "Filter" an: Auswahlfelder und formelbasierte Trefferliste.

Ohne Makro und ohne Datenschnitt - die Liste rechnet mit INDEX/VERGLEICH und
funktioniert damit in jeder Excel-Fassung. Gepflegt wird weiterhin nur im
Register "Anforderungen"; dieser Reiter zeigt es nur gefiltert an.

Aufruf:  python3 build_filter_sheet.py <mappe.xlsx>
"""
import sys
import openpyxl
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

QUELLE, ERSTE, LETZTE = 'Anforderungen', 7, 2000
RANG = 'O'   # Hilfsspalte im Register
TREFFER = 300                                  # Anzahl anzeigbarer Zeilen
ALLE = '(alle)'

C_TITEL, C_KOPF, C_FELD = 'FF1F3864', 'FF1F4E78', 'FFFFF2CC'
STATUSFARBEN = [('erfasst', 'FFE7E6E6', None, None),
                ('bewertet', 'FFD9EAF7', None, None),
                ('erfüllt', 'FFE2F0D9', 'FF006100', True),
                ('trifft nicht zu', 'FFD9D9D9', 'FF666666', None),
                ('außerhalb Umfang', 'FFEDEDED', 'FF9C6500', None)]
RAHMEN = Border(*[Side(style='thin', color='FFBFBFBF')] * 4)

# Spalten im Register: (Buchstabe, Ueberschrift, Breite)
ANZEIGE = [('B', 'ID', 11), ('C', 'Dokument', 18), ('D', 'Fundstelle', 15),
           ('E', 'Themenfeld', 22), ('F', 'Kernaussage', 52), ('G', 'Originaltext', 60),
           ('I', 'Partner', 18), ('J', 'Komponente(n)', 30), ('K', 'Status', 15),
           ('L', 'Erfüllt am', 12), ('M', 'Nachweis / Notiz', 40)]
# Filterfelder: (Zelle, Beschriftung, Spalte im Register, Vergleichsart, Listenspalte)
GRUNDFELDER = [('Status', 'K', 'exakt', 'V'), ('Dokument', 'C', 'exakt', 'W'),
               ('Themenfeld', 'E', 'exakt', 'X'), ('Komponente', 'J', 'teil', 'Y'),
               ('Partner', 'I', 'teil', 'Z'), ('Suchbegriff', '', 'text', None)]
ZELLEN = ['B5', 'E5', 'H5', 'B7', 'E7', 'H7', 'B9']


def feldliste(quelle):
    """Mappen mit eigener Umfang-Spalte bekommen ein zusaetzliches Feld."""
    felder = list(GRUNDFELDER)
    if str(quelle['N6'].value or '') == 'Umfang':
        felder.insert(0, ('Umfang', 'N', 'exakt', 'U'))
    return [(ZELLEN[i], *f) for i, f in enumerate(felder)]


def werte(ws, spalte, zusatz=()):
    """Sortierte Liste der tatsaechlich vorkommenden Werte einer Registerspalte."""
    index = openpyxl.utils.column_index_from_string(spalte)
    gefunden = {str(ws.cell(r, index).value).strip()
                for r in range(ERSTE, LETZTE + 1) if ws.cell(r, index).value}
    return [ALLE] + sorted(gefunden | set(zusatz))


def regel(formel, fuellung, farbe=None, fett=None):
    stil = DifferentialStyle(font=Font(color=farbe, b=fett) if (farbe or fett) else None,
                             fill=PatternFill(bgColor=fuellung))
    return Rule(type='expression', formula=[formel], dxf=stil, stopIfTrue=False)


def main():
    pfad = sys.argv[1]
    wb = openpyxl.load_workbook(pfad)
    quelle = wb[QUELLE]
    if 'Filter' in wb.sheetnames:
        del wb['Filter']
    ws = wb.create_sheet('Filter', wb.sheetnames.index(QUELLE) + 1)
    ws.sheet_view.showGridLines = False

    ws['A1'] = 'Filter – Anforderungen gezielt anzeigen'
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color=C_TITEL)
    ws['A2'] = ('Auswahl in den gelben Feldern treffen; die Liste unten aktualisiert sich sofort. '
                '„(alle)“ lässt das Feld unberücksichtigt. Komponente und Partner suchen als '
                'Teiltext, finden also auch Mehrfachzuordnungen. Gepflegt wird weiterhin nur im '
                'Register „Anforderungen“.')
    ws['A2'].font = Font(name='Arial', size=9, color='FF808080')
    ws.merge_cells('A2:K2')
    ws['A2'].alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30

    # --- Auswahllisten in den Spalten V bis Z -------------------------------------
    # Komponente und Partner suchen als Teiltext - deshalb reichen die Grundbegriffe
    # aus dem Reiter "Komponenten"; Kombinationen werden davon mit erfasst.
    komp = wb['Komponenten']
    grundkomponenten = [str(komp.cell(r, 4).value) for r in range(6, 30)
                        if komp.cell(r, 4).value and komp.cell(r, 3).value == 'Komponente']
    grundpartner = [str(komp.cell(r, 14).value) for r in range(6, 30) if komp.cell(r, 14).value]
    FELDER = feldliste(quelle)
    listen = {'V': werte(quelle, 'K'), 'W': werte(quelle, 'C'), 'X': werte(quelle, 'E'),
              'Y': [ALLE] + sorted(set(grundkomponenten) | {'Zuordnung prüfen', '–'}),
              'Z': [ALLE] + sorted(set(grundpartner) | {'offen', 'Zuordnung prüfen'})}
    if any(f[1] == 'Umfang' for f in FELDER):
        listen['U'] = werte(quelle, 'N')
    ws['V1'] = 'Auswahllisten (Grundlage der Dropdowns)'
    ws['V1'].font = Font(name='Arial', size=9, bold=True, color='FF808080')
    for spalte, eintraege in listen.items():
        ws.column_dimensions[spalte].width = 30
        for i, wert in enumerate(eintraege):
            zelle = ws[f'{spalte}{i + 2}']
            zelle.value = wert
            zelle.font = Font(name='Arial', size=8, color='FF808080')

    # --- Filterfelder -------------------------------------------------------------
    for zelle, titel, _spalte, art, liste in FELDER:
        beschriftung = ws[f'{zelle[0]}{int(zelle[1:]) - 1}']
        beschriftung.value = titel
        beschriftung.font = Font(name='Arial', size=9, bold=True, color=C_KOPF)
        feld = ws[zelle]
        feld.value = '' if art == 'text' else ALLE
        feld.fill = PatternFill('solid', fgColor=C_FELD)
        feld.border = RAHMEN
        feld.font = Font(name='Arial', size=10)
        feld.alignment = Alignment(vertical='center')
        ws.merge_cells(f'{zelle}:{get_column_letter(feld.column + 1)}{feld.row}')
        if liste:
            pruefung = DataValidation(
                type='list', formula1=f'${liste}$2:${liste}${len(listen[liste]) + 1}',
                allow_blank=False, showErrorMessage=True, showInputMessage=True)
            pruefung.promptTitle, pruefung.prompt = titel, f'{titel} auswählen oder „{ALLE}“'
            ws.add_data_validation(pruefung)
            pruefung.add(zelle)
    ws['A10'] = ('Der Suchbegriff wirkt auf Kernaussage und Originaltext. „(alle)“ lässt ein '
                 'Feld unberücksichtigt; Komponente und Partner suchen als Teiltext.')
    ws['A10'].font = Font(name='Arial', size=8, italic=True, color='FF808080')
    ws.merge_cells('A10:K10')

    # --- Trefferzahl und Statusverteilung der Auswahl ------------------------------
    ws['A12'] = 'Treffer'
    ws['A12'].font = Font(name='Arial', size=9, bold=True, color=C_KOPF)
    ws['B12'] = f'=COUNTIF({QUELLE}!${RANG}${ERSTE}:${RANG}${LETZTE},">0")'
    ws['B12'].font = Font(name='Arial', size=14, bold=True, color=C_TITEL)
    vorhanden = [s for s, *_ in STATUSFARBEN if s in listen['V']]
    for i, status in enumerate(vorhanden):
        ws.cell(12, 4 + i * 2).value = status
        ws.cell(12, 4 + i * 2).font = Font(name='Arial', size=9, color='FF808080')
        ws.cell(12, 5 + i * 2).value = (
            f'=SUMPRODUCT(({QUELLE}!${RANG}${ERSTE}:${RANG}${LETZTE}>0)*'
            f'({QUELLE}!$K${ERSTE}:$K${LETZTE}="{status}"))')
        ws.cell(12, 5 + i * 2).font = Font(name='Arial', size=10, bold=True)
    ws['A13'] = (f'Angezeigt werden die ersten {TREFFER} Treffer. Bei mehr Treffern die Auswahl '
                 f'weiter eingrenzen.')
    ws['A13'].font = Font(name='Arial', size=8, italic=True, color='FF808080')
    ws.merge_cells('A13:K13')

    # --- Hilfsspalte T: laufende Nummer je passender Registerzeile ----------------
    bedingungen = []
    for zelle, _t, spalte, art, _l in FELDER:
        bezug = f'Filter!${zelle[0]}${zelle[1:]}'
        if art == 'exakt':
            bedingungen.append(f'OR({bezug}="{ALLE}",${spalte}{{z}}={bezug})')
        elif art == 'teil':
            bedingungen.append(f'OR({bezug}="{ALLE}",ISNUMBER(SEARCH({bezug},${spalte}{{z}})))')
        else:
            bedingungen.append(f'OR({bezug}="",ISNUMBER(SEARCH({bezug},$F{{z}}&" "&$G{{z}})))')
    muster = 'AND(' + ','.join(bedingungen) + ')'
    # Die Rangspalte liegt im Register, nicht hier: LibreOffice raeumt beim Speichern
    # Zeilen ohne sonstigen Inhalt ab und wuerde die Rangfolge sonst abschneiden.
    ws.column_dimensions['R'].hidden = True
    quelle.column_dimensions[RANG].hidden = True
    quelle[f'{RANG}{ERSTE - 1}'] = 'Filter-Rang'
    quelle[f'{RANG}{ERSTE - 1}'].font = Font(name='Arial', size=8, color='FF808080')
    for z in range(ERSTE, LETZTE + 1):
        quelle[f'{RANG}{z}'] = (
            f'=IF($B{z}="",0,IF({muster.format(z=z)},'
            f'COUNTIF(${RANG}${ERSTE - 1}:${RANG}{z - 1},">0")+1,0))')

    # --- Ergebnisliste ------------------------------------------------------------
    kopf = 15
    for i, (_q, titel, breite) in enumerate(ANZEIGE):
        zelle = ws.cell(kopf, i + 1)
        zelle.value = titel
        zelle.font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
        zelle.fill = PatternFill('solid', fgColor=C_KOPF)
        zelle.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        zelle.border = RAHMEN
        ws.column_dimensions[get_column_letter(i + 1)].width = breite
    ws.row_dimensions[kopf].height = 28
    status_spalte = get_column_letter(1 + [t for _q, t, _b in ANZEIGE].index('Status'))
    for n in range(TREFFER):
        zeile = kopf + 1 + n
        ws[f'R{zeile}'] = (f'=IFERROR(MATCH(ROWS($A${kopf + 1}:$A{zeile}),'
                           f'{QUELLE}!${RANG}${ERSTE}:${RANG}${LETZTE},0)+{ERSTE - 1},"")')
        for i, (spalte, _t, _b) in enumerate(ANZEIGE):
            zelle = ws.cell(zeile, i + 1)
            zelle.value = (f'=IF($R{zeile}="","",INDEX({QUELLE}!${spalte}:${spalte},$R{zeile}))')
            zelle.font = Font(name='Arial', size=9)
            zelle.alignment = Alignment(wrap_text=True, vertical='top')
            zelle.border = RAHMEN
            if spalte == 'L':
                zelle.number_format = 'DD.MM.YYYY'
        ws.row_dimensions[zeile].height = 30
    ws.conditional_formatting = ConditionalFormattingList()
    for status, fuellung, farbe, fett in STATUSFARBEN:
        ws.conditional_formatting.add(
            f'A{kopf + 1}:{get_column_letter(len(ANZEIGE))}{kopf + TREFFER}',
            regel(f'${status_spalte}{kopf + 1}="{status}"', fuellung, farbe, fett))
    ws.freeze_panes = f'A{kopf + 1}'

    lese = wb['Lesehilfe']
    zeile = max((r for r in range(1, lese.max_row + 2) if lese.cell(r, 1).value), default=20) + 1
    lese.cell(zeile, 1).value = 'Filter'
    lese.cell(zeile, 1)._style = lese.cell(zeile - 1, 1)._style
    lese.cell(zeile, 2).value = ('Eigener Reiter zum gezielten Anzeigen: Status, Dokument, '
                                 'Themenfeld, Komponente, Partner und Freitext. Die Trefferliste '
                                 'rechnet sich aus dem Register und wird dort nicht gepflegt.')
    lese.cell(zeile, 2)._style = lese.cell(zeile - 1, 2)._style
    lese.row_dimensions[zeile].height = 42

    # Der Reiter wird nach dem Neuberechnen erzeugt: LibreOffice kuerzt beim Speichern
    # Formelzellen weg, deren Ergebnis leer ist, und zerstoert damit die Rangspalte.
    # Excel rechnet die Mappe beim Oeffnen deshalb einmal vollstaendig durch.
    wb.calculation.fullCalcOnLoad = True
    wb.save(pfad)
    print(f'Reiter „Filter“ angelegt: {len(FELDER)} Filterfelder, {TREFFER} Trefferzeilen, '
          f'Listen {", ".join(f"{k}={len(v)}" for k, v in listen.items())}')


if __name__ == '__main__':
    main()
