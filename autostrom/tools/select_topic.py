# -*- coding: utf-8 -*-
"""Grenzt das Register auf den eigenen Leistungsumfang ein.

Behalten wird, was Schranken, Kennzeichenerfassung oder die dynamische Anzeige
betrifft. Alles Uebrige erhaelt den Status "trifft nicht zu" mit Begruendung.

Nie angefasst werden Zeilen, die bereits einen anderen Status oder eine eigene
Notiz tragen - die Arbeit des Nutzers bleibt unberuehrt.

Aufruf:  python3 select_topic.py <mappe.xlsx>
"""
import re, sys
import openpyxl

SP = dict(id=2, kernaussage=6, text=7, bemerkung=8, komponente=10, status=11, nachweis=13)
KOMPONENTEN = ('Schrankenanlage', 'Schrankensteuerung', 'Kennzeichenerfassung',
               'Dynamische Anzeige')
# Sicherheitsnetz: Zeilen, die das Thema im Text tragen, bleiben auch dann stehen,
# wenn ihre Komponentenzuordnung noch offen ist.
STICHWORTE = re.compile(
    r'(schranke|poller|zufahrtsbeschr|kennzeichen|\bkamera|\banpr\b|nummernschild|'
    r'dynamisch\w* anzeige|\bled\b|anzeigetafel|\bdisplay)', re.I)
BEGRUENDUNG = ('Automatisch ausgewertet: Außerhalb des betrachteten Leistungsumfangs '
               '(Schranken, Kennzeichenerfassung, dynamische Anzeige).')


def main():
    pfad = sys.argv[1]
    wb = openpyxl.load_workbook(pfad)
    ws = wb['Anforderungen']
    gesetzt = behalten = geschuetzt = 0
    for r in range(7, 2100):
        if not ws.cell(r, SP['id']).value:
            continue
        if ws.cell(r, SP['status']).value != 'erfasst' or ws.cell(r, SP['nachweis']).value:
            geschuetzt += 1
            continue
        komponente = str(ws.cell(r, SP['komponente']).value or '')
        volltext = ' '.join(str(ws.cell(r, SP[k]).value or '')
                            for k in ('kernaussage', 'text', 'bemerkung'))
        if any(k in komponente for k in KOMPONENTEN) or STICHWORTE.search(volltext):
            behalten += 1
            continue
        ws.cell(r, SP['status']).value = 'trifft nicht zu'
        ws.cell(r, SP['nachweis']).value = BEGRUENDUNG
        gesetzt += 1
    print(f'im Umfang behalten: {behalten} | neu auf „trifft nicht zu": {gesetzt} | '
          f'unberührt (eigener Status/Notiz): {geschuetzt}')
    wb.save(pfad)
    print('gespeichert:', pfad)


if __name__ == '__main__':
    main()
