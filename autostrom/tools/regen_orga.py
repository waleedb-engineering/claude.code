# -*- coding: utf-8 -*-
"""Erzeugt Textquellen und Bauplan des Organigramms aus der AKTUELLEN Mappe neu.

Anders als build.py baut dieses Skript die Mappe nicht neu auf, sondern liest die
gepflegte Komponententabelle und richtet Spalte M sowie den Bauplan danach aus.
Damit bleiben alle Eintraege des Nutzers erhalten.

Aufruf:  python3 regen_orga.py <mappe.xlsx>
Danach:  neu berechnen lassen, dann inject_shapes.py <mappe.xlsx>
"""
import json, sys
import openpyxl
from openpyxl.styles import Font, Alignment

K_START, K_ENDE = 6, 29                       # Datenzeilen der Komponententabelle
BOX_B, ABSTAND = 200, 16
RASTER = BOX_B + ABSTAND
Y_WURZEL, H_WURZEL = 0, 52
Y_VE, H_VE = 86, 48
Y_PARTNER, H_PARTNER = 166, 48
Y_ART, H_ART = 246, 32
Y_KOMP, H_KOMP, RASTER_KOMP = 302, 70, 80


def main():
    pfad = sys.argv[1]
    wb = openpyxl.load_workbook(pfad)
    ws = wb['Komponenten']

    # Hierarchie logisch gruppieren (unabhaengig von der Zeilenreihenfolge)
    gruppen, reihenfolge = {}, []
    for r in range(K_START, K_ENDE + 1):
        komponente = ws.cell(r, 4).value
        if not komponente:
            continue
        schluessel = (ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value)
        if schluessel not in gruppen:
            gruppen[schluessel] = []
            reihenfolge.append(schluessel)
        gruppen[schluessel].append(r)
    # Spalten nach erstem Auftreten der Vergabeeinheit, dann des Partners ordnen
    ve_rang, partner_rang = {}, {}
    for ve, partner, _art in reihenfolge:
        ve_rang.setdefault(ve, len(ve_rang))
        partner_rang.setdefault((ve, partner), len(partner_rang))
    reihenfolge.sort(key=lambda s: (ve_rang[s[0]], partner_rang[(s[0], s[1])]))

    spalte_von_gruppe = {s: i for i, s in enumerate(reihenfolge)}
    max_komp = max(len(z) for z in gruppen.values())
    print(f'{len(reihenfolge)} Gruppen, {sum(len(z) for z in gruppen.values())} Komponenten, '
          f'max {max_komp} je Gruppe')
    for s in reihenfolge:
        print(f"   {s[0]:18s} | {s[1]:24s} | {s[2]:14s} | {len(gruppen[s])}")

    zaehler = {
        've': lambda r: f'COUNTIF($A${K_START}:$A${K_ENDE},$A{r})',
        'partner': lambda r: (f'COUNTIFS($A${K_START}:$A${K_ENDE},$A{r},'
                              f'$B${K_START}:$B${K_ENDE},$B{r})'),
        'art': lambda r: (f'COUNTIFS($A${K_START}:$A${K_ENDE},$A{r},'
                          f'$B${K_START}:$B${K_ENDE},$B{r},$C${K_START}:$C${K_ENDE},$C{r})'),
    }
    spalte_feld = {'ve': 'A', 'partner': 'B', 'art': 'C'}
    einheit = {'ve': ('" Komponente"', '" Komponenten"'),
               'partner': ('" Komponente"', '" Komponenten"'),
               'art': ('" Eintrag"', '" Einträge"')}

    knoten, m_zeile = [], 34
    stil_font = Font(name='Arial', size=8, color='FF808080')

    def textquelle(formel, ebene, **rest):
        nonlocal m_zeile
        zelle = f'M{m_zeile}'
        ws[zelle] = formel
        ws[zelle].font = stil_font
        ws[zelle].alignment = Alignment(vertical='center')
        knoten.append(dict(zelle=zelle, ebene=ebene, **rest))
        m_zeile += 1

    def spannweite(pruef):
        idx = [spalte_von_gruppe[s] for s in reihenfolge if pruef(s)]
        return min(idx), max(idx)

    textquelle(f'="SMARTINFRA"&CHAR(10)&COUNTA($D${K_START}:$D${K_ENDE})&'
               f'" Komponenten / Schnittstellen"', 'wurzel',
               spalte_von=0, spalte_bis=len(reihenfolge) - 1)

    for ebene, schluessel_laenge in (('ve', 1), ('partner', 2), ('art', 3)):
        gesehen = []
        for s in reihenfolge:
            teil = s[:schluessel_laenge]
            if teil in gesehen:
                continue
            gesehen.append(teil)
            r = gruppen[s][0]
            von, bis = spannweite(lambda x, t=teil, n=schluessel_laenge: x[:n] == t)
            ez, mz = einheit[ebene]
            z = zaehler[ebene](r)
            textquelle(f'=${spalte_feld[ebene]}{r}&CHAR(10)&{z}&IF({z}=1,{ez},{mz})',
                       ebene, spalte_von=von, spalte_bis=bis)

    for s in reihenfolge:
        for platz, r in enumerate(gruppen[s]):
            textquelle(f'=IF($D{r}="","",$D{r}&CHAR(10)&$E{r}&IF($E{r}=1," Anforderung",'
                       f'" Anforderungen")&" · "&$H{r}&"/"&$E{r}&" erfüllt"&CHAR(10)&'
                       f'"Status: "&$J{r})',
                       'komponente', spalte_von=spalte_von_gruppe[s], spalte_bis=spalte_von_gruppe[s],
                       platz=platz, status_zelle=f'J{r}')

    for r in range(m_zeile, m_zeile + 12):    # alte, nicht mehr benoetigte Textquellen leeren
        ws.cell(r, 13).value = None

    start_zeile = 34
    dia_hoehe = Y_KOMP + max_komp * RASTER_KOMP + 12
    noetige_zeilen = -(-dia_hoehe // 24) + 1
    for r in range(start_zeile, start_zeile + noetige_zeilen):
        if ws.row_dimensions[r].height is None:
            ws.row_dimensions[r].height = 18
    for r in range(1, start_zeile):
        if ws.row_dimensions[r].height is None:
            ws.row_dimensions[r].height = 15

    plan = dict(blatt='Komponenten', start_zeile=start_zeile,
                ende_zeile=start_zeile + noetige_zeilen - 1,
                box_breite=BOX_B, raster=RASTER,
                dia_breite=len(reihenfolge) * RASTER, dia_hoehe=dia_hoehe,
                ebenen=dict(wurzel=[Y_WURZEL, H_WURZEL], ve=[Y_VE, H_VE],
                            partner=[Y_PARTNER, H_PARTNER], art=[Y_ART, H_ART],
                            komponente=[Y_KOMP, H_KOMP, RASTER_KOMP]),
                y_versatz_pt=sum(ws.row_dimensions[r].height for r in range(1, start_zeile)),
                knoten=knoten)
    json.dump(plan, open('organigramm.json', 'w'), ensure_ascii=False, indent=1)
    wb.save(pfad)
    print(f'{len(knoten)} Textquellen in {pfad} geschrieben (M34..M{m_zeile - 1})')


if __name__ == '__main__':
    main()
