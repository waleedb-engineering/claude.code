# -*- coding: utf-8 -*-
"""Ergaenzt die Anforderungstabelle um die Vertragsdokumente.

Arbeitet auf der gepflegten Mappe: bestehende Zeilen bleiben unberuehrt, bereits
erfasste Fundstellen werden nicht doppelt angelegt.

Aufruf:  python3 add_docs.py <mappe.xlsx>
"""
import copy, json, re, sys
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

import classify                                   # Themenfeld-, Komponenten- und Partnerlogik

SPALTE = dict(id=2, dokument=3, fundstelle=4, themenfeld=5, kernaussage=6, text=7,
              bemerkung=8, partner=9, komponente=10, status=11, datum=12, nachweis=13)
NEUE_GRENZE = 2000


def inhaltsschluessel(text):
    """Vergleichsschluessel fuer die Dublettenpruefung: Anfang des Originaltextes."""
    normiert = re.sub(r'[^a-zäöüß0-9]+', '', str(text or '').lower())
    return normiert[:60]


def sauber(v):
    return ILLEGAL_CHARACTERS_RE.sub(' ', v) if isinstance(v, str) else v


def kernaussage(text):
    """Kurze fachliche Kernaussage: der tragende erste Satz der Regelung."""
    saetze = classify.saetze(text)
    if not saetze:
        return classify.kuerzen(text, 200)
    kern = saetze[0]
    if len(kern) < 45 and len(saetze) > 1:
        kern += ' ' + saetze[1]
    return classify.kuerzen(kern, 200)


def main():
    pfad = sys.argv[1]
    daten = json.load(open('vertragstexte.json', encoding='utf-8'))
    wb = openpyxl.load_workbook(pfad)
    ws = wb['Anforderungen']

    vorhanden, belegte_fundstellen, letzte_zeile, hoechste_id = set(), set(), 6, 0
    for r in range(7, 2000):
        kennung = ws.cell(r, SPALTE['id']).value
        if not kennung:
            continue
        letzte_zeile = r
        belegte_fundstellen.add((ws.cell(r, SPALTE['dokument']).value,
                                 ws.cell(r, SPALTE['fundstelle']).value))
        vorhanden.add(inhaltsschluessel(ws.cell(r, SPALTE['text']).value))
        treffer = re.search(r'(\d+)$', str(kennung))
        if treffer:
            hoechste_id = max(hoechste_id, int(treffer.group(1)))
    print(f'Bestand: {letzte_zeile - 6} Anforderungen, letzte ID ANF-{hoechste_id:03d}')

    stil = {s: copy.copy(ws.cell(7, s)._style) for s in SPALTE.values()}
    hoehe = ws.row_dimensions[7].height or 48

    zeile, angelegt, uebersprungen, gezaehlt, kollisionen = letzte_zeile + 1, 0, 0, {}, 0
    for satz in daten:
        schluessel = inhaltsschluessel(satz['text'])
        if schluessel in vorhanden:            # inhaltlich bereits erfasst
            uebersprungen += 1
            continue
        vorhanden.add(schluessel)
        doppelte_fundstelle = (satz['dokument'], satz['fundstelle']) in belegte_fundstellen
        kollisionen += doppelte_fundstelle
        volltext = f"{satz['kapitel']} {satz['text']}"
        komponenten, unsicher = classify.finde_komponenten(volltext)
        thema = classify.finde_themenfeld(volltext)
        if komponenten:
            komponente = ' / '.join(komponenten)
            partner = classify.finde_partner(komponenten) or 'Zuordnung prüfen'
        elif unsicher:
            komponente, partner = 'Zuordnung prüfen', 'Zuordnung prüfen'
        elif thema in classify.NICHT_TECHNISCH:
            komponente, partner = '–', 'offen'
        else:
            komponente, partner = 'Zuordnung prüfen', 'Zuordnung prüfen'

        hoechste_id += 1
        bemerkung = f"Kapitel: {satz['kapitel']}\nFundstelle im PDF: Seite {satz['seite']}"
        if doppelte_fundstelle:
            bemerkung += ('\nFundstelle im Bestand bereits belegt, Text weicht ab – '
                          'bitte gegen die alte Zeile prüfen')
        if 'Zuordnung prüfen' in (komponente, partner):
            bemerkung += '\nZuordnung prüfen (Komponente/Partner nicht eindeutig)'
        werte = {'id': f'ANF-{hoechste_id:03d}', 'dokument': satz['dokument'],
                 'fundstelle': satz['fundstelle'], 'themenfeld': thema,
                 'kernaussage': kernaussage(satz['text']), 'text': satz['text'],
                 'bemerkung': bemerkung, 'partner': partner, 'komponente': komponente,
                 'status': 'erfasst', 'datum': None, 'nachweis': None}
        for feld, spalte in SPALTE.items():
            zelle = ws.cell(zeile, spalte)
            zelle.value = sauber(werte[feld])
            zelle._style = copy.copy(stil[spalte])
        ws.row_dimensions[zeile].height = hoehe
        gezaehlt[satz['dokument']] = gezaehlt.get(satz['dokument'], 0) + 1
        zeile += 1
        angelegt += 1
    letzte_zeile = zeile - 1
    print(f'neu angelegt: {angelegt}  ({gezaehlt}), inhaltsgleich übersprungen: {uebersprungen}')
    print(f'davon mit bereits belegter Fundstelle (abweichender Text): {kollisionen}')
    print(f'Tabelle reicht jetzt bis Zeile {letzte_zeile}')

    # Auswertungsbereiche mitziehen
    for blatt in wb.worksheets:
        for row in blatt.iter_rows():
            for zelle in row:
                if isinstance(zelle.value, str) and zelle.value.startswith('='):
                    zelle.value = zelle.value.replace('$800', f'${NEUE_GRENZE}')
    for dv in ws.data_validations.dataValidation:
        neu = [str(b).replace('800', str(NEUE_GRENZE)) for b in dv.sqref.ranges]
        dv.sqref = ' '.join(neu)
    regeln = list(ws.conditional_formatting._cf_rules.items())
    from openpyxl.formatting.formatting import ConditionalFormattingList
    ws.conditional_formatting = ConditionalFormattingList()
    for bereich, liste in regeln:
        for regel in liste:
            ws.conditional_formatting.add(
                str(bereich.sqref).replace('800', str(NEUE_GRENZE)), regel)

    alt = ws.tables['tblAnforderungen']
    kopf = [c.name for c in alt.tableColumns]
    stil_tab = alt.tableStyleInfo
    del ws.tables['tblAnforderungen']
    tab = Table(displayName='tblAnforderungen', ref=f'B6:M{letzte_zeile}',
                tableColumns=[TableColumn(id=i + 1, name=n) for i, n in enumerate(kopf)])
    if stil_tab is not None:                      # Tabellenformat unveraendert uebernehmen
        tab.tableStyleInfo = TableStyleInfo(
            name=stil_tab.name, showFirstColumn=stil_tab.showFirstColumn,
            showLastColumn=stil_tab.showLastColumn, showRowStripes=stil_tab.showRowStripes,
            showColumnStripes=stil_tab.showColumnStripes)
    ws.add_table(tab)

    aus = wb['Auswertung']                        # Dokumentbezeichnung angleichen
    for r in range(50, 65):
        if aus.cell(r, 1).value == 'Anlage 13':
            aus.cell(r, 1).value = 'Vertragsanlage 13'
            print('Auswertung: „Anlage 13“ auf „Vertragsanlage 13“ berichtigt')

    wb.save(pfad)
    print('gespeichert:', pfad)


if __name__ == '__main__':
    main()
