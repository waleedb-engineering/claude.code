#!/usr/bin/env python3
"""Baut aus der Anforderungsliste "Autostrom / LKW-Laden" ein Arbeitscockpit.

Quelle : 20251208_Autostrom_LKWLaden_Anforderung.xlsx (Tabelle1 + Tabelle2)
Ziel   : Cockpit, Anforderungsregister, Auswertung, Komponenten,
         Anzeigekonzept, Lesehilfe

Es werden keine Anforderungen erfunden: Texte, Partner und Komponenten
stammen unveraendert aus der Quelldatei. Ergaenzt werden Ordnungsmerkmale
(ID, Fundstelle, Themenfeld) und leere Steuerungsspalten fuer die Pflege.
"""

import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.data_source import StrRef
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.table import Table, TableStyleInfo

FONT = "Arial"
C_HEAD = "1F3864"
C_SECTION = "D9E2F3"
C_INPUT = "FFF9E1"
C_CALC = "F2F2F2"
C_ACCENT = "E2EFDA"
C_WARN = "FCE4D6"
C_BORDER = "BFBFBF"

thin = Side(style="thin", color=C_BORDER)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
FILL_HEAD = PatternFill("solid", fgColor=C_HEAD)
FILL_SECTION = PatternFill("solid", fgColor=C_SECTION)
FILL_INPUT = PatternFill("solid", fgColor=C_INPUT)
FILL_CALC = PatternFill("solid", fgColor=C_CALC)
FILL_ACCENT = PatternFill("solid", fgColor=C_ACCENT)

# Themenfeld je Quellzeile – aus dem Inhalt der Anforderung abgeleitet
THEMA = {
    **{r: "Reservierung" for r in (2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 27, 29, 30,
                                   33, 34, 35, 36)},
    **{r: "Nutzungszeit & Entgelte" for r in (14, 15, 16, 17, 18, 19, 20, 21,
                                              22, 23, 24, 31, 32)},
    **{r: "Zufahrt & Schranke" for r in (11, 39, 40, 45, 46, 47, 48, 50, 51)},
    **{r: "Authentifizierung & Zugang" for r in (13, 42)},
    **{r: "Belegungsdetektion" for r in (28, 43, 44)},
    **{r: "Anzeige & Beschilderung" for r in (41, 49)},
    **{r: "Monitoring & Reporting" for r in (37, 38)},
    **{r: "Betrieb & Auslastung" for r in (25, 26)},
    **{r: "Bau & Dokumentation" for r in (52, 53)},
}
THEMEN = ["Reservierung", "Nutzungszeit & Entgelte", "Zufahrt & Schranke",
          "Belegungsdetektion", "Authentifizierung & Zugang",
          "Anzeige & Beschilderung", "Monitoring & Reporting",
          "Betrieb & Auslastung", "Bau & Dokumentation"]

STATUS = ["offen", "in Klärung", "in Umsetzung", "umgesetzt", "entfällt"]
PRIO = ["hoch", "mittel", "niedrig"]

COLS = [
    ("ID", 9), ("Dokument", 19), ("Fundstelle", 16), ("Themenfeld", 24),
    ("Kernaussage", 46), ("Anforderung (Originaltext)", 78),
    ("Bemerkung / Klärung", 34), ("Partner", 16), ("Komponente(n)", 30),
    ("Klärungsbedarf", 13), ("Status", 14), ("Priorität", 11),
    ("Verantwortlich", 18), ("Fällig bis", 13), ("Nachweis / Umsetzung", 34),
]
C_ID, C_DOK, C_FUND, C_THEMA, C_KURZ, C_TEXT, C_BEM, C_PARTNER, C_KOMP, \
    C_KLAER, C_STATUS, C_PRIO, C_VERANT, C_FAELL, C_NACHW = range(1, 16)

R_HEAD = 6
R_FIRST = 7


def f(size=10, bold=False, color="000000", italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)


def clean(text):
    """Weiche Trennzeichen aus dem PDF-Import entfernen, Leerraum normalisieren."""
    if text is None:
        return ""
    s = str(text).replace("_x0002_", "").replace("\x02", "")
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def first_line(text):
    return clean(text).split("\n")[0].strip()


# In der Quelldatei ist dieselbe Komponente unterschiedlich geschrieben.
# Ohne Vereinheitlichung zählt sie doppelt.
NORMALISIERUNG = {
    "authentifikationskomponente": "Authentifizierungskomponente",
    "dynamische anzeige": "Dynamische Anzeige",
}


def joined(text):
    parts = []
    for p in clean(text).split("\n"):
        p = p.strip()
        if not p:
            continue
        parts.append(NORMALISIERUNG.get(p.lower(), p))
    return " / ".join(dict.fromkeys(parts))


# ---------------------------------------------------------------------------
def read_source(path):
    ws = load_workbook(path)["Tabelle1"]
    rows = []
    for r in range(2, 54):
        text = clean(ws.cell(row=r, column=6).value)
        if not text:
            continue
        par = clean(ws.cell(row=r, column=3).value)
        abs_ = clean(ws.cell(row=r, column=4).value)
        unt = clean(ws.cell(row=r, column=5).value)
        fund = f"§ {par}" if par else ""
        if abs_:
            fund += f" ({abs_})"
        if unt:
            fund += f" {unt}"
        rows.append({
            "dok": clean(ws.cell(row=r, column=2).value).replace("BIeterfrage",
                                                                "Bieterfrage"),
            "fund": fund.strip(),
            "thema": THEMA.get(r, "Sonstiges"),
            "kurz": clean(ws.cell(row=r, column=10).value),
            "text": text,
            "bem": clean(ws.cell(row=r, column=7).value),
            "partner": joined(ws.cell(row=r, column=8).value),
            "komp": joined(ws.cell(row=r, column=9).value),
        })

    komponenten = []
    for r in range(57, 74):
        typ = clean(ws.cell(row=r, column=7).value)
        if not typ:
            continue
        komponenten.append({
            "typ": typ,
            "partner": joined(ws.cell(row=r, column=8).value),
            "komp": joined(ws.cell(row=r, column=9).value),
        })
    return rows, komponenten


# ---------------------------------------------------------------------------
def header_block(ws, title, subtitle, hint):
    ws.cell(row=1, column=1, value=title).font = f(16, True, C_HEAD)
    ws.cell(row=2, column=1, value=subtitle).font = f(10, color="595959")
    if hint:
        ws.cell(row=3, column=1, value=hint).font = f(9, italic=True,
                                                     color="808080")
    ws.sheet_view.showGridLines = False


def sheet_register(wb, rows):
    ws = wb.create_sheet("Anforderungen")
    header_block(
        ws, "Anforderungsregister – Autostrom / LKW-Laden",
        "Quelle: Betreibervertrag, Vertragsanlagen, Bieterfragen, Anlage 13 "
        "(Stand der Auswertung: 08.12.2025)",
        "Weiß = Inhalt aus der Ausschreibung (nicht ändern) · "
        "Gelb = eigene Steuerung · Grau = automatisch berechnet")

    for i, (name, width) in enumerate(COLS, start=1):
        c = ws.cell(row=R_HEAD, column=i, value=name)
        c.font = f(10, True, "FFFFFF")
        c.fill = FILL_HEAD
        c.border = BORDER
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[R_HEAD].height = 32

    for i, row in enumerate(rows):
        r = R_FIRST + i
        values = {
            C_ID: f"ANF-{i + 1:03d}",
            C_DOK: row["dok"],
            C_FUND: row["fund"],
            C_THEMA: row["thema"],
            C_KURZ: row["kurz"],
            C_TEXT: row["text"],
            C_BEM: row["bem"],
            C_PARTNER: row["partner"] or "offen",
            C_KOMP: row["komp"] or "–",
            C_KLAER: (f'=IF(OR($H{r}="offen",ISNUMBER(SEARCH("tbd",$H{r})),'
                      f'ISNUMBER(SEARCH("?",$G{r}))),"ja","nein")'),
        }
        for col in range(1, len(COLS) + 1):
            c = ws.cell(row=r, column=col, value=values.get(col))
            c.border = BORDER
            c.font = f(10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if col in (C_STATUS, C_PRIO, C_VERANT, C_FAELL, C_NACHW):
                c.fill = FILL_INPUT
            elif col == C_KLAER:
                c.fill = FILL_CALC
                c.alignment = Alignment(horizontal="center", vertical="top")
            elif col in (C_ID, C_FUND, C_THEMA):
                c.fill = FILL_CALC
        ws.cell(row=r, column=C_ID).font = f(10, True, C_HEAD)
        ws.cell(row=r, column=C_KURZ).font = f(10, True)
        ws.cell(row=r, column=C_TEXT).font = f(9, color="404040")
        ws.cell(row=r, column=C_FAELL).number_format = "DD.MM.YYYY"
        ws.row_dimensions[r].height = 92

    last = R_FIRST + len(rows) - 1
    tab = Table(displayName="tblAnf", ref=f"A{R_HEAD}:O{last}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight8",
                                        showRowStripes=False)
    ws.add_table(tab)

    dv_status = DataValidation(type="list", formula1=f'"{",".join(STATUS)}"',
                               allow_blank=True, showDropDown=False)
    dv_prio = DataValidation(type="list", formula1=f'"{",".join(PRIO)}"',
                             allow_blank=True, showDropDown=False)
    for dv, col in ((dv_status, "K"), (dv_prio, "L")):
        ws.add_data_validation(dv)
        dv.add(f"{col}{R_FIRST}:{col}{last + 100}")

    ws.conditional_formatting.add(
        f"J{R_FIRST}:J{last}",
        CellIsRule(operator="equal", formula=['"ja"'],
                   font=Font(name=FONT, size=10, bold=True, color="9C3400"),
                   fill=PatternFill("solid", bgColor=C_WARN)))
    ws.conditional_formatting.add(
        f"K{R_FIRST}:K{last}",
        CellIsRule(operator="equal", formula=['"umgesetzt"'],
                   font=Font(name=FONT, size=10, color="006100"),
                   fill=PatternFill("solid", bgColor="C6EFCE")))
    ws.conditional_formatting.add(
        f"K{R_FIRST}:K{last}",
        CellIsRule(operator="equal", formula=['"offen"'],
                   font=Font(name=FONT, size=10, color="9C0006"),
                   fill=PatternFill("solid", bgColor="FFC7CE")))

    ws.freeze_panes = f"E{R_FIRST}"
    ws.auto_filter.ref = f"A{R_HEAD}:O{last}"
    ws.print_title_rows = f"{R_HEAD}:{R_HEAD}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    return ws, last


# ---------------------------------------------------------------------------
def count_block(ws, row, title, labels, formula, width=38):
    """Auswertungsblock: Beschriftung, Anzahl, offene Punkte."""
    c = ws.cell(row=row, column=1, value=title)
    c.font = f(11, True, C_HEAD)
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col)
        cell.fill = FILL_SECTION
        cell.border = BORDER
    row += 1
    for head, col in (("Merkmal", 1), ("Anforderungen", 2),
                      ("davon Klärungsbedarf", 3), ("Anteil", 4)):
        cell = ws.cell(row=row, column=col, value=head)
        cell.font = f(10, True, "FFFFFF")
        cell.fill = FILL_HEAD
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left" if col == 1 else "center",
                                   wrap_text=True)
    first = row + 1
    for i, label in enumerate(labels):
        r = first + i
        a = ws.cell(row=r, column=1, value=label)
        a.font = f(10)
        a.fill = FILL_CALC
        a.border = BORDER
        b = ws.cell(row=r, column=2, value=formula(label, r))
        b.font = f(10, True)
        b.fill = FILL_CALC
        b.border = BORDER
        b.alignment = Alignment(horizontal="center")
        d = ws.cell(row=r, column=3, value=formula(label, r, klaer=True))
        d.font = f(10)
        d.fill = FILL_CALC
        d.border = BORDER
        d.alignment = Alignment(horizontal="center")
        e = ws.cell(row=r, column=4,
                    value=f"=IFERROR(B{r}/COUNTA(tblAnf[ID]),0)")
        e.font = f(10)
        e.fill = FILL_CALC
        e.border = BORDER
        e.number_format = "0 %"
        e.alignment = Alignment(horizontal="center")
    last = first + len(labels) - 1
    ws.conditional_formatting.add(
        f"B{first}:B{last}",
        DataBarRule(start_type="num", start_value=0, end_type="max",
                    color="4472C4", showValue=True))
    ws.column_dimensions["A"].width = width
    return first, last


def sheet_auswertung(wb, rows):
    ws = wb.create_sheet("Auswertung")
    header_block(ws, "Auswertung", "Alle Zahlen zählen automatisch aus dem "
                 "Register „Anforderungen“.",
                 "Neue Zeilen im Register werden automatisch mitgezählt.")
    for col, width in zip("ABCD", (38, 15, 20, 12)):
        ws.column_dimensions[col].width = width

    partner = sorted({p for row in rows for p in row["partner"].split(" / ")
                      if p and p != "offen"} | {"offen"})
    komponenten = sorted({k for row in rows for k in row["komp"].split(" / ")
                          if k and k != "–"})
    dokumente = sorted({row["dok"] for row in rows})

    def by(column):
        def maker(label, r, klaer=False):
            base = f'=COUNTIFS(tblAnf[{column}],"*"&A{r}&"*"'
            if klaer:
                base += ',tblAnf[Klärungsbedarf],"ja"'
            return base + ")"
        return maker

    blocks = []
    row = 5
    for title, labels, column in (
            ("1  Nach Themenfeld", THEMEN, "Themenfeld"),
            ("2  Nach Partner", partner, "Partner"),
            ("3  Nach Komponente", komponenten, "Komponente(n)"),
            ("4  Nach Dokument", dokumente, "Dokument"),
    ):
        first, last = count_block(ws, row, title, labels, by(column))
        blocks.append((title, first, last))
        row = last + 3

    ws.print_area = f"A1:D{row}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    return ws, blocks


# ---------------------------------------------------------------------------
def sheet_cockpit(wb, ws_aus, blocks, n_rows):
    ws = wb.create_sheet("Cockpit", 0)
    header_block(ws, "Cockpit – Anforderungen Autostrom / LKW-Laden",
                 "Reservierung, Zufahrt, Entgelte und Nachweise aus der "
                 "Ausschreibung auf einen Blick",
                 "Alle Kacheln und Diagramme rechnen live aus dem Register "
                 "„Anforderungen“.")

    kpis = [
        ("Anforderungen gesamt", "=COUNTA(tblAnf[ID])", "0"),
        ("mit Klärungsbedarf", '=COUNTIF(tblAnf[Klärungsbedarf],"ja")', "0"),
        ("ohne zugeordneten Partner",
         '=COUNTIF(tblAnf[Partner],"offen")+COUNTIF(tblAnf[Partner],"*tbd*")', "0"),
        ("Themenfelder", f"={len(THEMEN)}", "0"),
        ("Status offen", '=COUNTIF(tblAnf[Status],"offen")', "0"),
        ("Status in Klärung", '=COUNTIF(tblAnf[Status],"in Klärung")', "0"),
        ("Status in Umsetzung", '=COUNTIF(tblAnf[Status],"in Umsetzung")', "0"),
        ("Status umgesetzt", '=COUNTIF(tblAnf[Status],"umgesetzt")', "0"),
        ("Bearbeitungsgrad",
         '=IFERROR(COUNTIF(tblAnf[Status],"umgesetzt")/COUNTA(tblAnf[ID]),0)',
         "0 %"),
    ]
    r = 5
    ws.cell(row=r, column=1, value="Kennzahl").font = f(10, True, "FFFFFF")
    ws.cell(row=r, column=2, value="Wert").font = f(10, True, "FFFFFF")
    for col in (1, 2):
        c = ws.cell(row=r, column=col)
        c.fill = FILL_HEAD
        c.border = BORDER
    for i, (label, formula, fmt) in enumerate(kpis):
        rr = r + 1 + i
        a = ws.cell(row=rr, column=1, value=label)
        b = ws.cell(row=rr, column=2, value=formula)
        highlight = label in ("Anforderungen gesamt", "mit Klärungsbedarf",
                              "Bearbeitungsgrad")
        for c in (a, b):
            c.fill = FILL_ACCENT if highlight else FILL_CALC
            c.border = BORDER
            c.font = f(10, highlight)
        b.number_format = fmt
        b.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14

    # Diagramme aus den Auswertungsblöcken
    colors = ["4472C4", "ED7D31", "70AD47", "7030A0"]
    # Ankerzeile je Spalte mitführen, damit sich hohe Diagramme nicht
    # gegenseitig überdecken (eine Standardzeile ist rund 0,53 cm hoch).
    next_row = {"D": 5, "L": 5}
    spalten = ["D", "D", "L", "L"]
    for (title, first, last), spalte, color in zip(blocks, spalten, colors):
        anchor = f"{spalte}{next_row[spalte]}"
        ch = BarChart()
        ch.type = "bar"
        ch.title = title.split("  ", 1)[-1]
        # Höhe an die Zahl der Kategorien koppeln, sonst lässt Excel
        # Beschriftungen weg.
        n = last - first + 1
        ch.height, ch.width = max(8.0, 0.62 * n + 2.6), 14.5
        ch.gapWidth = 40
        ch.legend = None
        ch.add_data(Reference(ws_aus, min_col=2, min_row=first, max_row=last))
        ch.set_categories(Reference(ws_aus, min_col=1, min_row=first,
                                    max_row=last))
        ch.series[0].graphicalProperties.solidFill = color
        ch.series[0].tx = SeriesLabel(v="Anforderungen")
        gp = GraphicalProperties(solidFill="FFFFFF")
        gp.line.solidFill = C_BORDER
        ch.graphical_properties = gp
        ch.x_axis.axPos, ch.y_axis.axPos = "l", "b"
        ch.x_axis.tickLblSkip = 1        # jede Kategorie beschriften
        ch.x_axis.tickMarkSkip = 1
        ch.x_axis.delete, ch.y_axis.delete = False, False
        ch.y_axis.majorGridlines = None
        ws.add_chart(ch, anchor)
        next_row[spalte] += int(ch.height / 0.53) + 3

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    return ws


# ---------------------------------------------------------------------------
def sheet_komponenten(wb, komponenten):
    ws = wb.create_sheet("Komponenten")
    header_block(ws, "Komponenten & Schnittstellen",
                 "Wer liefert was – und wie viele Anforderungen hängen daran",
                 "Die Anzahl rechnet automatisch aus dem Register.")
    heads = ["Art", "Partner", "Komponente(n)", "Anforderungen",
             "davon Klärungsbedarf", "Bemerkung"]
    widths = (16, 24, 44, 15, 20, 40)
    for i, (h, w) in enumerate(zip(heads, widths), start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = f(10, True, "FFFFFF")
        c.fill = FILL_HEAD
        c.border = BORDER
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[5].height = 30

    for i, k in enumerate(komponenten):
        r = 6 + i
        key = k["komp"].split(" / ")[0]
        values = [k["typ"], k["partner"], k["komp"],
                  f'=COUNTIFS(tblAnf[Komponente(n)],"*{key}*")',
                  f'=COUNTIFS(tblAnf[Komponente(n)],"*{key}*",'
                  f'tblAnf[Klärungsbedarf],"ja")', None]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.border = BORDER
            c.font = f(10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.fill = FILL_INPUT if col == 6 else FILL_CALC
            if col in (4, 5):
                c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 30
    last = 5 + len(komponenten)
    ws.conditional_formatting.add(
        f"D6:D{last}", DataBarRule(start_type="num", start_value=0,
                                   end_type="max", color="4472C4",
                                   showValue=True))
    ws.auto_filter.ref = f"A5:F{last}"
    ws.freeze_panes = "A6"
    return ws


def sheet_anzeige(wb):
    ws = wb.create_sheet("Anzeigekonzept")
    header_block(
        ws, "Anzeigekonzept Ladeplatz",
        "Dynamische Anzeige je Ladeplatz – Vorgabe aus Vertragsanlage 4, "
        "Ziffer 3.13.4 (ANF-040)",
        "Offener Punkt aus der Quelldatei: die Breite der Anzeige reicht für "
        "die geforderten Inhalte nicht aus.")
    spalten = ["Zeile", "Geforderter Inhalt", "MCS", "CCS – Zwischenladen",
               "CCS – Übernachtladen"]
    for i, (h, w) in enumerate(zip(spalten, (10, 52, 22, 24, 24)), start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = f(10, True, "FFFFFF")
        c.fill = FILL_HEAD
        c.border = BORDER
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[5].height = 30

    daten = [
        ("Zeile 1", "Typ des Ladeplatzes", "MCS", "CCS-Zwischenladen",
         "CCS-Übernachtladen"),
        ("Zeile 2", "Leistung der Ladesäule", "1000 kW", "400 kW",
         "200 kW  (siehe Hinweis)"),
        ("Zeile 3", "Status (verfügbar, reserviert ab XX:XX Uhr, belegt, "
                    "außer Betrieb)", "verfügbar", "reserviert ab 15:00 Uhr",
         "belegt / außer Betrieb"),
        ("Zeile 4", "optional: Kennzeichen bzw. Reservierungsnummer",
         "bis 15:00 Uhr", "HH-MM 1234", "–"),
    ]
    for i, zeile in enumerate(daten):
        r = 6 + i
        for col, v in enumerate(zeile, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.border = BORDER
            c.font = f(10, col == 1)
            c.fill = FILL_CALC if col <= 2 else FILL_ACCENT
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 26

    r = 12
    ws.cell(row=r, column=1, value="Zusätzlich gefordert").font = f(11, True,
                                                                   C_HEAD)
    weitere = [
        "Ladeplatznummer als Schild: schwarze Nummer auf weißem, "
        "retroreflektierendem Grund, Reflexionsklasse RA1 A "
        "(Vertragsanlage 4, 3.13.4)",
        "Beschilderung mit Hinweis, dass der Stellplatz nach dem Ladevorgang "
        "unverzüglich freizugeben ist – Text bedarf der Zustimmung des "
        "Auftraggebers 2 (Betreibervertrag § 38 Abs. 9)",
        "Display an der Ladeeinrichtung: Ad-hoc-Preis, geladene Energiemenge, "
        "Blockadeentgelt, Bedienhinweise; ablesbar bei Sonne und Dunkelheit, "
        "Menü mindestens Deutsch und Englisch (Vertragsanlage 6, Ziffer 4)",
        "Offener Punkt: die vorgesehene Anzeigenbreite reicht laut Prüfung "
        "nicht für alle vier Zeilen – Layout oder Hardware klären.",
        "Offener Punkt: die Quelldatei nennt für CCS-Übernachtladen an einer "
        "Stelle 250 kW und an anderer 200 kW – zu vereinheitlichen.",
    ]
    for i, t in enumerate(weitere):
        c = ws.cell(row=r + 1 + i, column=1, value="•  " + t)
        c.font = f(10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r + 1 + i, start_column=1,
                       end_row=r + 1 + i, end_column=5)
        ws.row_dimensions[r + 1 + i].height = 30
    return ws


def sheet_lesehilfe(wb, n_rows, n_komp):
    ws = wb.create_sheet("Lesehilfe")
    header_block(ws, "Lesehilfe", "Wie diese Mappe aufgebaut ist und wie damit "
                 "gearbeitet wird", None)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 96

    inhalt = [
        ("Zweck", "Die Mappe führt alle vertraglichen Anforderungen aus der "
                  "Ausschreibung „Autostrom / LKW-Laden“ an einer Stelle "
                  "zusammen: was gefordert ist, wo es steht, wer es liefert "
                  "und wie weit die Umsetzung ist."),
        ("Cockpit", "Kennzahlen und vier Diagramme: Anforderungen je "
                    "Themenfeld, Partner, Komponente und Quelldokument. "
                    "Rechnet live mit."),
        ("Anforderungen", f"Das Register mit {n_rows} Anforderungen. Weiße "
                          "Spalten enthalten den Inhalt aus der "
                          "Ausschreibung, gelbe Spalten sind eure Steuerung "
                          "(Status, Priorität, Verantwortlich, Fällig, "
                          "Nachweis)."),
        ("Auswertung", "Zählt Anforderungen und offene Klärungen je "
                       "Themenfeld, Partner, Komponente und Dokument."),
        ("Komponenten", f"Die {n_komp} Komponenten und Schnittstellen aus der "
                        "Quelldatei mit der Anzahl daran hängender "
                        "Anforderungen."),
        ("Anzeigekonzept", "Der Inhalt der dynamischen Ladeplatzanzeige je "
                           "Ladeplatztyp inklusive des offenen Punktes zur "
                           "Anzeigenbreite."),
        (None, None),
        ("Spalte „Fundstelle“", "Paragraf, Absatz und Unterpunkt des "
                                "Quelldokuments, z. B. § 25 (2) c."),
        ("Spalte „Themenfeld“", "Fachliche Bündelung: " + ", ".join(THEMEN) + "."),
        ("Spalte „Klärungsbedarf“", "Wird automatisch auf „ja“ gesetzt, wenn "
                                    "kein Partner feststeht (offen oder tbd) "
                                    "oder die Bemerkung eine Frage enthält."),
        ("Spalte „Status“", "Auswahl: " + ", ".join(STATUS) + "."),
        (None, None),
        ("Neue Anforderung", "Direkt unter der letzten Zeile des Registers "
                             "eintragen – die Tabelle wächst automatisch, "
                             "alle Zählungen und Diagramme ziehen mit."),
        ("Originaltexte", "Die Anforderungstexte wurden unverändert "
                          "übernommen; lediglich die Trennzeichen aus dem "
                          "PDF-Import (_x0002_) wurden entfernt."),
    ]
    r = 5
    for label, text in inhalt:
        if label is None:
            r += 1
            continue
        a = ws.cell(row=r, column=1, value=label)
        a.font = f(10, True, C_HEAD)
        a.fill = FILL_SECTION
        a.border = BORDER
        a.alignment = Alignment(vertical="top", wrap_text=True)
        b = ws.cell(row=r, column=2, value=text)
        b.font = f(10)
        b.fill = FILL_CALC
        b.border = BORDER
        b.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 34
        r += 1
    return ws


# ---------------------------------------------------------------------------
def main():
    src = sys.argv[1]
    dst = sys.argv[2]
    rows, komponenten = read_source(src)

    wb = Workbook()
    wb.remove(wb.active)
    _, last = sheet_register(wb, rows)
    ws_aus, blocks = sheet_auswertung(wb, rows)
    sheet_cockpit(wb, ws_aus, blocks, len(rows))
    sheet_komponenten(wb, komponenten)
    sheet_anzeige(wb)
    sheet_lesehilfe(wb, len(rows), len(komponenten))
    wb.move_sheet("Cockpit", offset=-wb.sheetnames.index("Cockpit"))
    wb.save(dst)
    print(f"geschrieben: {dst} | {len(rows)} Anforderungen, "
          f"{len(komponenten)} Komponenten | Reiter: {wb.sheetnames}")


if __name__ == "__main__":
    main()
