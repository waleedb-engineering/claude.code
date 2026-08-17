#!/usr/bin/env python3
"""Verlaengert die produktive Cashflow-Datei von Dez 2027 auf Dez 2030.

Die Datei wird nicht neu gebaut, sondern in ihrer vorhandenen Struktur
erweitert: Auftragsdaten, Formelaufbau, Tabellen, Formate und Kostenstellen
bleiben unveraendert. Ergaenzt werden 36 Monatsspalten, drei Jahres-
Summenspalten und die zugehoerigen Bloecke in der Uebersicht.

Aufruf:  python extend_live_file_to_2030.py QUELLE.xlsx ZIEL.xlsx
"""

import sys
from copy import copy
from datetime import date

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.data_source import StrRef
from openpyxl.chart.series import SeriesLabel
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import TableColumn

# ---------------------------------------------------------------------------
# Bestehendes Raster der Quelldatei
# ---------------------------------------------------------------------------
C_FIRST_M = 3                 # C = Jul 26
C_OLD_LAST_M = 20             # T = Dez 27
C_OLD_SUMS = (21, 22, 23)     # U/V/W = Summe H2 2026, Summe 2027, Gesamt

N_NEW_MONTHS = 36             # Jan 28 – Dez 30
C_NEW_LAST_M = C_OLD_LAST_M + N_NEW_MONTHS          # BD = Dez 30
C_SUM_FIRST = C_NEW_LAST_M + 1                      # BE
C_TOTAL = C_SUM_FIRST + 5                           # BJ

MONTH_ABBR = ["Jan", "Feb", "Mär", "Apr", "Mai", "Jun",
              "Jul", "Aug", "Sep", "Okt", "Nov", "Dez"]

# Zeilenraster (aus der Quelldatei ausgelesen)
R_MONTHKEY = 5
R_HEAD = 6
HEADER_ROWS = (6, 7, 41, 75, 81, 91)     # Zeilen mit Monatsbeschriftungen
R_IN_FIRST, R_IN_LAST = 8, 37
R_IN_REST, R_SUM_IN = 38, 39
R_OUT_FIRST, R_OUT_LAST = 42, 71
R_OUT_REST, R_SUM_OUT = 72, 73
R_KPI_IN, R_KPI_OUT, R_KPI_NET, R_KPI_CUM = 76, 77, 78, 79
FORMULA_ROWS = (list(range(R_IN_FIRST, R_SUM_IN + 1))
                + list(range(R_OUT_FIRST, R_SUM_OUT + 1))
                + [R_KPI_IN, R_KPI_OUT, R_KPI_NET, R_KPI_CUM]
                + list(range(82, 90)) + list(range(92, 100)))

TABLES = {"Tabelle3": (7, 38), "Tabelle4": (41, 72),
          "Tabelle5": (81, 89), "Tabelle6": (91, 99)}


def month_dates():
    """Monatsanfänge der 36 neuen Spalten: Jan 2028 bis Dez 2030."""
    out = []
    for year in (2028, 2029, 2030):
        out += [date(year, m, 1) for m in range(1, 13)]
    return out


NEW_DATES = month_dates()
assert len(NEW_DATES) == N_NEW_MONTHS

# Jahresgrenzen als Spaltenindizes (erste/letzte Monatsspalte je Periode)
PERIODS = [
    ("Summe H2 2026", "Summe H2 26", 3, 8),        # C:H
    ("Summe 2027", "Summe 27", 9, 20),             # I:T
    ("Summe 2028", "Summe 28", 21, 32),            # U:AF
    ("Summe 2029", "Summe 29", 33, 44),            # AG:AR
    ("Summe 2030", "Summe 30", 45, 56),            # AS:BD
    ("Gesamt", "Gesamt", 3, 56),                   # C:BD
]


def L(col):
    return get_column_letter(col)


def clone_style(src, dst):
    dst._style = copy(src._style)


# ---------------------------------------------------------------------------
def extend_cashflow(ws):
    # ---- 1. Monatsschlüssel in der ausgeblendeten Zeile 5 ------------------
    src = ws.cell(row=R_MONTHKEY, column=C_OLD_LAST_M)
    for i, d in enumerate(NEW_DATES):
        c = ws.cell(row=R_MONTHKEY, column=C_OLD_LAST_M + 1 + i, value=d)
        clone_style(src, c)

    # ---- 2. Monatsbeschriftungen in allen Kopfzeilen -----------------------
    for row in HEADER_ROWS:
        src = ws.cell(row=row, column=C_OLD_LAST_M)
        for i, d in enumerate(NEW_DATES):
            c = ws.cell(row=row, column=C_OLD_LAST_M + 1 + i,
                        value=f"{MONTH_ABBR[d.month - 1]} {d:%y}")
            clone_style(src, c)
        # Summenspalten-Überschriften: Zeile 6 ausgeschrieben, sonst kurz
        style_src = ws.cell(row=row, column=C_OLD_SUMS[0])
        total_src = ws.cell(row=row, column=C_OLD_SUMS[2])
        for i, (lang, kurz, _, _) in enumerate(PERIODS):
            c = ws.cell(row=row, column=C_SUM_FIRST + i,
                        value=lang if row == R_HEAD else kurz)
            clone_style(total_src if i == len(PERIODS) - 1 else style_src, c)

    # ---- 3. Monatsformeln fortschreiben ------------------------------------
    for row in FORMULA_ROWS:
        # Spalte D ist das allgemeine Muster (C enthält bei der kumulierten
        # Zeile den Startwert und darf nicht kopiert werden).
        pattern = ws.cell(row=row, column=C_FIRST_M + 1)
        style_src = ws.cell(row=row, column=C_OLD_LAST_M)
        for col in range(C_OLD_LAST_M + 1, C_NEW_LAST_M + 1):
            c = ws.cell(row=row, column=col)
            c.value = Translator(pattern.value,
                                 origin=pattern.coordinate).translate_formula(
                                     f"{L(col)}{row}")
            clone_style(style_src, c)

    # ---- 4. Summenspalten neu setzen ---------------------------------------
    style_sum = {row: copy(ws.cell(row=row, column=C_OLD_SUMS[0])._style)
                 for row in FORMULA_ROWS}
    style_total = {row: copy(ws.cell(row=row, column=C_OLD_SUMS[2])._style)
                   for row in FORMULA_ROWS}

    def sum_formula(row, first, last):
        return f"=SUM({L(first)}{row}:{L(last)}{row})"

    for row in FORMULA_ROWS:
        for i, (_, _, first, last) in enumerate(PERIODS):
            col = C_SUM_FIRST + i
            x = L(col)
            if row == R_KPI_CUM:                      # Periodenendstand
                value = f"={L(last)}{row}"
            elif row == R_IN_REST:
                value = f"={x}{R_SUM_IN}-SUM({x}{R_IN_FIRST}:{x}{R_IN_LAST})"
            elif row == R_OUT_REST:
                value = f"={x}{R_SUM_OUT}-SUM({x}{R_OUT_FIRST}:{x}{R_OUT_LAST})"
            elif row == R_KPI_IN:
                value = f"={x}{R_SUM_IN}"
            elif row == R_KPI_OUT:
                value = f"={x}{R_SUM_OUT}"
            elif row == R_KPI_NET:
                value = f"={x}{R_KPI_IN}-{x}{R_KPI_OUT}"
            else:
                value = sum_formula(row, first, last)
            c = ws.cell(row=row, column=col, value=value)
            c._style = copy(style_total[row] if i == len(PERIODS) - 1
                            else style_sum[row])

    # ---- 5. Spaltenbreiten --------------------------------------------------
    for col in range(C_OLD_LAST_M + 1, C_NEW_LAST_M + 1):
        ws.column_dimensions[L(col)].width = 13
    for col in range(C_SUM_FIRST, C_TOTAL + 1):
        ws.column_dimensions[L(col)].width = 16

    # ---- 6. Bedingte Formatierung auf die neue Breite ziehen ----------------
    new_end = L(C_TOTAL)

    def split_ref(ref):
        col = "".join(ch for ch in ref if ch.isalpha())
        return column_index_from_string(col), "".join(ch for ch in ref
                                                      if ch.isdigit())

    rules = [(str(rng.sqref), list(rng.rules)) for rng in ws.conditional_formatting]
    ws.conditional_formatting._cf_rules.clear()
    for sqref, rule_list in rules:
        parts = []
        for part in sqref.split():
            a, _, b = part.partition(":")
            end = b or a
            col, row = split_ref(end)
            # alles, was bisher am rechten Rand endete, auf die neue Breite ziehen
            if col >= C_OLD_LAST_M:
                part = f"{a}:{new_end}{row}"
            parts.append(part)
        for rule in rule_list:
            ws.conditional_formatting.add(" ".join(parts), rule)

    # ---- 7. Tabellenbereiche erweitern -------------------------------------
    for name, (head_row, last_row) in TABLES.items():
        tab = ws.tables[name]
        tab.ref = f"B{head_row}:{L(C_TOTAL)}{last_row}"
        tab.autoFilter.ref = tab.ref
        names = [ws.cell(row=head_row, column=col).value
                 for col in range(2, C_TOTAL + 1)]
        tab.tableColumns = [TableColumn(id=i + 1, name=str(n))
                            for i, n in enumerate(names)]

    # ---- 8. Druckbereich ----------------------------------------------------
    ws.print_area = f"A1:{L(C_TOTAL)}{ws.max_row}"
    ws.print_title_cols = "$A:$B"
    ws.page_setup.fitToWidth = 3
    ws.page_setup.fitToHeight = 0


# ---------------------------------------------------------------------------
def rebuild_overview(ov, cf):
    """Kennzahlenblock auf sechs Perioden erweitern; Stile werden aus den
    vorhandenen Zeilen übernommen."""
    st_sec = (copy(ov["A6"]._style), copy(ov["B6"]._style))
    st_row = (copy(ov["A7"]._style), copy(ov["B7"]._style))
    st_bold = (copy(ov["A9"]._style), copy(ov["B9"]._style))

    ov["A2"] = "SmartInfra / DuSL  ·  Juli 2026 – Dezember 2030"

    groups = [
        ("H2 2026 (Jul – Dez 2026)", "H2 2026", C_SUM_FIRST),
        ("Geschäftsjahr 2027", "2027", C_SUM_FIRST + 1),
        ("Geschäftsjahr 2028", "2028", C_SUM_FIRST + 2),
        ("Geschäftsjahr 2029", "2029", C_SUM_FIRST + 3),
        ("Geschäftsjahr 2030", "2030", C_SUM_FIRST + 4),
        ("Gesamtzeitraum (54 Monate)", None, C_TOTAL),
    ]

    row = 6
    for title, short, col in groups:
        a, b = ov.cell(row=row, column=1, value=title), ov.cell(row=row, column=2)
        a._style, b._style = copy(st_sec[0]), copy(st_sec[1])
        b.value = None
        row += 1
        if short is None:
            labels = ["Gesamteinnahmen", "Gesamtausgaben", "Gesamt-Netto-Cashflow"]
        else:
            labels = [f"Einnahmen {short}", f"Ausgaben {short}",
                      f"Netto-Cashflow {short}"]
        for i, (label, src_row) in enumerate(zip(labels,
                                                 (R_KPI_IN, R_KPI_OUT, R_KPI_NET))):
            style = st_bold if i == 2 else st_row
            a = ov.cell(row=row, column=1, value=label)
            b = ov.cell(row=row, column=2, value=f"=Cashflow!{L(col)}{src_row}")
            a._style, b._style = copy(style[0]), copy(style[1])
            row += 1

    # ---- Bedingte Formatierung auf alle Kennzahlenzeilen ausdehnen ---------
    # Sie endete bisher bei Zeile 17; die neuen Blöcke 2029/2030/Gesamt
    # blieben dadurch ungefärbt und der Block wirkte uneinheitlich.
    last_row = row - 1
    rules = [(str(rng.sqref), list(rng.rules)) for rng in ov.conditional_formatting]
    ov.conditional_formatting._cf_rules.clear()
    for sqref, rule_list in rules:
        parts = []
        for part in sqref.split():
            a, _, b = part.partition(":")
            if b and b.startswith("B"):
                b = f"B{last_row}"
            parts.append(f"{a}:{b}" if b else a)
        for rule in rule_list:
            ov.conditional_formatting.add(" ".join(parts), rule)

    # ---- Diagramme neu aufsetzen -------------------------------------------
    # Die bisherigen Datenreihen begannen in Spalte B, also bei der
    # Beschriftungszelle – dadurch waren alle Werte um einen Monat gegenüber
    # der Achsenbeschriftung verschoben. Hier beginnen sie bei Spalte C.
    ov._charts = []
    cats = Reference(cf, min_col=C_FIRST_M, max_col=C_NEW_LAST_M,
                     min_row=R_HEAD, max_row=R_HEAD)

    bar = BarChart()
    bar.type, bar.grouping = "col", "clustered"
    bar.title = "Einnahmen vs. Ausgaben (monatlich)"
    bar.y_axis.title, bar.y_axis.numFmt = "EUR", "#,##0"
    bar.height, bar.width = 9, 32
    bar.gapWidth, bar.overlap = 60, -10
    bar.add_data(Reference(cf, min_col=C_FIRST_M, max_col=C_NEW_LAST_M,
                           min_row=R_KPI_IN, max_row=R_KPI_OUT), from_rows=True)
    bar.set_categories(cats)
    bar.x_axis.tickLblSkip = 3          # 54 Monate: nur jedes Quartal beschriften
    bar.x_axis.tickMarkSkip = 3
    for ser, r in zip(bar.series, (R_KPI_IN, R_KPI_OUT)):
        ser.tx = SeriesLabel(strRef=StrRef(f"Cashflow!$B${r}"))
    bar.series[0].graphicalProperties.solidFill = "4472C4"
    bar.series[1].graphicalProperties.solidFill = "C00000"
    tidy_axes(bar)
    ov.add_chart(bar, "D5")

    line = LineChart()
    line.title = "Netto-Cashflow und kumulierter Cashflow (monatlich)"
    line.y_axis.title, line.y_axis.numFmt = "EUR", "#,##0"
    line.height, line.width = 9, 32
    line.add_data(Reference(cf, min_col=C_FIRST_M, max_col=C_NEW_LAST_M,
                            min_row=R_KPI_NET, max_row=R_KPI_CUM), from_rows=True)
    line.set_categories(cats)
    line.x_axis.tickLblSkip = 3
    line.x_axis.tickMarkSkip = 3
    for ser, r in zip(line.series, (R_KPI_NET, R_KPI_CUM)):
        ser.tx = SeriesLabel(strRef=StrRef(f"Cashflow!$B${r}"))
    line.series[0].graphicalProperties.line.solidFill = "1F3864"
    line.series[1].graphicalProperties.line.solidFill = "70AD47"
    line.series[1].graphicalProperties.line.dashStyle = "dash"
    for ser in line.series:            # keine Spline-Glättung: der kumulierte
        ser.smooth = False             # Verlauf ist eine Treppe, keine Kurve
    tidy_axes(line)
    ov.add_chart(line, "D24")


def tidy_axes(chart):
    """Achsen so setzen, wie Excel sie für ein stehendes Diagramm erwartet.

    openpyxl legt die Kategorienachse standardmäßig links an (axPos "l") und
    zeichnet ihre Beschriftungen am Achsenschnittpunkt – bei negativen Werten
    also mitten durch die Zeichenfläche. Beides wird hier korrigiert."""
    chart.x_axis.axPos = "b"
    chart.y_axis.axPos = "l"
    chart.x_axis.tickLblPos = "low"      # Monatsnamen immer unterhalb
    chart.y_axis.tickLblPos = "nextTo"
    chart.x_axis.delete = False          # sonst blendet Excel die Achse aus
    chart.y_axis.delete = False
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"


def restore_kostenstellen_dropdown(ws):
    """Die Kostenstellen-Auswahl in Spalte C lag als x14-Erweiterung vor, die
    beim Speichern verlorengeht – hier als normale Gültigkeitsprüfung neu."""
    dv = DataValidation(type="list", formula1="=Cashflow!$B$8:$B$37",
                        allow_blank=True, showDropDown=False)
    dv.errorTitle = "Unbekannte Kostenstelle"
    dv.error = ("Bitte eine Kostenstelle aus dem Cashflow-Reiter wählen "
                "(Zeilen 8–37).")
    ws.add_data_validation(dv)
    dv.add("C7:C156")


def main():
    src, dst = sys.argv[1], sys.argv[2]
    wb = load_workbook(src)
    cf, ov, au = wb["Cashflow"], wb["Übersicht"], wb["Aufträge"]
    extend_cashflow(cf)
    rebuild_overview(ov, cf)
    restore_kostenstellen_dropdown(au)
    wb.save(dst)
    print(f"geschrieben: {dst}")


if __name__ == "__main__":
    main()
