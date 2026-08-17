#!/usr/bin/env python3
"""Erzeugt die Cashflow-Mappe fuer SmartInfra / IoT-Buero.

Projekt : DusL Plus Autostrom
Zeitraum: Juli 2026 - Dezember 2030 (54 Monate, monatlich)

Datenfluss:  Auftraege  ->  Cashflow  ->  Uebersicht

    Auftraege  zentrale Eingabe (echte Excel-Tabelle "tblAuftraege")
    Cashflow   monatliche Aggregation per SUMIFS ueber echte Datumsgrenzen
    Uebersicht Management-Zusammenfassung, zieht aus dem Cashflow

Aufruf:
    python build_cashflow_template.py                  # leere Vorlage
    python build_cashflow_template.py --source ALT.xlsx
        -> Positionsbezeichnungen (Kostenstellen-Zeilen) werden aus ALT.xlsx
           uebernommen. Betraege werden bewusst NICHT uebernommen: der
           Cashflow wird nicht mehr manuell gepflegt, alle Betraege stammen
           ab sofort aus dem Reiter "Auftraege".

Es werden keine Auftraege, Umsaetze oder Kosten erfunden.
"""

import argparse
import re
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.data_source import StrRef
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = Path(__file__).resolve().parent
OUT = HERE / "Cashflow_SmartInfra_IoT_DusL_Plus_Autostrom_2026_2030.xlsx"

# ----------------------------------------------------------------------------
# Layout-Konstanten
# ----------------------------------------------------------------------------
FONT = "Arial"

N_INCOME_ROWS = 50          # Auftragszeilen im Einnahmenblock
N_EXPENSE_ROWS = 50         # Auftragszeilen im Ausgabenblock
N_KST_ROWS = 15             # Zeilen je Kostenstellen-Auswertung
N_ORDER_ROWS = 300          # vorbereitete Zeilen der Auftragstabelle

PERIOD_START = date(2026, 7, 1)
PERIOD_END = date(2030, 12, 1)          # letzter Monat des Planungszeitraums

MONTH_STARTS = [PERIOD_START]
while MONTH_STARTS[-1] < PERIOD_END:
    _y, _m = MONTH_STARTS[-1].year, MONTH_STARTS[-1].month + 1
    MONTH_STARTS.append(date(_y + (_m > 12), 1 if _m > 12 else _m, 1))

MONTH_ABBR = ["Jan", "Feb", "Mär", "Apr", "Mai", "Jun",
              "Jul", "Aug", "Sep", "Okt", "Nov", "Dez"]
MONTHS = [f"{MONTH_ABBR[d.month - 1]} {d:%y}" for d in MONTH_STARTS]
N_MONTHS = len(MONTHS)


def _periods():
    """Summenspalten je Geschäftsjahr; ein angebrochenes erstes Jahr wird als
    Halbjahr ausgewiesen. Liefert (Beschriftung, erster Index, letzter Index)."""
    out = []
    for year in sorted({d.year for d in MONTH_STARTS}):
        idx = [i for i, d in enumerate(MONTH_STARTS) if d.year == year]
        full = len(idx) == 12
        if full:
            label = f"Summe {year}"
        elif MONTH_STARTS[idx[0]].month >= 7:
            label = f"Summe H2 {year}"
        else:
            label = f"Summe {year} (Teiljahr)"
        out.append((label, idx[0], idx[-1]))
    return out


PERIODS = _periods()

THEMENFELDER = ["IoT", "EDM", "Parking", "LLM"]
KATEGORIEN = ["Hardware", "Hardware + Dienstleistung", "Lizenzen", "Dienstleistung"]

COL_POS = 1                      # A = Auftrag (Schlüssel, Eingabe)
COL_KST = 2                      # B = Kostenstelle   (automatisch nachgeschlagen)
COL_TFD = 3                      # C = Themenfeld     (automatisch nachgeschlagen)
COL_KTG = 4                      # D = Kategorie      (automatisch nachgeschlagen)
COL_LABELS = (COL_POS, COL_KST, COL_TFD, COL_KTG)
COL_M1 = 5                                  # E = erster Monat
COL_LAST_M = COL_M1 + N_MONTHS - 1          # letzter Monat
COL_PERIOD = {label: COL_LAST_M + 1 + i     # eine Summenspalte je Periode
              for i, (label, _, _) in enumerate(PERIODS)}
COL_TOTAL = COL_LAST_M + len(PERIODS) + 1   # Gesamtspalte

L_KEY = get_column_letter(COL_POS)
L_M1 = get_column_letter(COL_M1)
L_LAST_M = get_column_letter(COL_LAST_M)
L_TOTAL = get_column_letter(COL_TOTAL)

# Spalte -> (erster Monatsindex, letzter Monatsindex) der zugehörigen Periode
PERIOD_BY_COL = {COL_PERIOD[label]: (first, last)
                 for label, first, last in PERIODS}
PERIOD_BY_COL[COL_TOTAL] = (0, N_MONTHS - 1)
SUM_COLS = tuple(COL_PERIOD[label] for label, _, _ in PERIODS) + (COL_TOTAL,)

# Zeilenraster des Cashflow-Blatts (aus den Blockgrößen abgeleitet)
R_TITLE = 1
R_MONTHKEY = 5                                     # ausgeblendet: echte Monatsanfänge
R_HEAD = 6
R_SEC_IN = R_HEAD + 1
R_IN_FIRST = R_SEC_IN + 1
R_IN_LAST = R_IN_FIRST + N_INCOME_ROWS - 1
R_IN_REST = R_IN_LAST + 1                          # ohne Kostenstellen-Zuordnung
R_SUM_IN = R_IN_REST + 1
R_SEC_OUT = R_SUM_IN + 2
R_OUT_FIRST = R_SEC_OUT + 1
R_OUT_LAST = R_OUT_FIRST + N_EXPENSE_ROWS - 1
R_OUT_REST = R_OUT_LAST + 1
R_SUM_OUT = R_OUT_REST + 1
R_SEC_KPI = R_SUM_OUT + 2
R_KPI_IN = R_SEC_KPI + 1
R_KPI_OUT = R_KPI_IN + 1
R_KPI_NET = R_KPI_OUT + 1
R_KPI_CUM = R_KPI_NET + 1
R_SEC_TF = R_KPI_CUM + 2                           # Auswertung nach Themenfeld
R_TF_FIRST = R_SEC_TF + 1
R_SEC_KAT = R_TF_FIRST + 3 * len(THEMENFELDER) + 1  # Auswertung nach Kategorie
R_KAT_FIRST = R_SEC_KAT + 1
R_SEC_FC = R_KAT_FIRST + 2 * len(KATEGORIEN) + 1       # Forecast-Bereich
R_FCBOX = R_SEC_FC + 1                                 # Titel "Pipeline Forecast"
R_FCBOX_FIRST = R_FCBOX + 1                            # Kennzahlen der Box
N_FCBOX_ROWS = 15
R_FC_SUB1 = R_FCBOX_FIRST + N_FCBOX_ROWS               # Zwischenüberschrift
R_FC_FIRST = R_FC_SUB1 + 1                             # 4 gewichtete Monatszeilen
R_FC_NET = R_FC_FIRST + 2
R_FC_CUM = R_FC_FIRST + 3
R_FC_SUB2 = R_FC_FIRST + 4                             # Zwischenüberschrift 100 %
R_SAFE_FIRST = R_FC_SUB2 + 1                           # 3 sichere Monatszeilen
R_SAFE_NET = R_SAFE_FIRST + 2
R_SEC_KST_IN = R_SAFE_NET + 2                          # Einnahmen je Kostenstelle
R_KST_IN_FIRST = R_SEC_KST_IN + 1
R_SEC_KST_OUT = R_KST_IN_FIRST + N_KST_ROWS + 1        # Ausgaben je Kostenstelle
R_KST_OUT_FIRST = R_SEC_KST_OUT + 1
R_MEMO_LAST = R_KST_OUT_FIRST + N_KST_ROWS - 1
R_LEGEND = R_MEMO_LAST + 3

# Auftragsblatt
A_TITLE = 1
A_HEAD = 6
A_FIRST = A_HEAD + 1
A_LAST = A_HEAD + N_ORDER_ROWS
TBL = "tblAuftraege"
ORDER_COLS = ["Auftrag", "Themenfeld", "Kategorie", "Kostenstelle",
              "Datum Einnahmen", "Datum Ausgaben", "Einnahmen", "Ausgaben",
              "Auftragswahrscheinlichkeit"]
A_COL_AUFTRAG = 1     # A – Schlüsselspalte, entspricht "Position" im Cashflow
A_COL_TFD = 2         # B – Dropdown Themenfeld
A_COL_KTG = 3         # C – Dropdown Kategorie
C_EINNAHMEN = f"{TBL}[Einnahmen]"
C_AUSGABEN = f"{TBL}[Ausgaben]"
C_DAT_EIN = f"{TBL}[Datum Einnahmen]"
C_DAT_AUS = f"{TBL}[Datum Ausgaben]"
C_KST = f"{TBL}[Kostenstelle]"
C_AUFTRAG = f"{TBL}[Auftrag]"
C_TF = f"{TBL}[Themenfeld]"
C_KAT = f"{TBL}[Kategorie]"
C_WSK = f"{TBL}[Auftragswahrscheinlichkeit]"

PERIOD_LABEL = (f"Juli {MONTH_STARTS[0].year} – "
                f"Dezember {MONTH_STARTS[-1].year}")
META_DEFAULT = [
    "Cashflow-Planung",
    "Bereich:  SmartInfra / IoT-Büro",
    f"Zeitraum:  {PERIOD_LABEL}  ({N_MONTHS} Monate, monatliche Granularität)",
]
META_HINT = ("Alle Beträge werden automatisch aus dem Reiter „Aufträge“ berechnet – "
             "hier wird links nur die Auftragsnummer (gelb) eingetragen.")

# Farben (dezentes Controlling-Layout)
C_HEAD = "1F3864"        # dunkelblau
C_SECTION = "D9E2F3"     # helles blau
C_INPUT = "FFF9E1"       # helles gelb  -> Eingabezellen
C_CALC = "F2F2F2"        # hellgrau     -> berechnete Zellen
C_RESULT = "E2EFDA"      # helles gruen -> Ergebniszeilen
C_FORECAST = "EDF2F9"    # sehr helles blau -> Forecast (gewichtet)
C_BORDER = "BFBFBF"

EUR = '#,##0.00\\ "€";[Red]-#,##0.00\\ "€";"–"'
DATE_FMT = "DD.MM.YYYY"
PCT = "0 %"
COUNT_FMT = "#,##0"

thin = Side(style="thin", color=C_BORDER)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
TOP_LINE = Border(left=thin, right=thin, top=Side(style="medium", color="8EA9DB"),
                  bottom=thin)

FILL_INPUT = PatternFill("solid", fgColor=C_INPUT)
FILL_CALC = PatternFill("solid", fgColor=C_CALC)
FILL_RESULT = PatternFill("solid", fgColor=C_RESULT)
FILL_SECTION = PatternFill("solid", fgColor=C_SECTION)
FILL_FORECAST = PatternFill("solid", fgColor=C_FORECAST)
FILL_HEAD = PatternFill("solid", fgColor=C_HEAD)
NO_FILL = PatternFill(fill_type=None)


def f(size=10, bold=False, color="000000", italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)


def period_sum(col, row):
    """Formel einer Summenspalte: Summe über die Monate ihrer Periode."""
    first, last = PERIOD_BY_COL[col]
    a = get_column_letter(COL_M1 + first)
    b = get_column_letter(COL_M1 + last)
    return f"=SUM({a}{row}:{b}{row})"


def period_end_ref(col, row):
    """Letzter Monat einer Periode – für kumulierte Zeilen (Endstand)."""
    _, last = PERIOD_BY_COL[col]
    return f"={get_column_letter(COL_M1 + last)}{row}"


def month_window(col):
    """Datumsgrenzen des Monats einer Spalte: >= Monatsanfang, < Folgemonat.
    Die Monatsanfänge stehen als echte Datumswerte in der ausgeblendeten
    Zeile 5, verglichen wird also nie mit Text wie "Okt 26"."""
    letter = get_column_letter(col)
    return (f'">="&{letter}${R_MONTHKEY}',
            f'"<"&EDATE({letter}${R_MONTHKEY},1)')


def weighted(value_col, date_col, col, extra=None):
    """Betrag x Auftragswahrscheinlichkeit, auf den Monat der Spalte begrenzt.

    Die Mehrargument-Form von SUMMENPRODUKT behandelt leere und nicht
    numerische Zellen als 0 – eine leere Wahrscheinlichkeit wirkt damit wie
    0 % und ein leeres Datum fällt aus dem Zeitfenster."""
    letter = get_column_letter(col)
    parts = [value_col, C_WSK,
             f"({date_col}>={letter}${R_MONTHKEY})*1",
             f"({date_col}<EDATE({letter}${R_MONTHKEY},1))*1"]
    if extra:
        parts += list(extra)
    return "=SUMPRODUCT(" + ",".join(parts) + ")"


def sumifs(value_col, date_col, col, extra=None):
    lo, hi = month_window(col)
    parts = [value_col, date_col, lo, date_col, hi]
    if extra:
        parts += list(extra)
    return "=SUMIFS(" + ",".join(parts) + ")"


# ----------------------------------------------------------------------------
# Uebernahme aus einer bestehenden Datei
# ----------------------------------------------------------------------------
def read_source(path):
    """Liest Kopfzeilen und Positionsbezeichnungen aus einer frueheren Fassung.
    Das Zeilenraster wird ueber die Abschnittsmarken erkannt."""
    ws = load_workbook(path, data_only=False)["Cashflow"]

    marks = {}
    for row in ws.iter_rows(min_col=1, max_col=1):
        v = row[0].value
        if isinstance(v, str):
            key = v.strip().lower()
            if key.startswith("1 "):
                marks.setdefault("in", row[0].row)
            if key.startswith("2 "):
                marks.setdefault("out", row[0].row)
            if key.startswith("summe einnahmen"):
                marks.setdefault("sum_in", row[0].row)
            if key.startswith("summe ausgaben"):
                marks.setdefault("sum_out", row[0].row)
            if key.startswith("position"):
                marks.setdefault("head", row[0].row)
    missing = {"in", "out", "sum_in", "sum_out", "head"} - marks.keys()
    if missing:
        raise SystemExit(f"Quelldatei: Abschnittsmarken nicht gefunden: {missing}")

    def labels(first, last):
        out = []
        for r in range(first, last + 1):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and v.strip() and not v.lower().startswith(
                    ("summe", "nicht zugeordnet", "aufträge ohne",
                     "noch nicht erfasste")):
                out.append(v.strip())
        return out

    meta = []
    for r in range(1, marks["head"]):
        v = ws.cell(row=r, column=COL_POS).value
        if isinstance(v, str) and v.strip() and not v.startswith("Alle Beträge"):
            meta.append(v.strip())

    # Kostenstellen aus den bisherigen Zeilenbeschriftungen ableiten
    # ("87004006 - LKW / Autostrom+"  ->  Nummer "87004006", Name "LKW / ...").
    kostenstellen, seen = [], {}
    for label in labels(marks["in"] + 1, marks["sum_in"] - 1) + \
            labels(marks["out"] + 1, marks["sum_out"] - 1):
        m = re.match(r"^\s*(\d{4,})\s*[-–—]\s*(.+)$", label)
        nummer, name = (m.group(1), m.group(2).strip()) if m else (label, "")
        if nummer in seen:
            if name and name not in seen[nummer]:
                seen[nummer].append(name)
            continue
        seen[nummer] = [name] if name else []
        kostenstellen.append(nummer)
    # Zeitraumzeile stammt immer aus den aktuellen Konstanten, nie aus der
    # Quelldatei – sonst würde ein alter Planungszeitraum weitergeschleppt.
    meta = [m for m in meta if not m.lower().startswith("zeitraum")][:2]
    meta = (meta or META_DEFAULT[:2]) + [META_DEFAULT[2]]
    return {
        "meta": meta,
        "kostenstellen": [(n, " / ".join(seen[n])) for n in kostenstellen],
    }


# ----------------------------------------------------------------------------
# Blatt "Aufträge"
# ----------------------------------------------------------------------------
def build_orders(wb):
    ws = wb.create_sheet("Aufträge")

    ws.cell(row=A_TITLE, column=1, value="Aufträge – zentrale Datenerfassung"
            ).font = f(16, True, C_HEAD)
    ws.cell(row=2, column=1,
            value="SmartInfra / IoT-Büro  ·  DusL Plus Autostrom").font = f(10)
    ws.cell(row=3, column=1,
            value="Jeder Auftrag bzw. finanzielle Vorgang wird hier einmalig erfasst. "
                  "Der Reiter „Cashflow“ aggregiert daraus automatisch nach Monat."
            ).font = f(9, italic=True, color="808080")
    ws.cell(row=4, column=1,
            value="Datum Einnahmen und Datum Ausgaben sind unabhängig – ein Auftrag "
                  "kann z. B. im September Ausgaben und erst im November Einnahmen haben. "
                  "LLM = Lastmanagement."
            ).font = f(9, italic=True, color="808080")

    for i, name in enumerate(ORDER_COLS, start=1):
        c = ws.cell(row=A_HEAD, column=i, value=name)
        c.font = f(10, True, "FFFFFF")
        c.fill = FILL_HEAD
        c.border = BORDER
        c.alignment = Alignment(horizontal="left" if i <= 4 else "center",
                                vertical="center", wrap_text=i > 4)
    ws.row_dimensions[A_HEAD].height = 30

    fmt = {5: DATE_FMT, 6: DATE_FMT, 7: EUR, 8: EUR, 9: PCT}
    for r in range(A_FIRST, A_LAST + 1):
        for col in range(1, len(ORDER_COLS) + 1):
            c = ws.cell(row=r, column=col)
            c.fill = FILL_INPUT
            c.border = BORDER
            c.font = f(10, color="0000FF")      # blau = manuelle Eingabe
            if col in fmt:
                c.number_format = fmt[col]
            if col == A_COL_AUFTRAG:
                c.alignment = Alignment(horizontal="left", indent=1)
            elif col <= 4:
                c.alignment = Alignment(horizontal="left")

    tab = Table(displayName=TBL, ref=f"A{A_HEAD}:I{A_LAST}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight8", showRowStripes=False,
                                        showColumnStripes=False,
                                        showFirstColumn=False, showLastColumn=False)
    ws.add_table(tab)

    # Dropdowns / Wertebereiche – bewusst etwas über die Tabelle hinaus, damit
    # sie auch beim Erweitern der Tabelle nach unten sofort greifen.
    extra = A_LAST + 200
    dv_tf = DataValidation(type="list", formula1=f'"{",".join(THEMENFELDER)}"',
                           allow_blank=True, showDropDown=False)
    dv_tf.error = "Bitte IoT, EDM, Parking oder LLM auswählen."
    dv_tf.errorTitle = "Ungültiges Themenfeld"
    dv_kat = DataValidation(type="list", formula1=f'"{",".join(KATEGORIEN)}"',
                            allow_blank=True, showDropDown=False)
    dv_kat.error = ("Bitte Hardware, Hardware + Dienstleistung, Lizenzen "
                    "oder Dienstleistung auswählen.")
    dv_kat.errorTitle = "Ungültige Kategorie"
    dv_pct = DataValidation(type="decimal", operator="between",
                            formula1="0", formula2="1", allow_blank=True)
    dv_pct.errorTitle = "Ungültige Wahrscheinlichkeit"
    dv_pct.error = ("Zulässig sind 0 % bis 100 %. Eingabe z. B. als 20% "
                    "oder als 0,2.")
    dv_dat = DataValidation(type="date", operator="between",
                            formula1="DATE(2020,1,1)", formula2="DATE(2040,12,31)",
                            allow_blank=True)
    dv_dat.errorTitle = "Ungültiges Datum"
    dv_dat.error = "Bitte ein Datum im Format TT.MM.JJJJ eintragen."

    for dv, rng in ((dv_tf, f"B{A_FIRST}:B{extra}"),
                    (dv_kat, f"C{A_FIRST}:C{extra}"),
                    (dv_dat, f"E{A_FIRST}:F{extra}"),
                    (dv_pct, f"I{A_FIRST}:I{extra}")):
        ws.add_data_validation(dv)
        dv.add(rng)

    for col, width in zip("ABCDEFGHI",
                          (28, 14, 26, 20, 17, 17, 16, 16, 22)):
        ws.column_dimensions[col].width = width

    ws.freeze_panes = f"B{A_FIRST}"          # Kopfzeile + Spalte "Auftrag"
    ws.sheet_view.showGridLines = False
    ws.print_area = f"A1:I{A_LAST}"
    ws.print_title_rows = f"{A_HEAD}:{A_HEAD}"
    ws.print_title_cols = "$A:$A"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.oddFooter.center.text = "DusL Plus Autostrom – Aufträge  |  Seite &P von &N"
    ws.oddFooter.center.size = 8
    return ws


# ----------------------------------------------------------------------------
# Blatt "Cashflow"
# ----------------------------------------------------------------------------
def build_cashflow(wb, meta, kostenstellen):
    if len(kostenstellen) > N_KST_ROWS:
        raise SystemExit("Mehr Kostenstellen als Zeilen – N_KST_ROWS erhöhen.")

    ws = wb.active
    ws.title = "Cashflow"

    for i, text in enumerate(meta[:3]):
        c = ws.cell(row=R_TITLE + i, column=COL_POS, value=text)
        c.font = f(16, True, C_HEAD) if i == 0 else f(10)
        c.alignment = Alignment(vertical="center")
    ws.cell(row=4, column=COL_POS, value=META_HINT).font = f(
        9, italic=True, color="808080")

    # ---- technische Monatsgrenzen (ausgeblendet) ----------------------------
    for i, start in enumerate(MONTH_STARTS):
        c = ws.cell(row=R_MONTHKEY, column=COL_M1 + i, value=start)
        c.number_format = DATE_FMT
        c.font = f(8, color="808080")
    ws.row_dimensions[R_MONTHKEY].hidden = True

    # ---- Tabellenkopf -------------------------------------------------------
    headers = ["Auftrag", "Kostenstelle", "Themenfeld", "Kategorie"] + MONTHS + [
        *[label for label, _, _ in PERIODS], "Gesamt"]
    for i, text in enumerate(headers, start=1):
        c = ws.cell(row=R_HEAD, column=i, value=text)
        c.font = f(10, True, "FFFFFF")
        c.fill = FILL_HEAD
        c.border = BORDER
        c.alignment = Alignment(horizontal="left" if i in COL_LABELS else "center",
                                vertical="center", wrap_text=i not in COL_LABELS)
    ws.row_dimensions[R_HEAD].height = 30

    def section(row, title):
        for col in range(COL_POS, COL_TOTAL + 1):
            c = ws.cell(row=row, column=col)
            c.fill = FILL_SECTION
            c.border = BORDER
        c = ws.cell(row=row, column=COL_POS, value=title)
        c.font = f(11, True, C_HEAD)
        ws.row_dimensions[row].height = 20

    def lookup(row, table_col):
        """Holt eine Zusatzangabe zur Auftragsnummer aus dem Reiter „Aufträge“."""
        return (f'=IF(${L_KEY}{row}="","",IFERROR(INDEX({table_col},'
                f'MATCH(${L_KEY}{row},{C_AUFTRAG},0)),""))')

    def detail_row(row, value_col, date_col):
        """Auftragszeile: nur Spalte A ist Eingabe, alles andere ist Formel."""
        k = ws.cell(row=row, column=COL_POS)
        k.font = f(10, color="0000FF")
        k.fill = FILL_INPUT
        k.border = BORDER
        k.alignment = Alignment(indent=1)

        for col, table_col in ((COL_KST, C_KST), (COL_TFD, C_TF), (COL_KTG, C_KAT)):
            c = ws.cell(row=row, column=col, value=lookup(row, table_col))
            c.font = f(10, color="595959")
            c.fill = FILL_CALC
            c.border = BORDER

        for col in range(COL_M1, COL_LAST_M + 1):
            formula = sumifs(value_col, date_col, col,
                             extra=[C_AUFTRAG, f"${L_KEY}{row}"])
            # ohne Auftragsnummer bleibt die Zeile bei 0
            c = ws.cell(row=row, column=col,
                        value=f'=IF(${L_KEY}{row}="",0,{formula[1:]})')
            c.fill = FILL_CALC
            c.border = BORDER
            c.number_format = EUR
            c.font = f(10)

        for col in SUM_COLS:
            c = ws.cell(row=row, column=col, value=period_sum(col, row))
            c.fill = FILL_CALC
            c.border = BORDER
            c.number_format = EUR
            c.font = f(10, bold=(col == COL_TOTAL))

    def key_row(row, value_col, date_col, field):
        """Auswertungszeile mit frei eingetragenem Schlüssel (z. B. Kostenstelle)."""
        k = ws.cell(row=row, column=COL_POS)
        k.font = f(10, color="0000FF")
        k.fill = FILL_INPUT
        k.border = BORDER
        k.alignment = Alignment(indent=1)
        for col in (COL_KST, COL_TFD, COL_KTG):
            c = ws.cell(row=row, column=col)
            c.fill = FILL_CALC
            c.border = BORDER
        for col in range(COL_M1, COL_LAST_M + 1):
            formula = sumifs(value_col, date_col, col,
                             extra=[field, f"${L_KEY}{row}"])
            c = ws.cell(row=row, column=col,
                        value=f'=IF(${L_KEY}{row}="",0,{formula[1:]})')
            c.fill = FILL_CALC
            c.border = BORDER
            c.number_format = EUR
            c.font = f(10)
        for col in SUM_COLS:
            c = ws.cell(row=row, column=col, value=period_sum(col, row))
            c.fill = FILL_CALC
            c.border = BORDER
            c.number_format = EUR
            c.font = f(10, bold=(col == COL_TOTAL))

    def calc_row(row, label, formula_for_col, fill, bold=False, border=BORDER,
                 note=None):
        p = ws.cell(row=row, column=COL_POS, value=label)
        p.font = f(10, bold, C_HEAD)
        p.fill = fill
        p.border = border
        n = ws.cell(row=row, column=COL_KST, value=note)
        n.font = f(9, color="808080")
        n.fill = fill
        n.border = border
        for col in (COL_TFD, COL_KTG):
            c = ws.cell(row=row, column=col)
            c.fill = fill
            c.border = border
        for col in range(COL_M1, COL_TOTAL + 1):
            c = ws.cell(row=row, column=col, value=formula_for_col(col))
            c.font = f(10, bold)
            c.fill = fill
            c.border = border
            c.number_format = EUR

    # ---- 1. Einnahmen -------------------------------------------------------
    section(R_SEC_IN, "1  EINNAHMEN JE AUFTRAG  (Quelle: Reiter „Aufträge“)")
    for i in range(N_INCOME_ROWS):
        detail_row(R_IN_FIRST + i, C_EINNAHMEN, C_DAT_EIN)
    calc_row(R_IN_REST, "Noch nicht erfasste Aufträge",
             lambda col: (f"={get_column_letter(col)}{R_SUM_IN}"
                          f"-SUM({get_column_letter(col)}{R_IN_FIRST}:"
                          f"{get_column_letter(col)}{R_IN_LAST})"),
             FILL_CALC, note="Differenz")
    calc_row(R_SUM_IN, "Summe Einnahmen",
             lambda col: (sumifs(C_EINNAHMEN, C_DAT_EIN, col)
                          if col <= COL_LAST_M else period_sum(col, R_SUM_IN)),
             FILL_RESULT, bold=True, border=TOP_LINE)

    # ---- 2. Ausgaben --------------------------------------------------------
    section(R_SEC_OUT, "2  AUSGABEN JE AUFTRAG  (Quelle: Reiter „Aufträge“)")
    for i in range(N_EXPENSE_ROWS):
        detail_row(R_OUT_FIRST + i, C_AUSGABEN, C_DAT_AUS)
    calc_row(R_OUT_REST, "Noch nicht erfasste Aufträge",
             lambda col: (f"={get_column_letter(col)}{R_SUM_OUT}"
                          f"-SUM({get_column_letter(col)}{R_OUT_FIRST}:"
                          f"{get_column_letter(col)}{R_OUT_LAST})"),
             FILL_CALC, note="Differenz")
    calc_row(R_SUM_OUT, "Summe Ausgaben",
             lambda col: (sumifs(C_AUSGABEN, C_DAT_AUS, col)
                          if col <= COL_LAST_M else period_sum(col, R_SUM_OUT)),
             FILL_RESULT, bold=True, border=TOP_LINE)

    # ---- 3. Kennzahlen ------------------------------------------------------
    section(R_SEC_KPI, "3  KENNZAHLEN (automatisch berechnet)")
    calc_row(R_KPI_IN, "Summe Einnahmen",
             lambda col: f"={get_column_letter(col)}{R_SUM_IN}", FILL_CALC)
    calc_row(R_KPI_OUT, "Summe Ausgaben",
             lambda col: f"={get_column_letter(col)}{R_SUM_OUT}", FILL_CALC)
    calc_row(R_KPI_NET, "Netto-Cashflow (Monat)",
             lambda col: (f"={get_column_letter(col)}{R_KPI_IN}"
                          f"-{get_column_letter(col)}{R_KPI_OUT}"),
             FILL_RESULT, bold=True)

    def cum_formula(col):
        """Kumuliert über die Monate; in den Summenspalten der Periodenendstand."""
        if col == COL_M1:
            return f"={L_M1}{R_KPI_NET}"
        if col <= COL_LAST_M:
            return (f"={get_column_letter(col - 1)}{R_KPI_CUM}"
                    f"+{get_column_letter(col)}{R_KPI_NET}")
        return period_end_ref(col, R_KPI_CUM)      # Periodenendstand

    calc_row(R_KPI_CUM, "Kumulierter Cashflow", cum_formula, FILL_RESULT, bold=True)

    # ---- 4./5. Auswertung nach Themenfeld und Kategorie (nachrichtlich) -----
    def dimension_block(sec_row, first_row, title, field, values, weighted_net=False):
        section(sec_row, title)
        r = first_row
        for value in values:
            for label, vcol, dcol in (("Einnahmen", C_EINNAHMEN, C_DAT_EIN),
                                      ("Ausgaben", C_AUSGABEN, C_DAT_AUS)):
                calc_row(r, f"{value} – {label}",
                         lambda col, v=value, vc=vcol, dc=dcol, rr=r: (
                             sumifs(vc, dc, col, extra=[field, f'"{v}"'])
                             if col <= COL_LAST_M else period_sum(col, rr)),
                         FILL_CALC)
                r += 1
            if weighted_net:
                # gewichtete Einnahmen minus gewichtete Ausgaben des Themenfelds
                calc_row(r, f"{value} – gewichteter Netto-Forecast",
                         lambda col, v=value, rr=r: (
                             (weighted(C_EINNAHMEN, C_DAT_EIN, col,
                                       extra=[f"({field}=\"{v}\")*1"])
                              + "-" +
                              weighted(C_AUSGABEN, C_DAT_AUS, col,
                                       extra=[f"({field}=\"{v}\")*1"])[1:])
                             if col <= COL_LAST_M else period_sum(col, rr)),
                         FILL_FORECAST)
                r += 1

    dimension_block(R_SEC_TF, R_TF_FIRST,
                    "4  AUSWERTUNG NACH THEMENFELD (nachrichtlich, nicht in der "
                    "Cashflow-Summe enthalten)", C_TF, THEMENFELDER,
                    weighted_net=True)
    dimension_block(R_SEC_KAT, R_KAT_FIRST,
                    "5  AUSWERTUNG NACH KATEGORIE (nachrichtlich, nicht in der "
                    "Cashflow-Summe enthalten)", C_KAT, KATEGORIEN)

    # ---- 6. Forecast nach Auftragswahrscheinlichkeit ------------------------
    section(R_SEC_FC, "6  FORECAST NACH AUFTRAGSWAHRSCHEINLICHKEIT")

    def subsection(row, title):
        for col in range(COL_POS, COL_TOTAL + 1):
            c = ws.cell(row=row, column=col)
            c.fill = FILL_FORECAST
            c.border = BORDER
        c = ws.cell(row=row, column=COL_POS, value=title)
        c.font = f(10, True, C_HEAD)

    # Management-Box: Kennzahlen über den gesamten Auftragsbestand
    subsection(R_FCBOX, "Pipeline Forecast  (Zusammenfassung über alle Aufträge)")

    def box_ref(i):
        return f"{get_column_letter(COL_KTG)}{R_FCBOX_FIRST + i}"

    box = [
        ("Pipeline Einnahmen gesamt", f"=SUM({C_EINNAHMEN})", EUR, False),
        ("Pipeline Ausgaben gesamt", f"=SUM({C_AUSGABEN})", EUR, False),
        ("Pipeline Netto-Cashflow", f"={box_ref(0)}-{box_ref(1)}", EUR, True),
        ("Gewichtete Einnahmen", f"=SUMPRODUCT({C_EINNAHMEN},{C_WSK})", EUR, False),
        ("Gewichtete Ausgaben", f"=SUMPRODUCT({C_AUSGABEN},{C_WSK})", EUR, False),
        ("Gewichteter Netto-Cashflow", f"={box_ref(3)}-{box_ref(4)}", EUR, True),
        ("Sichere Einnahmen (100 %)", f"=SUMIFS({C_EINNAHMEN},{C_WSK},1)", EUR, False),
        ("Sichere Ausgaben (100 %)", f"=SUMIFS({C_AUSGABEN},{C_WSK},1)", EUR, False),
        ("Sicherer Netto-Cashflow", f"={box_ref(6)}-{box_ref(7)}", EUR, True),
        ("Durchschnittliche Auftragswahrscheinlichkeit",
         f"=IFERROR(AVERAGE({C_WSK}),0)", PCT, False),
        ("Anzahl offene Aufträge (> 0 % und < 100 %)",
         f'=COUNTIFS({C_WSK},">0",{C_WSK},"<1")', COUNT_FMT, False),
        ("Anzahl sichere Aufträge (100 %)", f"=COUNTIF({C_WSK},1)", COUNT_FMT, False),
        ("Pipeline-Wert (Einnahmen gesamt)", f"=SUM({C_EINNAHMEN})", EUR, False),
        ("Gewichteter Pipeline-Wert", f"=SUMPRODUCT({C_EINNAHMEN},{C_WSK})", EUR, False),
        ("Conversion Gap (gefährdeter Umsatz)", f"={box_ref(12)}-{box_ref(13)}",
         EUR, True),
    ]
    assert len(box) == N_FCBOX_ROWS
    for i, (label, formula, fmt, bold) in enumerate(box):
        row = R_FCBOX_FIRST + i
        fill = FILL_FORECAST if bold else FILL_CALC
        for col in (COL_POS, COL_KST, COL_TFD):
            c = ws.cell(row=row, column=col)
            c.fill = fill
            c.border = BORDER
        lab = ws.cell(row=row, column=COL_POS, value=label)
        lab.font = f(10, bold, C_HEAD if bold else "000000")
        lab.alignment = Alignment(indent=1)
        val = ws.cell(row=row, column=COL_KTG, value=formula)
        val.font = f(10, bold)
        val.fill = fill
        val.border = BORDER
        val.number_format = fmt
        val.alignment = Alignment(horizontal="right")

    # Gewichteter Forecast je Monat
    subsection(R_FC_SUB1, "Gewichteter Forecast je Monat  "
                          "(Betrag × Auftragswahrscheinlichkeit)")
    calc_row(R_FC_FIRST, "Gewichtete Einnahmen",
             lambda col: (weighted(C_EINNAHMEN, C_DAT_EIN, col)
                          if col <= COL_LAST_M else period_sum(col, R_FC_FIRST)),
             FILL_CALC)
    calc_row(R_FC_FIRST + 1, "Gewichtete Ausgaben",
             lambda col: (weighted(C_AUSGABEN, C_DAT_AUS, col)
                          if col <= COL_LAST_M else period_sum(col, R_FC_FIRST + 1)),
             FILL_CALC)
    calc_row(R_FC_NET, "Gewichteter Netto-Cashflow",
             lambda col: (f"={get_column_letter(col)}{R_FC_FIRST}"
                          f"-{get_column_letter(col)}{R_FC_FIRST + 1}"),
             FILL_FORECAST, bold=True)

    def fc_cum(col):
        if col == COL_M1:
            return f"={L_M1}{R_FC_NET}"
        if col <= COL_LAST_M:
            return (f"={get_column_letter(col - 1)}{R_FC_CUM}"
                    f"+{get_column_letter(col)}{R_FC_NET}")
        return period_end_ref(col, R_FC_CUM)       # Periodenendstand

    calc_row(R_FC_CUM, "Kumulierter gewichteter Cashflow", fc_cum,
             FILL_FORECAST, bold=True)

    # Sichere Aufträge (100 %)
    subsection(R_FC_SUB2, "Cashflow nur 100%-Aufträge  (bereits gesicherte Aufträge)")
    calc_row(R_SAFE_FIRST, "Sichere Einnahmen (100 %)",
             lambda col: (sumifs(C_EINNAHMEN, C_DAT_EIN, col, extra=[C_WSK, "1"])
                          if col <= COL_LAST_M else period_sum(col, R_SAFE_FIRST)),
             FILL_CALC)
    calc_row(R_SAFE_FIRST + 1, "Sichere Ausgaben (100 %)",
             lambda col: (sumifs(C_AUSGABEN, C_DAT_AUS, col, extra=[C_WSK, "1"])
                          if col <= COL_LAST_M else period_sum(col, R_SAFE_FIRST + 1)),
             FILL_CALC)
    calc_row(R_SAFE_NET, "Sicherer Netto-Cashflow",
             lambda col: (f"={get_column_letter(col)}{R_SAFE_FIRST}"
                          f"-{get_column_letter(col)}{R_SAFE_FIRST + 1}"),
             FILL_RESULT, bold=True)

    def kst_block(sec_row, first_row, title, value_col, date_col):
        section(sec_row, title)
        for i in range(N_KST_ROWS):
            row = first_row + i
            key_row(row, value_col, date_col, C_KST)
            if i < len(kostenstellen):
                nummer, name = kostenstellen[i]
                ws.cell(row=row, column=COL_POS).value = nummer
                b = ws.cell(row=row, column=COL_KST, value=name)
                b.font = f(10, color="595959")
                b.fill = FILL_INPUT

    kst_block(R_SEC_KST_IN, R_KST_IN_FIRST,
              "7  EINNAHMEN NACH KOSTENSTELLE (nachrichtlich, nicht in der "
              "Cashflow-Summe enthalten)", C_EINNAHMEN, C_DAT_EIN)
    kst_block(R_SEC_KST_OUT, R_KST_OUT_FIRST,
              "8  AUSGABEN NACH KOSTENSTELLE (nachrichtlich, nicht in der "
              "Cashflow-Summe enthalten)", C_AUSGABEN, C_DAT_AUS)

    # ---- Legende ------------------------------------------------------------
    legend = [
        ("Legende", None, f(10, True, C_HEAD)),
        ("Eingabezelle – hier wird die Auftragsnummer bzw. die Kostenstelle "
         "eingetragen", C_INPUT, f(9)),
        ("Automatisch aus dem Reiter „Aufträge“ berechnet – nicht überschreiben",
         C_CALC, f(9)),
        ("Ergebniszeile – automatisch berechnet", C_RESULT, f(9)),
        ("Datenfluss:  Aufträge  →  Cashflow  →  Übersicht. Alle Beträge werden "
         "ausschließlich im Reiter „Aufträge“ erfasst.",
         None, f(9, italic=True, color="808080")),
        ("Spalte A: Auftragsnummer aus dem Reiter „Aufträge“ eintragen – "
         "Kostenstelle, Themenfeld und Kategorie werden automatisch daneben "
         "ergänzt, die Monatswerte über das jeweilige Datum zugeordnet.",
         None, f(9, italic=True, color="808080")),
        (f"Zeile {R_IN_REST} bzw. {R_OUT_REST} zeigt Beträge aus Aufträgen, die "
         "oben noch keine eigene Zeile haben. Die Summenzeilen enthalten immer "
         "den vollständigen Monatswert, es geht also nie ein Betrag verloren.",
         None, f(9, italic=True, color="808080")),
        (f"Zeile {R_MONTHKEY} ist ausgeblendet und enthält die echten Monatsanfänge "
         "(Datumswerte), mit denen die SUMMEWENNS-Formeln die Monatsgrenzen bilden.",
         None, f(9, italic=True, color="808080")),
        ("Hellblau hinterlegt – gewichteter Forecast (Betrag × Auftrags"
         "wahrscheinlichkeit)", C_FORECAST, f(9)),
        (f"Bereich 6 ist eine reine Analyse und wird nirgends zum Cashflow "
         f"addiert: Zeilen {R_FC_FIRST}–{R_FC_CUM} zeigen den gewichteten "
         f"Forecast, Zeilen {R_SAFE_FIRST}–{R_SAFE_NET} nur die Aufträge mit "
         "100 % Wahrscheinlichkeit. Die Kennzahlen in Bereich 3 bleiben "
         "unverändert der vollständige Cashflow.",
         None, f(9, italic=True, color="808080")),
        ("Eine leere Auftragswahrscheinlichkeit wirkt im Forecast wie 0 % – der "
         "normale Cashflow zeigt den Betrag trotzdem vollständig.",
         None, f(9, italic=True, color="808080")),
    ]
    for i, (text, fill, font) in enumerate(legend):
        r = R_LEGEND + i
        if fill:
            marker = ws.cell(row=r, column=COL_POS)
            marker.fill = PatternFill("solid", fgColor=fill)
            marker.border = BORDER
            c = ws.cell(row=r, column=COL_KST, value=text)
        else:
            c = ws.cell(row=r, column=COL_POS, value=text)
        c.font = font

    # ---- Bedingte Formatierung ---------------------------------------------
    ws.conditional_formatting.add(
        f"{L_M1}{R_IN_FIRST}:{L_TOTAL}{R_MEMO_LAST}",
        CellIsRule(operator="lessThan", formula=["0"],
                   font=Font(name=FONT, size=10, color="9C0006"),
                   fill=PatternFill("solid", bgColor="FFC7CE")))
    result_ranges = [f"{L_M1}{R_KPI_NET}:{L_TOTAL}{R_KPI_CUM}",
                     f"{L_M1}{R_FC_NET}:{L_TOTAL}{R_FC_CUM}",
                     f"{L_M1}{R_SAFE_NET}:{L_TOTAL}{R_SAFE_NET}",
                     f"{get_column_letter(COL_KTG)}{R_FCBOX_FIRST}:"
                     f"{get_column_letter(COL_KTG)}{R_FCBOX_FIRST + 8}"]
    result_range = " ".join(result_ranges)
    ws.conditional_formatting.add(
        result_range, CellIsRule(operator="greaterThan", formula=["0"],
                                 font=Font(name=FONT, size=10, bold=True,
                                           color="006100"),
                                 fill=PatternFill("solid", bgColor="C6EFCE")))
    ws.conditional_formatting.add(
        result_range, CellIsRule(operator="lessThan", formula=["0"],
                                 font=Font(name=FONT, size=10, bold=True,
                                           color="9C0006"),
                                 fill=PatternFill("solid", bgColor="FFC7CE")))

    # ---- Spaltenbreiten / Fixierung / Druck ---------------------------------
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 24
    for col in range(COL_M1, COL_LAST_M + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13
    for col in SUM_COLS:
        ws.column_dimensions[get_column_letter(col)].width = 16

    ws.freeze_panes = f"{L_M1}{R_HEAD + 1}"
    ws.sheet_view.showGridLines = False
    ws.print_area = f"A1:{L_TOTAL}{R_LEGEND + len(legend) - 1}"
    ws.print_title_rows = f"{R_HEAD}:{R_HEAD}"
    ws.print_title_cols = "$A:$D"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 3      # 54 Monate: 3 Seiten breit, sonst unlesbar
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5
    ws.oddFooter.center.text = "DusL Plus Autostrom – Cashflow  |  Seite &P von &N"
    ws.oddFooter.center.size = 8
    return ws


# ----------------------------------------------------------------------------
# Blatt "Übersicht"
# ----------------------------------------------------------------------------
def build_overview(wb, ws_cf):
    ov = wb.create_sheet("Übersicht")

    ov["A1"] = "Übersicht Cashflow – DusL Plus Autostrom"
    ov["A1"].font = f(16, True, C_HEAD)
    ov["A2"] = f"SmartInfra / IoT-Büro  ·  {PERIOD_LABEL}"
    ov["A2"].font = f(10, color="595959")
    ov["A3"] = "Alle Werte werden automatisch aus dem Blatt „Cashflow“ übernommen."
    ov["A3"].font = f(9, italic=True, color="808080")

    kpis = []
    for label, first, last in PERIODS:
        name = label.replace("Summe ", "")
        letter = get_column_letter(COL_PERIOD[label])
        von, bis = MONTHS[first], MONTHS[last]
        kpis.append((f"{name}  ({von} – {bis})", None, None))
        kpis.append((f"Einnahmen {name}", f"=Cashflow!{letter}{R_KPI_IN}", False))
        kpis.append((f"Ausgaben {name}", f"=Cashflow!{letter}{R_KPI_OUT}", False))
        kpis.append((f"Netto-Cashflow {name}", f"=Cashflow!{letter}{R_KPI_NET}", True))
    kpis += [
        (f"Gesamtzeitraum ({N_MONTHS} Monate)", None, None),
        ("Gesamteinnahmen", f"=Cashflow!{L_TOTAL}{R_KPI_IN}", False),
        ("Gesamtausgaben", f"=Cashflow!{L_TOTAL}{R_KPI_OUT}", False),
        ("Gesamt-Netto-Cashflow", f"=Cashflow!{L_TOTAL}{R_KPI_NET}", True),
    ]

    start = 5
    ov.cell(row=start, column=1, value="Kennzahl").font = f(10, True, "FFFFFF")
    ov.cell(row=start, column=2, value="Betrag").font = f(10, True, "FFFFFF")
    for col in (1, 2):
        c = ov.cell(row=start, column=col)
        c.fill = FILL_HEAD
        c.border = BORDER
        c.alignment = Alignment(horizontal="left" if col == 1 else "center")

    r = start + 1
    for label, formula, bold in kpis:
        if formula is None:
            c = ov.cell(row=r, column=1, value=label)
            c.font = f(10, True, C_HEAD)
            c.fill = FILL_SECTION
            c.border = BORDER
            c2 = ov.cell(row=r, column=2)
            c2.fill = FILL_SECTION
            c2.border = BORDER
        else:
            c = ov.cell(row=r, column=1, value=label)
            c.font = f(10, bold)
            c.border = BORDER
            c.alignment = Alignment(indent=1)
            v = ov.cell(row=r, column=2, value=formula)
            v.font = f(10, bold)
            v.fill = FILL_RESULT if bold else FILL_CALC
            v.border = BORDER
            v.number_format = EUR
        r += 1
    last_kpi_row = r - 1

    ov.conditional_formatting.add(
        f"B{start + 1}:B{last_kpi_row}",
        CellIsRule(operator="lessThan", formula=["0"],
                   font=Font(name=FONT, size=10, bold=True, color="9C0006"),
                   fill=PatternFill("solid", bgColor="FFC7CE")))
    ov.conditional_formatting.add(
        f"B{start + 1}:B{last_kpi_row}",
        CellIsRule(operator="greaterThan", formula=["0"],
                   font=Font(name=FONT, size=10, bold=True, color="006100"),
                   fill=PatternFill("solid", bgColor="C6EFCE")))

    ov.column_dimensions["A"].width = 34
    ov.column_dimensions["B"].width = 20
    ov.sheet_view.showGridLines = False
    ov.page_setup.orientation = "portrait"
    ov.page_setup.paperSize = ov.PAPERSIZE_A4
    ov.page_setup.fitToWidth = 1
    ov.page_setup.fitToHeight = 0
    ov.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    cats = Reference(ws_cf, min_col=COL_M1, max_col=COL_LAST_M,
                     min_row=R_HEAD, max_row=R_HEAD)

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.title = "Einnahmen vs. Ausgaben (monatlich)"
    bar.y_axis.title = "EUR"
    bar.y_axis.numFmt = '#,##0'
    bar.height, bar.width = 8, 26
    bar.gapWidth = 40
    # Werte erst ab der ersten Monatsspalte, sonst kippen die drei
    # Beschriftungsspalten als Datenpunkte in die Reihe und verschieben sie
    # gegenüber den Monatskategorien.
    bar.add_data(Reference(ws_cf, min_col=COL_M1, max_col=COL_LAST_M,
                           min_row=R_KPI_IN, max_row=R_KPI_OUT), from_rows=True)
    bar.set_categories(cats)
    for ser, row in zip(bar.series, (R_KPI_IN, R_KPI_OUT)):
        ser.tx = SeriesLabel(strRef=StrRef(f"Cashflow!$A${row}"))
    bar.series[0].graphicalProperties.solidFill = "4472C4"
    bar.series[1].graphicalProperties.solidFill = "C00000"
    ov.add_chart(bar, "D5")

    line = LineChart()
    line.title = "Netto-Cashflow und kumulierter Cashflow (monatlich)"
    line.y_axis.title = "EUR"
    line.y_axis.numFmt = '#,##0'
    line.height, line.width = 8, 26
    line.add_data(Reference(ws_cf, min_col=COL_M1, max_col=COL_LAST_M,
                            min_row=R_KPI_NET, max_row=R_KPI_CUM), from_rows=True)
    line.set_categories(cats)
    for ser, row in zip(line.series, (R_KPI_NET, R_KPI_CUM)):
        ser.tx = SeriesLabel(strRef=StrRef(f"Cashflow!$A${row}"))
    line.series[0].graphicalProperties.line.solidFill = "1F3864"
    line.series[1].graphicalProperties.line.solidFill = "70AD47"
    line.series[1].graphicalProperties.line.dashStyle = "dash"
    ov.add_chart(line, "D22")
    return ov


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", help="bestehende Datei, deren Positionen "
                                     "übernommen werden")
    ap.add_argument("--out", default=str(OUT))
    args = ap.parse_args()

    if args.source:
        src = read_source(args.source)
        meta, kostenstellen = src["meta"], src["kostenstellen"]
        print(f"übernommen: {len(kostenstellen)} Kostenstellen "
              f"({', '.join(n for n, _ in kostenstellen)})")
    else:
        meta, kostenstellen = META_DEFAULT, []

    wb = Workbook()
    ws_cf = build_cashflow(wb, meta, kostenstellen)
    build_orders(wb)                       # direkt nach "Cashflow"
    build_overview(wb, ws_cf)
    wb.save(args.out)
    print(f"geschrieben: {args.out}  |  Reiter: {wb.sheetnames}")


if __name__ == "__main__":
    main()
